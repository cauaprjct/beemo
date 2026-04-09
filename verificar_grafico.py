"""Script para verificar se o gráfico foi criado corretamente no Excel."""

from openpyxl import load_workbook
from pathlib import Path

def verificar_grafico():
    """Verifica se o gráfico foi criado no arquivo vendas_2025.xlsx."""
    
    file_path = r"C:\Users\ngb\Documents\GeminiOfficeFiles\vendas_2025.xlsx"
    
    print("=" * 80)
    print("VERIFICAÇÃO DE GRÁFICO NO EXCEL")
    print("=" * 80)
    print(f"\nArquivo: {file_path}")
    
    if not Path(file_path).exists():
        print("❌ ERRO: Arquivo não encontrado!")
        return
    
    try:
        # Carregar o workbook
        wb = load_workbook(file_path)
        
        # Verificar sheets disponíveis
        print(f"\n📊 Sheets disponíveis: {wb.sheetnames}")
        
        # Pegar a sheet principal
        sheet = wb['Vendas']
        
        print(f"\n📈 Verificando gráficos na sheet 'Vendas'...")
        
        # Verificar se há gráficos
        charts = sheet._charts
        
        if not charts:
            print("❌ NENHUM GRÁFICO ENCONTRADO!")
            print("\nPossíveis causas:")
            print("  1. O comando não foi executado com sucesso")
            print("  2. Houve erro durante a criação do gráfico")
            print("  3. O gráfico foi criado em outra sheet")
        else:
            print(f"✅ GRÁFICO(S) ENCONTRADO(S): {len(charts)} gráfico(s)")
            
            for i, chart in enumerate(charts, 1):
                print(f"\n--- Gráfico {i} ---")
                print(f"  Tipo: {type(chart).__name__}")
                
                # Extrair título de forma mais limpa
                if hasattr(chart, 'title') and chart.title:
                    try:
                        title_text = chart.title.tx.rich.p[0].r[0].t if chart.title.tx and chart.title.tx.rich else "Sem título"
                    except:
                        title_text = "Título presente (não extraível)"
                    print(f"  Título: {title_text}")
                else:
                    print(f"  Título: Sem título")
                
                # Extrair posição de forma mais clara
                if hasattr(chart, 'anchor') and hasattr(chart.anchor, '_from'):
                    col = chart.anchor._from.col
                    row = chart.anchor._from.row
                    # Converter coluna numérica para letra (0=A, 1=B, 7=H, etc)
                    col_letter = chr(65 + col)  # 65 é 'A' em ASCII
                    position = f"{col_letter}{row + 1}"  # +1 porque Excel usa 1-based
                    print(f"  Posição: {position}")
                else:
                    print(f"  Posição: Não identificada")
                
                print(f"  Largura: {chart.width} cm")
                print(f"  Altura: {chart.height} cm")
                
                # Verificar se é BarChart (column ou bar)
                if hasattr(chart, 'type'):
                    print(f"  Subtipo: {chart.type}")
                
                # Verificar títulos dos eixos (se aplicável)
                if hasattr(chart, 'x_axis') and hasattr(chart.x_axis, 'title'):
                    x_title = chart.x_axis.title if chart.x_axis.title else "Sem título"
                    print(f"  Título eixo X: {x_title}")
                
                if hasattr(chart, 'y_axis') and hasattr(chart.y_axis, 'title'):
                    y_title = chart.y_axis.title if chart.y_axis.title else "Sem título"
                    print(f"  Título eixo Y: {y_title}")
        
        # Verificar dados na planilha
        print(f"\n📋 Verificando dados da planilha...")
        print(f"\nPrimeiras 5 linhas de dados:")
        print(f"{'ID':<5} {'Data':<12} {'Produto':<12} {'Qtd':<6} {'Preço':<10} {'Total':<10}")
        print("-" * 65)
        
        for row in range(2, 7):  # Linhas 2 a 6
            id_val = sheet.cell(row, 1).value
            data_val = sheet.cell(row, 2).value
            produto_val = sheet.cell(row, 3).value
            qtd_val = sheet.cell(row, 4).value
            preco_val = sheet.cell(row, 5).value
            total_val = sheet.cell(row, 6).value
            
            print(f"{str(id_val):<5} {str(data_val):<12} {str(produto_val):<12} "
                  f"{str(qtd_val):<6} {str(preco_val):<10} {str(total_val):<10}")
        
        print("\n" + "=" * 80)
        print("RESUMO DA VERIFICAÇÃO")
        print("=" * 80)
        
        if charts:
            print(f"✅ Status: GRÁFICO CRIADO COM SUCESSO!")
            print(f"✅ Quantidade: {len(charts)} gráfico(s)")
            
            # Extrair posição real
            if hasattr(charts[0], 'anchor') and hasattr(charts[0].anchor, '_from'):
                col = charts[0].anchor._from.col
                row = charts[0].anchor._from.row
                col_letter = chr(65 + col)
                position_real = f"{col_letter}{row + 1}"
            else:
                position_real = "Desconhecida"
            
            print(f"✅ Posição esperada: H2")
            print(f"✅ Posição real: {position_real}")
            
            # Verificar se a posição está correta
            if position_real == "H2":
                print("✅ Posição: CORRETA!")
            else:
                print(f"⚠️  Posição: Esperado H2, encontrado {position_real}")
            
            # Verificar tipo
            chart_type = type(charts[0]).__name__
            if "Bar" in chart_type:
                # Verificar se é column (vertical) ou bar (horizontal)
                if hasattr(charts[0], 'type'):
                    if charts[0].type == "col":
                        print("✅ Tipo: Gráfico de COLUNAS (vertical) - CORRETO!")
                    elif charts[0].type == "bar":
                        print("⚠️  Tipo: Gráfico de BARRAS (horizontal)")
                    else:
                        print(f"✅ Tipo: BarChart ({charts[0].type})")
                else:
                    print("✅ Tipo: Gráfico de colunas/barras (BarChart)")
            else:
                print(f"⚠️  Tipo: {chart_type}")
            
            # Verificar título
            if hasattr(charts[0], 'title') and charts[0].title:
                try:
                    title_text = charts[0].title.tx.rich.p[0].r[0].t if charts[0].title.tx and charts[0].title.tx.rich else None
                    if title_text and "Vendas por Produto" in title_text:
                        print(f"✅ Título: CORRETO ('{title_text}')")
                    elif title_text:
                        print(f"⚠️  Título: '{title_text}' (esperado 'Vendas por Produto')")
                    else:
                        print("⚠️  Título: Presente mas não extraível")
                except:
                    print("⚠️  Título: Presente mas não extraível")
            else:
                print("⚠️  Título: Não encontrado")
        else:
            print("❌ Status: GRÁFICO NÃO ENCONTRADO")
            print("\nVerifique os logs do Streamlit para identificar o erro.")
        
        wb.close()
        
    except Exception as e:
        print(f"\n❌ ERRO ao ler arquivo: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    verificar_grafico()
