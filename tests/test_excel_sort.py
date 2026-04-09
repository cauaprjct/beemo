"""Tests for Excel sort functionality."""

import pytest
from pathlib import Path
from openpyxl import load_workbook
from src.excel_tool import ExcelTool
from src.exceptions import CorruptedFileError


@pytest.fixture
def excel_tool():
    """Create an ExcelTool instance for testing."""
    return ExcelTool()


@pytest.fixture
def sample_data_file(tmp_path):
    """Create a sample Excel file with unsorted data for testing."""
    file_path = tmp_path / "test_sort.xlsx"
    
    tool = ExcelTool()
    data = {
        "Sheet1": [
            ["Nome", "Data", "Valor", "Categoria"],
            ["João", "2025-03-15", 150, "A"],
            ["Ana", "2025-01-10", 300, "B"],
            ["Pedro", "2025-02-20", 100, "A"],
            ["Maria", "2025-03-01", 250, "C"],
            ["Carlos", "2025-01-25", 200, "B"]
        ]
    }
    tool.create_excel(str(file_path), data)
    
    return file_path


class TestSortData:
    """Test suite for sort_data method."""
    
    def test_sort_by_column_letter_ascending(self, excel_tool, sample_data_file):
        """Test sorting by column letter in ascending order."""
        sort_config = {
            'column': 'C',  # Valor column
            'order': 'asc',
            'has_header': True
        }
        
        excel_tool.sort_data(str(sample_data_file), 'Sheet1', sort_config)
        
        # Verify the sort
        workbook = load_workbook(str(sample_data_file))
        sheet = workbook['Sheet1']
        
        # Check that values are sorted (100, 150, 200, 250, 300)
        assert sheet.cell(2, 3).value == 100  # Pedro
        assert sheet.cell(3, 3).value == 150  # João
        assert sheet.cell(4, 3).value == 200  # Carlos
        assert sheet.cell(5, 3).value == 250  # Maria
        assert sheet.cell(6, 3).value == 300  # Ana
        
        workbook.close()
    
    def test_sort_by_column_letter_descending(self, excel_tool, sample_data_file):
        """Test sorting by column letter in descending order."""
        sort_config = {
            'column': 'C',  # Valor column
            'order': 'desc',
            'has_header': True
        }
        
        excel_tool.sort_data(str(sample_data_file), 'Sheet1', sort_config)
        
        # Verify the sort
        workbook = load_workbook(str(sample_data_file))
        sheet = workbook['Sheet1']
        
        # Check that values are sorted (300, 250, 200, 150, 100)
        assert sheet.cell(2, 3).value == 300  # Ana
        assert sheet.cell(3, 3).value == 250  # Maria
        assert sheet.cell(4, 3).value == 200  # Carlos
        assert sheet.cell(5, 3).value == 150  # João
        assert sheet.cell(6, 3).value == 100  # Pedro
        
        workbook.close()
    
    def test_sort_by_column_number(self, excel_tool, sample_data_file):
        """Test sorting by column number."""
        sort_config = {
            'column': 1,  # Nome column (first column)
            'order': 'asc',
            'has_header': True
        }
        
        excel_tool.sort_data(str(sample_data_file), 'Sheet1', sort_config)
        
        # Verify the sort
        workbook = load_workbook(str(sample_data_file))
        sheet = workbook['Sheet1']
        
        # Check that names are sorted alphabetically
        assert sheet.cell(2, 1).value == "Ana"
        assert sheet.cell(3, 1).value == "Carlos"
        assert sheet.cell(4, 1).value == "João"
        assert sheet.cell(5, 1).value == "Maria"
        assert sheet.cell(6, 1).value == "Pedro"
        
        workbook.close()
    
    def test_sort_with_header(self, excel_tool, sample_data_file):
        """Test that header row is not included in sort."""
        sort_config = {
            'column': 'A',
            'order': 'asc',
            'has_header': True
        }
        
        excel_tool.sort_data(str(sample_data_file), 'Sheet1', sort_config)
        
        # Verify header is still in row 1
        workbook = load_workbook(str(sample_data_file))
        sheet = workbook['Sheet1']
        
        assert sheet.cell(1, 1).value == "Nome"
        assert sheet.cell(1, 2).value == "Data"
        assert sheet.cell(1, 3).value == "Valor"
        assert sheet.cell(1, 4).value == "Categoria"
        
        workbook.close()
    
    def test_sort_without_header(self, excel_tool, tmp_path):
        """Test sorting data without header row."""
        file_path = tmp_path / "no_header.xlsx"
        
        data = {
            "Sheet1": [
                [3, "C"],
                [1, "A"],
                [2, "B"]
            ]
        }
        excel_tool.create_excel(str(file_path), data)
        
        sort_config = {
            'column': 1,
            'order': 'asc',
            'has_header': False
        }
        
        excel_tool.sort_data(str(file_path), 'Sheet1', sort_config)
        
        # Verify the sort
        workbook = load_workbook(str(file_path))
        sheet = workbook['Sheet1']
        
        assert sheet.cell(1, 1).value == 1
        assert sheet.cell(2, 1).value == 2
        assert sheet.cell(3, 1).value == 3
        
        workbook.close()
    
    def test_sort_specific_range(self, excel_tool, sample_data_file):
        """Test sorting a specific range of rows."""
        sort_config = {
            'column': 'C',
            'start_row': 2,
            'end_row': 4,  # Only sort first 3 data rows
            'order': 'asc',
            'has_header': True
        }
        
        excel_tool.sort_data(str(sample_data_file), 'Sheet1', sort_config)
        
        # Verify only rows 2-4 are sorted
        workbook = load_workbook(str(sample_data_file))
        sheet = workbook['Sheet1']
        
        # First 3 rows should be sorted (100, 150, 300)
        assert sheet.cell(2, 3).value == 100  # Pedro
        assert sheet.cell(3, 3).value == 150  # João
        assert sheet.cell(4, 3).value == 300  # Ana
        
        # Last 2 rows should remain unchanged
        assert sheet.cell(5, 3).value == 250  # Maria
        assert sheet.cell(6, 3).value == 200  # Carlos
        
        workbook.close()
    
    def test_sort_by_date_column(self, excel_tool, sample_data_file):
        """Test sorting by date column."""
        sort_config = {
            'column': 'B',  # Data column
            'order': 'asc',
            'has_header': True
        }
        
        excel_tool.sort_data(str(sample_data_file), 'Sheet1', sort_config)
        
        # Verify dates are sorted
        workbook = load_workbook(str(sample_data_file))
        sheet = workbook['Sheet1']
        
        # Check chronological order
        assert sheet.cell(2, 2).value == "2025-01-10"  # Ana
        assert sheet.cell(3, 2).value == "2025-01-25"  # Carlos
        assert sheet.cell(4, 2).value == "2025-02-20"  # Pedro
        assert sheet.cell(5, 2).value == "2025-03-01"  # Maria
        assert sheet.cell(6, 2).value == "2025-03-15"  # João
        
        workbook.close()
    
    def test_sort_by_text_column(self, excel_tool, sample_data_file):
        """Test sorting by text column."""
        sort_config = {
            'column': 'D',  # Categoria column
            'order': 'asc',
            'has_header': True
        }
        
        excel_tool.sort_data(str(sample_data_file), 'Sheet1', sort_config)
        
        # Verify categories are sorted
        workbook = load_workbook(str(sample_data_file))
        sheet = workbook['Sheet1']
        
        # Check alphabetical order (A, A, B, B, C)
        assert sheet.cell(2, 4).value == "A"
        assert sheet.cell(3, 4).value == "A"
        assert sheet.cell(4, 4).value == "B"
        assert sheet.cell(5, 4).value == "B"
        assert sheet.cell(6, 4).value == "C"
        
        workbook.close()
    
    def test_sort_invalid_column_letter(self, excel_tool, sample_data_file):
        """Test error handling for invalid column letter."""
        sort_config = {
            'column': 'XYZ',
            'order': 'asc'
        }
        
        with pytest.raises(ValueError, match="Invalid column letter"):
            excel_tool.sort_data(str(sample_data_file), 'Sheet1', sort_config)
    
    def test_sort_invalid_column_type(self, excel_tool, sample_data_file):
        """Test error handling for invalid column type."""
        sort_config = {
            'column': 3.14,  # Float not allowed
            'order': 'asc'
        }
        
        with pytest.raises(ValueError, match="Column must be string"):
            excel_tool.sort_data(str(sample_data_file), 'Sheet1', sort_config)
    
    def test_sort_invalid_order(self, excel_tool, sample_data_file):
        """Test error handling for invalid order."""
        sort_config = {
            'column': 'A',
            'order': 'invalid'
        }
        
        with pytest.raises(ValueError, match="Order must be 'asc' or 'desc'"):
            excel_tool.sort_data(str(sample_data_file), 'Sheet1', sort_config)
    
    def test_sort_invalid_sheet(self, excel_tool, sample_data_file):
        """Test error handling for non-existent sheet."""
        sort_config = {
            'column': 'A',
            'order': 'asc'
        }
        
        with pytest.raises(ValueError, match="Sheet 'NonExistent' not found"):
            excel_tool.sort_data(str(sample_data_file), 'NonExistent', sort_config)
    
    def test_sort_missing_column(self, excel_tool, sample_data_file):
        """Test error handling when column is missing from config."""
        sort_config = {
            'order': 'asc'
        }
        
        with pytest.raises(ValueError, match="'column' is required"):
            excel_tool.sort_data(str(sample_data_file), 'Sheet1', sort_config)
    
    def test_sort_file_not_found(self, excel_tool):
        """Test error handling for non-existent file."""
        sort_config = {
            'column': 'A',
            'order': 'asc'
        }
        
        with pytest.raises(FileNotFoundError):
            excel_tool.sort_data('nonexistent.xlsx', 'Sheet1', sort_config)
    
    def test_sort_invalid_row_range(self, excel_tool, sample_data_file):
        """Test error handling for invalid row range."""
        sort_config = {
            'column': 'A',
            'start_row': 10,
            'end_row': 5,  # end_row < start_row
            'order': 'asc'
        }
        
        with pytest.raises(ValueError, match="end_row .* must be >= start_row"):
            excel_tool.sort_data(str(sample_data_file), 'Sheet1', sort_config)
    
    def test_sort_preserves_row_integrity(self, excel_tool, sample_data_file):
        """Test that sorting preserves entire rows (all columns move together)."""
        sort_config = {
            'column': 'A',  # Sort by Nome
            'order': 'asc',
            'has_header': True
        }
        
        excel_tool.sort_data(str(sample_data_file), 'Sheet1', sort_config)
        
        # Verify that when Ana is in row 2, her data is also in row 2
        workbook = load_workbook(str(sample_data_file))
        sheet = workbook['Sheet1']
        
        # Ana's row
        assert sheet.cell(2, 1).value == "Ana"
        assert sheet.cell(2, 2).value == "2025-01-10"
        assert sheet.cell(2, 3).value == 300
        assert sheet.cell(2, 4).value == "B"
        
        # Carlos's row
        assert sheet.cell(3, 1).value == "Carlos"
        assert sheet.cell(3, 2).value == "2025-01-25"
        assert sheet.cell(3, 3).value == 200
        assert sheet.cell(3, 4).value == "B"
        
        workbook.close()
