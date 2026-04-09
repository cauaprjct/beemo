"""Tests for freeze_panes and unfreeze_panes functionality in ExcelTool."""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from src.excel_tool import ExcelTool
from src.exceptions import CorruptedFileError


@pytest.fixture
def excel_tool():
    """Create ExcelTool instance."""
    return ExcelTool()


@pytest.fixture
def sample_file(tmp_path):
    """Create a sample Excel file with data."""
    file_path = tmp_path / "test.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add header row and data (10 rows, 10 columns)
    headers = [f"Col{i}" for i in range(1, 11)]
    ws.append(headers)
    
    for row_num in range(2, 11):
        row_data = [f"R{row_num}C{col}" for col in range(1, 11)]
        ws.append(row_data)
    
    wb.save(file_path)
    wb.close()
    
    return str(file_path)


class TestFreezePanes:
    """Test freezing panes."""
    
    def test_freeze_header_row(self, excel_tool, sample_file):
        """Test freezing the header row (row 1)."""
        excel_tool.freeze_panes(sample_file, 'Sheet1', row=2)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Check freeze is set at B2 (freezes row 1)
        assert ws.freeze_panes == 'A2'
        
        wb.close()
    
    def test_freeze_first_column(self, excel_tool, sample_file):
        """Test freezing the first column."""
        excel_tool.freeze_panes(sample_file, 'Sheet1', col='B')
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Check freeze is set at B1 (freezes column A)
        assert ws.freeze_panes == 'B1'
        
        wb.close()
    
    def test_freeze_first_column_by_number(self, excel_tool, sample_file):
        """Test freezing first column using number."""
        excel_tool.freeze_panes(sample_file, 'Sheet1', col=2)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Check freeze is set at B1 (freezes column A)
        assert ws.freeze_panes == 'B1'
        
        wb.close()
    
    def test_freeze_both_row_and_column(self, excel_tool, sample_file):
        """Test freezing both header row and first column."""
        excel_tool.freeze_panes(sample_file, 'Sheet1', row=2, col='B')
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Check freeze is set at B2 (freezes row 1 and column A)
        assert ws.freeze_panes == 'B2'
        
        wb.close()
    
    def test_freeze_multiple_rows(self, excel_tool, sample_file):
        """Test freezing multiple rows."""
        excel_tool.freeze_panes(sample_file, 'Sheet1', row=4)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Check freeze is set at A4 (freezes rows 1-3)
        assert ws.freeze_panes == 'A4'
        
        wb.close()
    
    def test_freeze_multiple_columns(self, excel_tool, sample_file):
        """Test freezing multiple columns."""
        excel_tool.freeze_panes(sample_file, 'Sheet1', col='D')
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Check freeze is set at D1 (freezes columns A-C)
        assert ws.freeze_panes == 'D1'
        
        wb.close()
    
    def test_freeze_replaces_existing_freeze(self, excel_tool, sample_file):
        """Test that new freeze replaces existing freeze."""
        # First freeze
        excel_tool.freeze_panes(sample_file, 'Sheet1', row=2)
        
        # Second freeze (should replace first)
        excel_tool.freeze_panes(sample_file, 'Sheet1', col='C')
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Should only have the second freeze
        assert ws.freeze_panes == 'C1'
        
        wb.close()


class TestUnfreezePanes:
    """Test unfreezing panes."""
    
    def test_unfreeze_panes(self, excel_tool, sample_file):
        """Test removing freeze panes."""
        # First freeze
        excel_tool.freeze_panes(sample_file, 'Sheet1', row=2, col='B')
        
        # Verify freeze is set
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.freeze_panes == 'B2'
        wb.close()
        
        # Unfreeze
        excel_tool.unfreeze_panes(sample_file, 'Sheet1')
        
        # Verify freeze is removed
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.freeze_panes is None
        
        wb.close()
    
    def test_unfreeze_when_no_freeze(self, excel_tool, sample_file):
        """Test unfreezing when there's no freeze (should not error)."""
        # Should not raise error
        excel_tool.unfreeze_panes(sample_file, 'Sheet1')
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.freeze_panes is None
        wb.close()


class TestFreezePanesValidation:
    """Test validation and error handling."""
    
    def test_freeze_no_parameters(self, excel_tool, sample_file):
        """Test error when neither row nor col is provided."""
        with pytest.raises(ValueError, match="At least one of 'row' or 'col' must be specified"):
            excel_tool.freeze_panes(sample_file, 'Sheet1')
    
    def test_freeze_invalid_row(self, excel_tool, sample_file):
        """Test error for invalid row."""
        with pytest.raises(ValueError, match="row must be >= 1"):
            excel_tool.freeze_panes(sample_file, 'Sheet1', row=0)
    
    def test_freeze_invalid_col_number(self, excel_tool, sample_file):
        """Test error for invalid column number."""
        with pytest.raises(ValueError, match="col must be >= 1"):
            excel_tool.freeze_panes(sample_file, 'Sheet1', col=0)
    
    def test_freeze_invalid_col_type(self, excel_tool, sample_file):
        """Test error for invalid column type."""
        with pytest.raises(ValueError, match="Invalid column type"):
            excel_tool.freeze_panes(sample_file, 'Sheet1', col=3.14)
    
    def test_freeze_invalid_sheet(self, excel_tool, sample_file):
        """Test error for invalid sheet name."""
        with pytest.raises(ValueError, match="Sheet .* not found"):
            excel_tool.freeze_panes(sample_file, 'NonExistent', row=2)
    
    def test_freeze_file_not_found(self, excel_tool):
        """Test error for non-existent file."""
        with pytest.raises(FileNotFoundError):
            excel_tool.freeze_panes('nonexistent.xlsx', 'Sheet1', row=2)
    
    def test_unfreeze_invalid_sheet(self, excel_tool, sample_file):
        """Test error for invalid sheet name when unfreezing."""
        with pytest.raises(ValueError, match="Sheet .* not found"):
            excel_tool.unfreeze_panes(sample_file, 'NonExistent')
    
    def test_unfreeze_file_not_found(self, excel_tool):
        """Test error for non-existent file when unfreezing."""
        with pytest.raises(FileNotFoundError):
            excel_tool.unfreeze_panes('nonexistent.xlsx', 'Sheet1')


class TestFreezePanesIntegration:
    """Test integration scenarios."""
    
    def test_freeze_after_data_operations(self, excel_tool, sample_file):
        """Test freezing after modifying data."""
        # Modify data
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        ws.cell(1, 1, "Modified Header")
        wb.save(sample_file)
        wb.close()
        
        # Freeze
        excel_tool.freeze_panes(sample_file, 'Sheet1', row=2)
        
        # Verify both modification and freeze
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.cell(1, 1).value == "Modified Header"
        assert ws.freeze_panes == 'A2'
        wb.close()
    
    def test_freeze_unfreeze_freeze_cycle(self, excel_tool, sample_file):
        """Test freeze -> unfreeze -> freeze cycle."""
        # First freeze
        excel_tool.freeze_panes(sample_file, 'Sheet1', row=2)
        
        # Unfreeze
        excel_tool.unfreeze_panes(sample_file, 'Sheet1')
        
        # Freeze again with different parameters
        excel_tool.freeze_panes(sample_file, 'Sheet1', row=3, col='C')
        
        # Verify final state
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.freeze_panes == 'C3'
        wb.close()
    
    def test_freeze_on_multiple_sheets(self, excel_tool, sample_file):
        """Test freezing on multiple sheets."""
        # Add second sheet
        wb = load_workbook(sample_file)
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Header1", "Header2", "Header3"])
        ws2.append(["Data1", "Data2", "Data3"])
        wb.save(sample_file)
        wb.close()
        
        # Freeze on both sheets
        excel_tool.freeze_panes(sample_file, 'Sheet1', row=2)
        excel_tool.freeze_panes(sample_file, 'Sheet2', row=2, col='B')
        
        # Verify both freezes
        wb = load_workbook(sample_file)
        assert wb['Sheet1'].freeze_panes == 'A2'
        assert wb['Sheet2'].freeze_panes == 'B2'
        wb.close()
    
    def test_freeze_with_large_dataset(self, excel_tool, tmp_path):
        """Test freezing with a large dataset."""
        file_path = tmp_path / "large.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Create large dataset (100 rows, 50 columns)
        headers = [f"Column{i}" for i in range(1, 51)]
        ws.append(headers)
        
        for row_num in range(2, 102):
            row_data = [f"R{row_num}C{col}" for col in range(1, 51)]
            ws.append(row_data)
        
        wb.save(file_path)
        wb.close()
        
        # Freeze header and first 2 columns
        excel_tool.freeze_panes(str(file_path), 'Data', row=2, col='C')
        
        # Verify
        wb = load_workbook(file_path)
        ws = wb['Data']
        assert ws.freeze_panes == 'C2'
        wb.close()
