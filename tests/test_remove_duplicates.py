"""Tests for remove_duplicates functionality in ExcelTool."""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from src.excel_tool import ExcelTool
from src.exceptions import CorruptedFileError


@pytest.fixture
def excel_tool():
    """Create ExcelTool instance."""
    return ExcelTool()


@pytest.fixture
def sample_file_with_duplicates(tmp_path):
    """Create a sample Excel file with duplicate data."""
    file_path = tmp_path / "duplicates.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add header
    ws.append(["ID", "Name", "Email", "City"])
    
    # Add data with duplicates
    ws.append([1, "Alice", "alice@example.com", "NYC"])
    ws.append([2, "Bob", "bob@example.com", "LA"])
    ws.append([3, "Charlie", "charlie@example.com", "Chicago"])
    ws.append([1, "Alice Duplicate", "alice@example.com", "Boston"])  # Duplicate email
    ws.append([4, "David", "david@example.com", "Seattle"])
    ws.append([2, "Bob Again", "bob@example.com", "Miami"])  # Duplicate email
    ws.append([5, "Eve", "eve@example.com", "Portland"])
    
    wb.save(file_path)
    wb.close()
    
    return str(file_path)


@pytest.fixture
def sample_file_with_duplicate_rows(tmp_path):
    """Create a sample Excel file with completely duplicate rows."""
    file_path = tmp_path / "duplicate_rows.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add header
    ws.append(["Product", "Quantity", "Price"])
    
    # Add data with duplicate rows
    ws.append(["Apple", 10, 1.50])
    ws.append(["Banana", 20, 0.75])
    ws.append(["Apple", 10, 1.50])  # Exact duplicate
    ws.append(["Orange", 15, 1.25])
    ws.append(["Banana", 20, 0.75])  # Exact duplicate
    ws.append(["Grape", 30, 2.00])
    
    wb.save(file_path)
    wb.close()
    
    return str(file_path)


class TestRemoveDuplicatesByColumn:
    """Test removing duplicates by specific column."""
    
    def test_remove_duplicates_by_email_keep_first(self, excel_tool, sample_file_with_duplicates):
        """Test removing duplicates by email column, keeping first occurrence."""
        config = {
            'column': 'C',  # Email column
            'has_header': True,
            'keep': 'first'
        }
        
        result = excel_tool.remove_duplicates(sample_file_with_duplicates, 'Sheet1', config)
        
        # Should remove 2 duplicates (alice and bob duplicates)
        assert result['removed_count'] == 2
        assert result['remaining_count'] == 6  # 1 header + 5 unique rows
        
        # Verify data
        wb = load_workbook(sample_file_with_duplicates)
        ws = wb['Sheet1']
        
        emails = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)]
        assert emails == ["alice@example.com", "bob@example.com", "charlie@example.com", 
                         "david@example.com", "eve@example.com"]
        
        # Verify first occurrences were kept
        names = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)]
        assert "Alice" in names
        assert "Bob" in names
        assert "Alice Duplicate" not in names
        assert "Bob Again" not in names
        
        wb.close()
    
    def test_remove_duplicates_by_email_keep_last(self, excel_tool, sample_file_with_duplicates):
        """Test removing duplicates by email column, keeping last occurrence."""
        config = {
            'column': 'C',
            'has_header': True,
            'keep': 'last'
        }
        
        result = excel_tool.remove_duplicates(sample_file_with_duplicates, 'Sheet1', config)
        
        assert result['removed_count'] == 2
        assert result['remaining_count'] == 6
        
        # Verify last occurrences were kept
        wb = load_workbook(sample_file_with_duplicates)
        ws = wb['Sheet1']
        
        names = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)]
        assert "Alice Duplicate" in names
        assert "Bob Again" in names
        assert "Alice" not in names
        assert "Bob" not in names
        
        wb.close()
    
    def test_remove_duplicates_by_id_column_number(self, excel_tool, sample_file_with_duplicates):
        """Test removing duplicates by ID using column number."""
        config = {
            'column': 1,  # ID column
            'has_header': True,
            'keep': 'first'
        }
        
        result = excel_tool.remove_duplicates(sample_file_with_duplicates, 'Sheet1', config)
        
        # Should remove 2 duplicates (ID 1 and 2 appear twice)
        assert result['removed_count'] == 2
        assert result['remaining_count'] == 6
        
        # Verify IDs are unique
        wb = load_workbook(sample_file_with_duplicates)
        ws = wb['Sheet1']
        
        ids = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
        assert ids == [1, 2, 3, 4, 5]
        
        wb.close()


class TestRemoveDuplicateRows:
    """Test removing completely duplicate rows."""
    
    def test_remove_duplicate_rows_keep_first(self, excel_tool, sample_file_with_duplicate_rows):
        """Test removing completely duplicate rows, keeping first occurrence."""
        config = {
            'has_header': True,
            'keep': 'first'
        }
        
        result = excel_tool.remove_duplicates(sample_file_with_duplicate_rows, 'Sheet1', config)
        
        # Should remove 2 duplicate rows
        assert result['removed_count'] == 2
        assert result['remaining_count'] == 5  # 1 header + 4 unique rows
        
        # Verify data
        wb = load_workbook(sample_file_with_duplicate_rows)
        ws = wb['Sheet1']
        
        products = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
        assert products == ["Apple", "Banana", "Orange", "Grape"]
        
        wb.close()
    
    def test_remove_duplicate_rows_keep_last(self, excel_tool, sample_file_with_duplicate_rows):
        """Test removing completely duplicate rows, keeping last occurrence."""
        config = {
            'has_header': True,
            'keep': 'last'
        }
        
        result = excel_tool.remove_duplicates(sample_file_with_duplicate_rows, 'Sheet1', config)
        
        assert result['removed_count'] == 2
        assert result['remaining_count'] == 5
        
        wb = load_workbook(sample_file_with_duplicate_rows)
        ws = wb['Sheet1']
        
        # Should still have 4 unique products
        products = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
        assert len(products) == 4
        assert set(products) == {"Apple", "Banana", "Orange", "Grape"}
        
        wb.close()


class TestRemoveDuplicatesNoHeader:
    """Test removing duplicates without header row."""
    
    def test_remove_duplicates_no_header(self, excel_tool, tmp_path):
        """Test removing duplicates when there's no header row."""
        file_path = tmp_path / "no_header.xlsx"
        
        wb = Workbook()
        ws = wb.active
        
        # Add data without header
        ws.append([1, "Alice"])
        ws.append([2, "Bob"])
        ws.append([1, "Alice"])  # Duplicate
        ws.append([3, "Charlie"])
        
        wb.save(file_path)
        wb.close()
        
        config = {
            'has_header': False,
            'keep': 'first'
        }
        
        result = excel_tool.remove_duplicates(str(file_path), 'Sheet', config)
        
        assert result['removed_count'] == 1
        assert result['remaining_count'] == 3


class TestRemoveDuplicatesEdgeCases:
    """Test edge cases for remove_duplicates."""
    
    def test_no_duplicates_found(self, excel_tool, tmp_path):
        """Test when there are no duplicates."""
        file_path = tmp_path / "no_duplicates.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email"])
        ws.append(["Alice", "alice@example.com"])
        ws.append(["Bob", "bob@example.com"])
        ws.append(["Charlie", "charlie@example.com"])
        
        wb.save(file_path)
        wb.close()
        
        config = {'column': 'B', 'has_header': True, 'keep': 'first'}
        result = excel_tool.remove_duplicates(str(file_path), 'Sheet', config)
        
        assert result['removed_count'] == 0
        assert result['remaining_count'] == 4  # Header + 3 rows
    
    def test_empty_sheet(self, excel_tool, tmp_path):
        """Test removing duplicates from empty sheet."""
        file_path = tmp_path / "empty.xlsx"
        
        wb = Workbook()
        ws = wb.active
        
        wb.save(file_path)
        wb.close()
        
        config = {'has_header': True, 'keep': 'first'}
        result = excel_tool.remove_duplicates(str(file_path), 'Sheet', config)
        
        assert result['removed_count'] == 0
        assert result['remaining_count'] == 0
    
    def test_only_header_row(self, excel_tool, tmp_path):
        """Test when sheet has only header row."""
        file_path = tmp_path / "only_header.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email"])
        
        wb.save(file_path)
        wb.close()
        
        config = {'column': 'B', 'has_header': True, 'keep': 'first'}
        result = excel_tool.remove_duplicates(str(file_path), 'Sheet', config)
        
        assert result['removed_count'] == 0
        assert result['remaining_count'] == 1
    
    def test_all_rows_duplicate(self, excel_tool, tmp_path):
        """Test when all rows are duplicates."""
        file_path = tmp_path / "all_duplicates.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.append(["Name"])
        ws.append(["Alice"])
        ws.append(["Alice"])
        ws.append(["Alice"])
        ws.append(["Alice"])
        
        wb.save(file_path)
        wb.close()
        
        config = {'column': 'A', 'has_header': True, 'keep': 'first'}
        result = excel_tool.remove_duplicates(str(file_path), 'Sheet', config)
        
        assert result['removed_count'] == 3
        assert result['remaining_count'] == 2  # Header + 1 row


class TestRemoveDuplicatesValidation:
    """Test validation and error handling."""
    
    def test_invalid_keep_value(self, excel_tool, sample_file_with_duplicates):
        """Test error for invalid keep value."""
        config = {'column': 'C', 'has_header': True, 'keep': 'middle'}
        
        with pytest.raises(ValueError, match="Invalid 'keep' value"):
            excel_tool.remove_duplicates(sample_file_with_duplicates, 'Sheet1', config)
    
    def test_invalid_column_type(self, excel_tool, sample_file_with_duplicates):
        """Test error for invalid column type."""
        config = {'column': 3.14, 'has_header': True, 'keep': 'first'}
        
        with pytest.raises(ValueError, match="Invalid column type"):
            excel_tool.remove_duplicates(sample_file_with_duplicates, 'Sheet1', config)
    
    def test_column_out_of_range(self, excel_tool, sample_file_with_duplicates):
        """Test error for column out of range."""
        config = {'column': 99, 'has_header': True, 'keep': 'first'}
        
        with pytest.raises(ValueError, match="Column .* out of range"):
            excel_tool.remove_duplicates(sample_file_with_duplicates, 'Sheet1', config)
    
    def test_invalid_sheet_name(self, excel_tool, sample_file_with_duplicates):
        """Test error for invalid sheet name."""
        config = {'column': 'C', 'has_header': True, 'keep': 'first'}
        
        with pytest.raises(ValueError, match="Sheet .* not found"):
            excel_tool.remove_duplicates(sample_file_with_duplicates, 'NonExistent', config)
    
    def test_file_not_found(self, excel_tool):
        """Test error for non-existent file."""
        config = {'column': 'C', 'has_header': True, 'keep': 'first'}
        
        with pytest.raises(FileNotFoundError):
            excel_tool.remove_duplicates('nonexistent.xlsx', 'Sheet1', config)


class TestRemoveDuplicatesIntegration:
    """Test integration scenarios."""
    
    def test_remove_duplicates_then_verify(self, excel_tool, sample_file_with_duplicates):
        """Test removing duplicates and verifying the result."""
        # Remove duplicates
        config = {'column': 'C', 'has_header': True, 'keep': 'first'}
        result = excel_tool.remove_duplicates(sample_file_with_duplicates, 'Sheet1', config)
        
        assert result['removed_count'] == 2
        
        # Verify by reading the file
        wb = load_workbook(sample_file_with_duplicates)
        ws = wb['Sheet1']
        
        # Check that all emails are now unique
        emails = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)]
        assert len(emails) == len(set(emails))  # All unique
        
        wb.close()
    
    def test_remove_duplicates_preserves_other_data(self, excel_tool, sample_file_with_duplicates):
        """Test that removing duplicates preserves other column data correctly."""
        config = {'column': 'C', 'has_header': True, 'keep': 'first'}
        excel_tool.remove_duplicates(sample_file_with_duplicates, 'Sheet1', config)
        
        wb = load_workbook(sample_file_with_duplicates)
        ws = wb['Sheet1']
        
        # Verify that each row still has all 4 columns
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, 5):
                assert ws.cell(row=row_idx, column=col_idx).value is not None
        
        wb.close()
