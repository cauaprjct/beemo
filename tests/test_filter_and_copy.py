"""Tests for filter_and_copy functionality in ExcelTool."""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from src.excel_tool import ExcelTool
from src.exceptions import CorruptedFileError


@pytest.fixture
def excel_tool():
    """Create ExcelTool instance."""
    return ExcelTool()


@pytest.fixture
def sample_sales_file(tmp_path):
    """Create a sample Excel file with sales data."""
    file_path = tmp_path / "sales.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add header
    ws.append(["ID", "Product", "Category", "Quantity", "Price", "Total"])
    
    # Add data
    ws.append([1, "Laptop", "Electronics", 2, 1000, 2000])
    ws.append([2, "Mouse", "Electronics", 10, 25, 250])
    ws.append([3, "Desk", "Furniture", 1, 500, 500])
    ws.append([4, "Chair", "Furniture", 4, 150, 600])
    ws.append([5, "Monitor", "Electronics", 3, 300, 900])
    ws.append([6, "Keyboard", "Electronics", 5, 80, 400])
    ws.append([7, "Table", "Furniture", 1, 800, 800])
    ws.append([8, "Headphones", "Electronics", 8, 50, 400])
    
    wb.save(file_path)
    wb.close()
    
    return str(file_path)


@pytest.fixture
def sample_customers_file(tmp_path):
    """Create a sample Excel file with customer data."""
    file_path = tmp_path / "customers.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add header
    ws.append(["ID", "Name", "Email", "City"])
    
    # Add data
    ws.append([1, "Alice", "alice@gmail.com", "NYC"])
    ws.append([2, "Bob", "bob@yahoo.com", "LA"])
    ws.append([3, "Charlie", "charlie@gmail.com", "Chicago"])
    ws.append([4, "David", "david@outlook.com", "Seattle"])
    ws.append([5, "Eve", "eve@gmail.com", "Portland"])
    
    wb.save(file_path)
    wb.close()
    
    return str(file_path)


class TestFilterNumeric:
    """Test filtering with numeric operators."""
    
    def test_filter_greater_than(self, excel_tool, sample_sales_file):
        """Test filtering values greater than threshold."""
        config = {
            'column': 'F',  # Total column
            'operator': '>',
            'value': 500,
            'destination_sheet': 'High Sales',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
        
        # Should find 4 rows: 2000, 600, 900, 800
        assert result['filtered_count'] == 4
        assert result['destination'] == 'High Sales'
        assert result['destination_type'] == 'sheet'
        
        # Verify data
        wb = load_workbook(sample_sales_file)
        ws = wb['High Sales']
        
        # Check header
        assert ws.cell(1, 1).value == "ID"
        
        # Check filtered data
        totals = [ws.cell(row=i, column=6).value for i in range(2, ws.max_row + 1)]
        assert all(total > 500 for total in totals)
        assert len(totals) == 4
        
        wb.close()
    
    def test_filter_less_than(self, excel_tool, sample_sales_file):
        """Test filtering values less than threshold."""
        config = {
            'column': 'F',
            'operator': '<',
            'value': 500,
            'destination_sheet': 'Low Sales',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
        
        # Should find 3 rows: 250, 400, 400
        assert result['filtered_count'] == 3
        
        wb = load_workbook(sample_sales_file)
        ws = wb['Low Sales']
        
        totals = [ws.cell(row=i, column=6).value for i in range(2, ws.max_row + 1)]
        assert all(total < 500 for total in totals)
        
        wb.close()
    
    def test_filter_greater_equal(self, excel_tool, sample_sales_file):
        """Test filtering values greater than or equal to threshold."""
        config = {
            'column': 'F',
            'operator': '>=',
            'value': 600,
            'destination_sheet': 'Medium Plus',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
        
        # Should find 4 rows: 2000, 600, 900, 800
        assert result['filtered_count'] == 4
    
    def test_filter_less_equal(self, excel_tool, sample_sales_file):
        """Test filtering values less than or equal to threshold."""
        config = {
            'column': 'F',
            'operator': '<=',
            'value': 500,
            'destination_sheet': 'Medium Minus',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
        
        # Should find 4 rows: 250, 500, 400, 400
        assert result['filtered_count'] == 4


class TestFilterEquality:
    """Test filtering with equality operators."""
    
    def test_filter_equals(self, excel_tool, sample_sales_file):
        """Test filtering exact matches."""
        config = {
            'column': 'C',  # Category column
            'operator': '==',
            'value': 'Electronics',
            'destination_sheet': 'Electronics',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
        
        # Should find 5 electronics items
        assert result['filtered_count'] == 5
        
        wb = load_workbook(sample_sales_file)
        ws = wb['Electronics']
        
        categories = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)]
        assert all(cat == 'Electronics' for cat in categories)
        
        wb.close()
    
    def test_filter_not_equals(self, excel_tool, sample_sales_file):
        """Test filtering non-matches."""
        config = {
            'column': 'C',
            'operator': '!=',
            'value': 'Electronics',
            'destination_sheet': 'Non Electronics',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
        
        # Should find 3 furniture items
        assert result['filtered_count'] == 3
        
        wb = load_workbook(sample_sales_file)
        ws = wb['Non Electronics']
        
        categories = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)]
        assert all(cat != 'Electronics' for cat in categories)
        
        wb.close()


class TestFilterText:
    """Test filtering with text operators."""
    
    def test_filter_contains(self, excel_tool, sample_customers_file):
        """Test filtering text that contains substring."""
        config = {
            'column': 'C',  # Email column
            'operator': 'contains',
            'value': '@gmail.com',
            'destination_sheet': 'Gmail Users',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_customers_file, 'Sheet1', config)
        
        # Should find 3 gmail users
        assert result['filtered_count'] == 3
        
        wb = load_workbook(sample_customers_file)
        ws = wb['Gmail Users']
        
        emails = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)]
        assert all('@gmail.com' in email.lower() for email in emails)
        
        wb.close()
    
    def test_filter_starts_with(self, excel_tool, sample_customers_file):
        """Test filtering text that starts with prefix."""
        config = {
            'column': 'B',  # Name column
            'operator': 'starts_with',
            'value': 'A',
            'destination_sheet': 'Names A',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_customers_file, 'Sheet1', config)
        
        # Should find 1 name starting with A (Alice)
        assert result['filtered_count'] == 1
        
        wb = load_workbook(sample_customers_file)
        ws = wb['Names A']
        
        name = ws.cell(2, 2).value
        assert name.startswith('A')
        
        wb.close()
    
    def test_filter_ends_with(self, excel_tool, sample_customers_file):
        """Test filtering text that ends with suffix."""
        config = {
            'column': 'C',
            'operator': 'ends_with',
            'value': '.com',
            'destination_sheet': 'Dot Com',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_customers_file, 'Sheet1', config)
        
        # All emails end with .com
        assert result['filtered_count'] == 5


class TestFilterToFile:
    """Test filtering to new file."""
    
    def test_filter_to_new_file(self, excel_tool, sample_sales_file, tmp_path):
        """Test filtering data to new file."""
        dest_file = tmp_path / "high_sales.xlsx"
        
        config = {
            'column': 'F',
            'operator': '>',
            'value': 500,
            'destination_file': str(dest_file),
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
        
        assert result['filtered_count'] == 4
        assert result['destination'] == str(dest_file)
        assert result['destination_type'] == 'file'
        
        # Verify new file exists and has correct data
        assert dest_file.exists()
        
        wb = load_workbook(dest_file)
        ws = wb.active
        
        # Check header
        assert ws.cell(1, 1).value == "ID"
        
        # Check data
        totals = [ws.cell(row=i, column=6).value for i in range(2, ws.max_row + 1)]
        assert all(total > 500 for total in totals)
        assert len(totals) == 4
        
        wb.close()
    
    def test_filter_without_header_copy(self, excel_tool, sample_sales_file):
        """Test filtering without copying header."""
        config = {
            'column': 'F',
            'operator': '>',
            'value': 500,
            'destination_sheet': 'No Header',
            'has_header': True,
            'copy_header': False
        }
        
        result = excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
        
        assert result['filtered_count'] == 4
        
        wb = load_workbook(sample_sales_file)
        ws = wb['No Header']
        
        # First row should be data, not header
        assert ws.cell(1, 1).value != "ID"
        assert isinstance(ws.cell(1, 1).value, int)
        
        wb.close()


class TestFilterEdgeCases:
    """Test edge cases for filter_and_copy."""
    
    def test_filter_no_matches(self, excel_tool, sample_sales_file):
        """Test filtering when no rows match."""
        config = {
            'column': 'F',
            'operator': '>',
            'value': 10000,
            'destination_sheet': 'No Matches',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
        
        assert result['filtered_count'] == 0
        
        wb = load_workbook(sample_sales_file)
        ws = wb['No Matches']
        
        # Should only have header
        assert ws.max_row == 1
        
        wb.close()
    
    def test_filter_all_match(self, excel_tool, sample_sales_file):
        """Test filtering when all rows match."""
        config = {
            'column': 'F',
            'operator': '>',
            'value': 0,
            'destination_sheet': 'All Match',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
        
        # All 8 data rows should match
        assert result['filtered_count'] == 8
    
    def test_filter_empty_sheet(self, excel_tool, tmp_path):
        """Test filtering empty sheet."""
        file_path = tmp_path / "empty.xlsx"
        
        wb = Workbook()
        ws = wb.active
        
        wb.save(file_path)
        wb.close()
        
        config = {
            'column': 'A',
            'operator': '>',
            'value': 0,
            'destination_sheet': 'Result',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(str(file_path), 'Sheet', config)
        
        assert result['filtered_count'] == 0
    
    def test_filter_overwrites_existing_sheet(self, excel_tool, sample_sales_file):
        """Test that filtering overwrites existing destination sheet."""
        # First filter
        config = {
            'column': 'F',
            'operator': '>',
            'value': 500,
            'destination_sheet': 'Filtered',
            'has_header': True
        }
        
        excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
        
        # Second filter with different criteria (should overwrite)
        config['value'] = 1000
        result = excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
        
        # Should only have 1 row (2000)
        assert result['filtered_count'] == 1
        
        wb = load_workbook(sample_sales_file)
        ws = wb['Filtered']
        
        # Should have header + 1 data row
        assert ws.max_row == 2
        
        wb.close()


class TestFilterValidation:
    """Test validation and error handling."""
    
    def test_missing_column(self, excel_tool, sample_sales_file):
        """Test error when column is missing."""
        config = {
            'operator': '>',
            'value': 500,
            'destination_sheet': 'Result'
        }
        
        with pytest.raises(ValueError, match="'column' parameter is required"):
            excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
    
    def test_missing_operator(self, excel_tool, sample_sales_file):
        """Test error when operator is missing."""
        config = {
            'column': 'F',
            'value': 500,
            'destination_sheet': 'Result'
        }
        
        with pytest.raises(ValueError, match="'operator' parameter is required"):
            excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
    
    def test_missing_value(self, excel_tool, sample_sales_file):
        """Test error when value is missing."""
        config = {
            'column': 'F',
            'operator': '>',
            'destination_sheet': 'Result'
        }
        
        with pytest.raises(ValueError, match="'value' parameter is required"):
            excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
    
    def test_invalid_operator(self, excel_tool, sample_sales_file):
        """Test error for invalid operator."""
        config = {
            'column': 'F',
            'operator': 'invalid',
            'value': 500,
            'destination_sheet': 'Result'
        }
        
        with pytest.raises(ValueError, match="Invalid operator"):
            excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
    
    def test_missing_destination(self, excel_tool, sample_sales_file):
        """Test error when no destination specified."""
        config = {
            'column': 'F',
            'operator': '>',
            'value': 500
        }
        
        with pytest.raises(ValueError, match="Either 'destination_sheet' or 'destination_file' must be specified"):
            excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
    
    def test_both_destinations(self, excel_tool, sample_sales_file):
        """Test error when both destinations specified."""
        config = {
            'column': 'F',
            'operator': '>',
            'value': 500,
            'destination_sheet': 'Sheet',
            'destination_file': 'file.xlsx'
        }
        
        with pytest.raises(ValueError, match="Cannot specify both"):
            excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
    
    def test_invalid_column_type(self, excel_tool, sample_sales_file):
        """Test error for invalid column type."""
        config = {
            'column': 3.14,
            'operator': '>',
            'value': 500,
            'destination_sheet': 'Result'
        }
        
        with pytest.raises(ValueError, match="Invalid column type"):
            excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
    
    def test_column_out_of_range(self, excel_tool, sample_sales_file):
        """Test error for column out of range."""
        config = {
            'column': 99,
            'operator': '>',
            'value': 500,
            'destination_sheet': 'Result'
        }
        
        with pytest.raises(ValueError, match="Column .* out of range"):
            excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
    
    def test_invalid_sheet_name(self, excel_tool, sample_sales_file):
        """Test error for invalid sheet name."""
        config = {
            'column': 'F',
            'operator': '>',
            'value': 500,
            'destination_sheet': 'Result'
        }
        
        with pytest.raises(ValueError, match="Sheet .* not found"):
            excel_tool.filter_and_copy(sample_sales_file, 'NonExistent', config)
    
    def test_file_not_found(self, excel_tool):
        """Test error for non-existent file."""
        config = {
            'column': 'F',
            'operator': '>',
            'value': 500,
            'destination_sheet': 'Result'
        }
        
        with pytest.raises(FileNotFoundError):
            excel_tool.filter_and_copy('nonexistent.xlsx', 'Sheet1', config)


class TestFilterIntegration:
    """Test integration scenarios."""
    
    def test_filter_by_column_number(self, excel_tool, sample_sales_file):
        """Test filtering using column number instead of letter."""
        config = {
            'column': 6,  # Total column (F)
            'operator': '>',
            'value': 500,
            'destination_sheet': 'High Sales',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_sales_file, 'Sheet1', config)
        
        assert result['filtered_count'] == 4
    
    def test_filter_case_insensitive(self, excel_tool, sample_customers_file):
        """Test that text filtering is case-insensitive."""
        config = {
            'column': 'C',
            'operator': 'contains',
            'value': '@GMAIL.COM',  # Uppercase
            'destination_sheet': 'Gmail',
            'has_header': True
        }
        
        result = excel_tool.filter_and_copy(sample_customers_file, 'Sheet1', config)
        
        # Should still find gmail users despite case difference
        assert result['filtered_count'] == 3
