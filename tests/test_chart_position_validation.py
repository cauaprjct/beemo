"""Tests for chart position validation functionality."""

import pytest
from pathlib import Path
from src.excel_tool import ExcelTool


@pytest.fixture
def excel_tool():
    """Create an ExcelTool instance for testing."""
    return ExcelTool()


@pytest.fixture
def sample_file_with_chart(tmp_path):
    """Create a sample Excel file with a chart for testing."""
    file_path = tmp_path / "test_chart_validation.xlsx"
    
    tool = ExcelTool()
    
    # Create file with data
    data = {
        "Sheet1": [
            ["Produto", "Vendas"],
            ["A", 100],
            ["B", 200],
            ["C", 150]
        ]
    }
    tool.create_excel(str(file_path), data)
    
    # Add a chart at position H2
    chart_config = {
        'type': 'column',
        'title': 'Vendas por Produto',
        'categories': 'A2:A4',
        'values': 'B2:B4',
        'position': 'H2'
    }
    tool.add_chart(str(file_path), 'Sheet1', chart_config)
    
    return file_path


class TestChartPositionValidation:
    """Test suite for chart position validation."""
    
    def test_add_chart_to_empty_position_succeeds(self, excel_tool, tmp_path):
        """Test that adding chart to empty position succeeds."""
        file_path = tmp_path / "test_empty_position.xlsx"
        
        data = {
            "Sheet1": [
                ["Produto", "Vendas"],
                ["A", 100],
                ["B", 200]
            ]
        }
        excel_tool.create_excel(str(file_path), data)
        
        chart_config = {
            'type': 'column',
            'title': 'Test Chart',
            'categories': 'A2:A3',
            'values': 'B2:B3',
            'position': 'D2'
        }
        
        # Should succeed without error
        excel_tool.add_chart(str(file_path), 'Sheet1', chart_config)
        
        # Verify chart was added
        from openpyxl import load_workbook
        wb = load_workbook(str(file_path))
        ws = wb['Sheet1']
        assert len(ws._charts) == 1
        wb.close()
    
    def test_add_chart_to_occupied_position_fails(self, excel_tool, sample_file_with_chart):
        """Test that adding chart to occupied position raises ValueError."""
        chart_config = {
            'type': 'pie',
            'title': 'Another Chart',
            'categories': 'A2:A4',
            'values': 'B2:B4',
            'position': 'H2'  # Same position as existing chart
        }
        
        with pytest.raises(ValueError) as exc_info:
            excel_tool.add_chart(str(sample_file_with_chart), 'Sheet1', chart_config)
        
        error_msg = str(exc_info.value)
        assert "Position 'H2' is already occupied" in error_msg
        assert "Vendas por Produto" in error_msg  # Chart title should be in error
    
    def test_add_chart_with_replace_existing_succeeds(self, excel_tool, sample_file_with_chart):
        """Test that adding chart with replace_existing=True replaces existing chart."""
        chart_config = {
            'type': 'pie',
            'title': 'New Chart',
            'categories': 'A2:A4',
            'values': 'B2:B4',
            'position': 'H2',
            'replace_existing': True
        }
        
        # Should succeed and replace existing chart
        excel_tool.add_chart(str(sample_file_with_chart), 'Sheet1', chart_config)
        
        # Verify only one chart exists
        from openpyxl import load_workbook
        wb = load_workbook(str(sample_file_with_chart))
        ws = wb['Sheet1']
        assert len(ws._charts) == 1
        
        # Verify it's the new chart (pie chart)
        chart = ws._charts[0]
        assert chart.__class__.__name__ == 'PieChart'
        wb.close()
    
    def test_add_multiple_charts_at_different_positions(self, excel_tool, tmp_path):
        """Test that adding multiple charts at different positions succeeds."""
        file_path = tmp_path / "test_multiple_charts.xlsx"
        
        data = {
            "Sheet1": [
                ["Produto", "Vendas"],
                ["A", 100],
                ["B", 200],
                ["C", 150]
            ]
        }
        excel_tool.create_excel(str(file_path), data)
        
        # Add first chart at H2
        chart_config_1 = {
            'type': 'column',
            'title': 'Chart 1',
            'categories': 'A2:A4',
            'values': 'B2:B4',
            'position': 'H2'
        }
        excel_tool.add_chart(str(file_path), 'Sheet1', chart_config_1)
        
        # Add second chart at K2 (different position)
        chart_config_2 = {
            'type': 'pie',
            'title': 'Chart 2',
            'categories': 'A2:A4',
            'values': 'B2:B4',
            'position': 'K2'
        }
        excel_tool.add_chart(str(file_path), 'Sheet1', chart_config_2)
        
        # Verify both charts exist
        from openpyxl import load_workbook
        wb = load_workbook(str(file_path))
        ws = wb['Sheet1']
        assert len(ws._charts) == 2
        wb.close()
    
    def test_error_message_includes_chart_title(self, excel_tool, sample_file_with_chart):
        """Test that error message includes the title of the existing chart."""
        chart_config = {
            'type': 'line',
            'title': 'Conflicting Chart',
            'categories': 'A2:A4',
            'values': 'B2:B4',
            'position': 'H2'
        }
        
        with pytest.raises(ValueError) as exc_info:
            excel_tool.add_chart(str(sample_file_with_chart), 'Sheet1', chart_config)
        
        error_msg = str(exc_info.value)
        assert 'Vendas por Produto' in error_msg
        assert 'replace_existing=True' in error_msg
    
    def test_chart_without_title_still_validates_position(self, excel_tool, tmp_path):
        """Test that validation works even when existing chart has no title."""
        file_path = tmp_path / "test_no_title.xlsx"
        
        data = {
            "Sheet1": [
                ["Produto", "Vendas"],
                ["A", 100],
                ["B", 200]
            ]
        }
        excel_tool.create_excel(str(file_path), data)
        
        # Add chart without title
        chart_config_1 = {
            'type': 'column',
            'categories': 'A2:A3',
            'values': 'B2:B3',
            'position': 'D2'
        }
        excel_tool.add_chart(str(file_path), 'Sheet1', chart_config_1)
        
        # Try to add another chart at same position
        chart_config_2 = {
            'type': 'pie',
            'title': 'New Chart',
            'categories': 'A2:A3',
            'values': 'B2:B3',
            'position': 'D2'
        }
        
        with pytest.raises(ValueError) as exc_info:
            excel_tool.add_chart(str(file_path), 'Sheet1', chart_config_2)
        
        error_msg = str(exc_info.value)
        assert "Position 'D2' is already occupied" in error_msg
    
    def test_find_chart_at_position_helper(self, excel_tool, sample_file_with_chart):
        """Test the _find_chart_at_position helper method."""
        from openpyxl import load_workbook
        
        wb = load_workbook(str(sample_file_with_chart))
        ws = wb['Sheet1']
        
        # Should find chart at H2
        chart = excel_tool._find_chart_at_position(ws, 'H2')
        assert chart is not None
        
        # Should not find chart at K2
        chart = excel_tool._find_chart_at_position(ws, 'K2')
        assert chart is None
        
        wb.close()
    
    def test_get_chart_title_helper(self, excel_tool, sample_file_with_chart):
        """Test the _get_chart_title helper method."""
        from openpyxl import load_workbook
        
        wb = load_workbook(str(sample_file_with_chart))
        ws = wb['Sheet1']
        
        chart = ws._charts[0]
        title = excel_tool._get_chart_title(chart)
        
        assert title == 'Vendas por Produto'
        
        wb.close()
    
    def test_validation_with_invalid_position_format(self, excel_tool, tmp_path):
        """Test that invalid position format is handled gracefully."""
        file_path = tmp_path / "test_invalid_position.xlsx"
        
        data = {
            "Sheet1": [
                ["Produto", "Vendas"],
                ["A", 100]
            ]
        }
        excel_tool.create_excel(str(file_path), data)
        
        # Invalid position format should not crash
        chart_config = {
            'type': 'column',
            'title': 'Test',
            'categories': 'A2:A2',
            'values': 'B2:B2',
            'position': 'INVALID'
        }
        
        # Should fail at openpyxl level, not at validation level
        with pytest.raises(Exception):  # Will fail when trying to add chart
            excel_tool.add_chart(str(file_path), 'Sheet1', chart_config)
