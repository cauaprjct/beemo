"""
Testes para a funcionalidade de deletar gráficos (delete_chart) do ExcelTool.
"""

import pytest
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

from src.excel_tool import ExcelTool
from src.exceptions import CorruptedFileError


@pytest.fixture
def excel_tool():
    """Fixture que retorna uma instância do ExcelTool."""
    return ExcelTool()


@pytest.fixture
def sample_file_with_charts(tmp_path):
    """Cria um arquivo Excel de teste com múltiplos gráficos."""
    file_path = tmp_path / "test_charts.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Adicionar dados
    data = [
        ["Produto", "Vendas"],
        ["A", 100],
        ["B", 200],
        ["C", 150],
        ["D", 300]
    ]
    
    for row in data:
        ws.append(row)
    
    # Adicionar 3 gráficos diferentes
    # Gráfico 1: BarChart com título
    chart1 = BarChart()
    chart1.title = "Vendas por Produto"
    chart1.type = "col"
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=5)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=5)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    ws.add_chart(chart1, "D2")
    
    # Gráfico 2: PieChart com título
    chart2 = PieChart()
    chart2.title = "Distribuição"
    data_ref2 = Reference(ws, min_col=2, min_row=2, max_row=5)
    cats_ref2 = Reference(ws, min_col=1, min_row=2, max_row=5)
    chart2.add_data(data_ref2, titles_from_data=False)
    chart2.set_categories(cats_ref2)
    ws.add_chart(chart2, "D15")
    
    # Gráfico 3: LineChart sem título
    chart3 = LineChart()
    data_ref3 = Reference(ws, min_col=2, min_row=1, max_row=5)
    chart3.add_data(data_ref3, titles_from_data=True)
    ws.add_chart(chart3, "H2")
    
    wb.save(file_path)
    wb.close()
    
    return str(file_path)


class TestDeleteChartFunctionality:
    """Testes de funcionalidade básica do delete_chart."""
    
    def test_delete_chart_by_index_first(self, excel_tool, sample_file_with_charts):
        """Testa deletar o primeiro gráfico por índice."""
        # Verificar que existem 3 gráficos
        result_before = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        assert result_before['total_count'] == 3
        
        # Deletar primeiro gráfico (índice 0)
        excel_tool.delete_chart(sample_file_with_charts, "Sheet1", 0)
        
        # Verificar que agora existem 2 gráficos
        result_after = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        assert result_after['total_count'] == 2
        
        # Verificar que o gráfico deletado era "Vendas por Produto"
        titles_after = [chart['title'] for chart in result_after['charts']]
        assert "Vendas por Produto" not in titles_after
        assert "Distribuição" in titles_after
    
    def test_delete_chart_by_index_middle(self, excel_tool, sample_file_with_charts):
        """Testa deletar o gráfico do meio por índice."""
        # Deletar segundo gráfico (índice 1)
        excel_tool.delete_chart(sample_file_with_charts, "Sheet1", 1)
        
        # Verificar que agora existem 2 gráficos
        result_after = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        assert result_after['total_count'] == 2
        
        # Verificar que o gráfico deletado era "Distribuição"
        titles_after = [chart['title'] for chart in result_after['charts']]
        assert "Distribuição" not in titles_after
        assert "Vendas por Produto" in titles_after
    
    def test_delete_chart_by_index_last(self, excel_tool, sample_file_with_charts):
        """Testa deletar o último gráfico por índice."""
        # Deletar terceiro gráfico (índice 2)
        excel_tool.delete_chart(sample_file_with_charts, "Sheet1", 2)
        
        # Verificar que agora existem 2 gráficos
        result_after = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        assert result_after['total_count'] == 2
        
        # Verificar que os dois primeiros gráficos ainda existem
        titles_after = [chart['title'] for chart in result_after['charts']]
        assert "Vendas por Produto" in titles_after
        assert "Distribuição" in titles_after
    
    def test_delete_chart_by_title(self, excel_tool, sample_file_with_charts):
        """Testa deletar gráfico por título."""
        # Deletar gráfico por título
        excel_tool.delete_chart(sample_file_with_charts, "Sheet1", "Distribuição")
        
        # Verificar que agora existem 2 gráficos
        result_after = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        assert result_after['total_count'] == 2
        
        # Verificar que o gráfico deletado não existe mais
        titles_after = [chart['title'] for chart in result_after['charts']]
        assert "Distribuição" not in titles_after
        assert "Vendas por Produto" in titles_after
    
    def test_delete_all_charts_sequentially(self, excel_tool, sample_file_with_charts):
        """Testa deletar todos os gráficos sequencialmente."""
        # Deletar todos os gráficos um por um (sempre deletando o índice 0)
        excel_tool.delete_chart(sample_file_with_charts, "Sheet1", 0)
        result1 = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        assert result1['total_count'] == 2
        
        excel_tool.delete_chart(sample_file_with_charts, "Sheet1", 0)
        result2 = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        assert result2['total_count'] == 1
        
        excel_tool.delete_chart(sample_file_with_charts, "Sheet1", 0)
        result3 = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        assert result3['total_count'] == 0
    
    def test_delete_chart_file_persists(self, excel_tool, sample_file_with_charts):
        """Testa que a deleção persiste após reabrir o arquivo."""
        # Deletar um gráfico
        excel_tool.delete_chart(sample_file_with_charts, "Sheet1", "Vendas por Produto")
        
        # Verificar imediatamente
        result1 = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        assert result1['total_count'] == 2
        
        # Criar nova instância do tool e verificar novamente
        new_tool = ExcelTool()
        result2 = new_tool.list_charts(sample_file_with_charts, "Sheet1")
        assert result2['total_count'] == 2
        
        titles = [chart['title'] for chart in result2['charts']]
        assert "Vendas por Produto" not in titles


class TestDeleteChartValidation:
    """Testes de validação e tratamento de erros."""
    
    def test_delete_chart_invalid_index_negative(self, excel_tool, sample_file_with_charts):
        """Testa erro ao tentar deletar com índice negativo."""
        with pytest.raises(ValueError) as exc_info:
            excel_tool.delete_chart(sample_file_with_charts, "Sheet1", -1)
        
        assert "out of range" in str(exc_info.value)
    
    def test_delete_chart_invalid_index_too_high(self, excel_tool, sample_file_with_charts):
        """Testa erro ao tentar deletar com índice muito alto."""
        with pytest.raises(ValueError) as exc_info:
            excel_tool.delete_chart(sample_file_with_charts, "Sheet1", 10)
        
        assert "out of range" in str(exc_info.value)
        assert "has 3 chart(s)" in str(exc_info.value)
    
    def test_delete_chart_invalid_title(self, excel_tool, sample_file_with_charts):
        """Testa erro ao tentar deletar com título inexistente."""
        with pytest.raises(ValueError) as exc_info:
            excel_tool.delete_chart(sample_file_with_charts, "Sheet1", "Gráfico Inexistente")
        
        assert "not found" in str(exc_info.value)
        assert "Available charts:" in str(exc_info.value)
    
    def test_delete_chart_invalid_sheet(self, excel_tool, sample_file_with_charts):
        """Testa erro ao especificar sheet inexistente."""
        with pytest.raises(ValueError) as exc_info:
            excel_tool.delete_chart(sample_file_with_charts, "SheetInexistente", 0)
        
        assert "not found" in str(exc_info.value)
    
    def test_delete_chart_file_not_found(self, excel_tool):
        """Testa erro ao especificar arquivo inexistente."""
        with pytest.raises(FileNotFoundError):
            excel_tool.delete_chart("arquivo_inexistente.xlsx", "Sheet1", 0)
    
    def test_delete_chart_invalid_identifier_type(self, excel_tool, sample_file_with_charts):
        """Testa erro ao usar tipo inválido de identificador."""
        with pytest.raises(ValueError) as exc_info:
            excel_tool.delete_chart(sample_file_with_charts, "Sheet1", 3.14)
        
        assert "Invalid identifier type" in str(exc_info.value)
        assert "Must be int (index) or str (title)" in str(exc_info.value)
    
    def test_delete_chart_from_empty_sheet(self, excel_tool, tmp_path):
        """Testa erro ao tentar deletar de sheet sem gráficos."""
        # Criar arquivo sem gráficos
        file_path = tmp_path / "no_charts.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Data", "Value"])
        ws.append([1, 100])
        wb.save(file_path)
        wb.close()
        
        # Tentar deletar gráfico
        with pytest.raises(ValueError) as exc_info:
            excel_tool.delete_chart(str(file_path), "Sheet1", 0)
        
        assert "out of range" in str(exc_info.value)
        assert "has 0 chart(s)" in str(exc_info.value)


class TestDeleteChartEdgeCases:
    """Testes de casos extremos."""
    
    def test_delete_chart_with_special_characters_in_title(self, excel_tool, tmp_path):
        """Testa deletar gráfico com caracteres especiais no título."""
        file_path = tmp_path / "special_chars.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Adicionar dados
        ws.append(["A", "B"])
        ws.append([1, 2])
        ws.append([3, 4])
        
        # Adicionar gráfico com título especial
        chart = BarChart()
        chart.title = "Vendas 2025 (R$) - 100%"
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=3)
        chart.add_data(data_ref, titles_from_data=True)
        ws.add_chart(chart, "D2")
        
        wb.save(file_path)
        wb.close()
        
        # Deletar por título
        excel_tool.delete_chart(str(file_path), "Sheet1", "Vendas 2025 (R$) - 100%")
        
        # Verificar que foi deletado
        result = excel_tool.list_charts(str(file_path), "Sheet1")
        assert result['total_count'] == 0
    
    def test_delete_chart_case_sensitive_title(self, excel_tool, sample_file_with_charts):
        """Testa que a busca por título é case-sensitive."""
        # Tentar deletar com título em minúsculas (deve falhar)
        with pytest.raises(ValueError) as exc_info:
            excel_tool.delete_chart(sample_file_with_charts, "Sheet1", "vendas por produto")
        
        assert "not found" in str(exc_info.value)
        
        # Deletar com título correto (deve funcionar)
        excel_tool.delete_chart(sample_file_with_charts, "Sheet1", "Vendas por Produto")
        
        result = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        assert result['total_count'] == 2
    
    def test_delete_chart_updates_indices(self, excel_tool, sample_file_with_charts):
        """Testa que os índices são atualizados após deleção."""
        # Listar gráficos antes
        result_before = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        chart1_title = result_before['charts'][1]['title']  # Segundo gráfico
        
        # Deletar primeiro gráfico
        excel_tool.delete_chart(sample_file_with_charts, "Sheet1", 0)
        
        # Verificar que o segundo gráfico agora é o primeiro (índice 0)
        result_after = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        assert result_after['charts'][0]['title'] == chart1_title
        assert result_after['charts'][0]['index'] == 0


class TestDeleteChartIntegration:
    """Testes de integração com outras operações."""
    
    def test_delete_then_add_chart(self, excel_tool, sample_file_with_charts):
        """Testa deletar e depois adicionar um novo gráfico."""
        # Deletar gráfico existente
        excel_tool.delete_chart(sample_file_with_charts, "Sheet1", "Vendas por Produto")
        
        # Adicionar novo gráfico na mesma posição
        chart_config = {
            'type': 'bar',
            'title': 'Novo Gráfico',
            'categories': 'A2:A5',
            'values': 'B2:B5',
            'position': 'D2'
        }
        excel_tool.add_chart(sample_file_with_charts, "Sheet1", chart_config)
        
        # Verificar que o novo gráfico foi adicionado
        result = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        assert result['total_count'] == 3
        
        titles = [chart['title'] for chart in result['charts']]
        assert "Novo Gráfico" in titles
        assert "Vendas por Produto" not in titles
    
    def test_list_then_delete_workflow(self, excel_tool, sample_file_with_charts):
        """Testa workflow de listar e depois deletar."""
        # Listar gráficos
        result = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        
        # Encontrar gráfico específico
        chart_to_delete = None
        for chart in result['charts']:
            if chart['title'] == "Distribuição":
                chart_to_delete = chart
                break
        
        assert chart_to_delete is not None
        
        # Deletar por índice encontrado
        excel_tool.delete_chart(sample_file_with_charts, "Sheet1", chart_to_delete['index'])
        
        # Verificar que foi deletado
        result_after = excel_tool.list_charts(sample_file_with_charts, "Sheet1")
        titles_after = [chart['title'] for chart in result_after['charts']]
        assert "Distribuição" not in titles_after
