"""Tests for insert_rows and insert_columns functionality in ExcelTool."""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from src.excel_tool import ExcelTool
from src.exceptions import CorruptedFileError


@pytest.fixture
def excel_tool():
    """Create ExcelTool instance."""
    return ExcelTool()


@pytest.fixture
def sample_file(tmp_path):
    """Create a sample Excel file with data."""
    file_path = tmp_path / "test.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add data (5 rows, 5 columns)
    ws.append(["A1", "B1", "C1", "D1", "E1"])
    ws.append(["A2", "B2", "C2", "D2", "E2"])
    ws.append(["A3", "B3", "C3", "D3", "E3"])
    ws.append(["A4", "B4", "C4", "D4", "E4"])
    ws.append(["A5", "B5", "C5", "D5", "E5"])
    
    wb.save(file_path)
    wb.close()
    
    return str(file_path)


class TestInsertRows:
    """Test inserting rows."""
    
    def test_insert_single_row_middle(self, excel_tool, sample_file):
        """Test inserting one row in the middle."""
        excel_tool.insert_rows(sample_file, 'Sheet1', 3, 1)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Should now have 6 rows
        assert ws.max_row == 6
        
        # Row 3 should be empty
        assert ws.cell(3, 1).value is None
        assert ws.cell(3, 2).value is None
        
        # Original row 3 should now be row 4
        assert ws.cell(4, 1).value == "A3"
        assert ws.cell(4, 2).value == "B3"
        
        # Rows 1-2 should be unchanged
        assert ws.cell(1, 1).value == "A1"
        assert ws.cell(2, 1).value == "A2"
        
        wb.close()
    
    def test_insert_multiple_rows(self, excel_tool, sample_file):
        """Test inserting multiple rows."""
        excel_tool.insert_rows(sample_file, 'Sheet1', 2, 3)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Should now have 8 rows (5 + 3)
        assert ws.max_row == 8
        
        # Rows 2, 3, 4 should be empty
        for row in [2, 3, 4]:
            assert ws.cell(row, 1).value is None
        
        # Original row 2 should now be row 5
        assert ws.cell(5, 1).value == "A2"
        
        wb.close()
    
    def test_insert_row_at_top(self, excel_tool, sample_file):
        """Test inserting row at the very top."""
        excel_tool.insert_rows(sample_file, 'Sheet1', 1, 1)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Row 1 should be empty
        assert ws.cell(1, 1).value is None
        
        # Original row 1 should now be row 2
        assert ws.cell(2, 1).value == "A1"
        
        wb.close()
    
    def test_insert_row_at_end(self, excel_tool, sample_file):
        """Test inserting row at the end."""
        excel_tool.insert_rows(sample_file, 'Sheet1', 6, 2)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Inserting beyond max_row doesn't change max_row until data is added
        # Original data should be unchanged
        assert ws.cell(5, 1).value == "A5"
        
        # Can add data to new rows
        ws.cell(6, 1, "NEW1")
        ws.cell(7, 1, "NEW2")
        
        # Now max_row should reflect the new data
        assert ws.max_row == 7
        
        wb.close()
    
    def test_insert_rows_preserves_data(self, excel_tool, sample_file):
        """Test that inserting rows preserves all data correctly."""
        excel_tool.insert_rows(sample_file, 'Sheet1', 3, 1)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Check all columns are preserved
        assert ws.cell(4, 1).value == "A3"
        assert ws.cell(4, 2).value == "B3"
        assert ws.cell(4, 3).value == "C3"
        assert ws.cell(4, 4).value == "D3"
        assert ws.cell(4, 5).value == "E3"
        
        wb.close()


class TestInsertColumns:
    """Test inserting columns."""
    
    def test_insert_single_column_by_letter(self, excel_tool, sample_file):
        """Test inserting one column using letter."""
        excel_tool.insert_columns(sample_file, 'Sheet1', 'C', 1)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Should now have 6 columns
        assert ws.max_column == 6
        
        # Column C should be empty
        assert ws.cell(1, 3).value is None
        assert ws.cell(2, 3).value is None
        
        # Original column C should now be column D
        assert ws.cell(1, 4).value == "C1"
        assert ws.cell(2, 4).value == "C2"
        
        # Columns A-B should be unchanged
        assert ws.cell(1, 1).value == "A1"
        assert ws.cell(1, 2).value == "B1"
        
        wb.close()
    
    def test_insert_single_column_by_number(self, excel_tool, sample_file):
        """Test inserting one column using number."""
        excel_tool.insert_columns(sample_file, 'Sheet1', 3, 1)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Should now have 6 columns
        assert ws.max_column == 6
        
        # Column 3 (C) should be empty
        assert ws.cell(1, 3).value is None
        
        # Original column 3 should now be column 4
        assert ws.cell(1, 4).value == "C1"
        
        wb.close()
    
    def test_insert_multiple_columns(self, excel_tool, sample_file):
        """Test inserting multiple columns."""
        excel_tool.insert_columns(sample_file, 'Sheet1', 'B', 2)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Should now have 7 columns (5 + 2)
        assert ws.max_column == 7
        
        # Columns B and C should be empty
        assert ws.cell(1, 2).value is None
        assert ws.cell(1, 3).value is None
        
        # Original column B should now be column D
        assert ws.cell(1, 4).value == "B1"
        
        wb.close()
    
    def test_insert_column_at_start(self, excel_tool, sample_file):
        """Test inserting column at the very start."""
        excel_tool.insert_columns(sample_file, 'Sheet1', 'A', 1)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Column A should be empty
        assert ws.cell(1, 1).value is None
        
        # Original column A should now be column B
        assert ws.cell(1, 2).value == "A1"
        
        wb.close()
    
    def test_insert_column_at_end(self, excel_tool, sample_file):
        """Test inserting column at the end."""
        excel_tool.insert_columns(sample_file, 'Sheet1', 6, 1)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Inserting beyond max_column doesn't change max_column until data is added
        # Original data should be unchanged
        assert ws.cell(1, 5).value == "E1"
        
        # Can add data to new column
        ws.cell(1, 6, "NEW")
        
        # Now max_column should reflect the new data
        assert ws.max_column == 6
        
        wb.close()
    
    def test_insert_columns_preserves_data(self, excel_tool, sample_file):
        """Test that inserting columns preserves all data correctly."""
        excel_tool.insert_columns(sample_file, 'Sheet1', 'C', 1)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Check all rows are preserved in shifted column
        assert ws.cell(1, 4).value == "C1"
        assert ws.cell(2, 4).value == "C2"
        assert ws.cell(3, 4).value == "C3"
        assert ws.cell(4, 4).value == "C4"
        assert ws.cell(5, 4).value == "C5"
        
        wb.close()


class TestInsertValidation:
    """Test validation and error handling."""
    
    def test_insert_rows_invalid_start_row(self, excel_tool, sample_file):
        """Test error for invalid start_row."""
        with pytest.raises(ValueError, match="start_row must be >= 1"):
            excel_tool.insert_rows(sample_file, 'Sheet1', 0, 1)
    
    def test_insert_rows_invalid_count(self, excel_tool, sample_file):
        """Test error for invalid count."""
        with pytest.raises(ValueError, match="count must be >= 1"):
            excel_tool.insert_rows(sample_file, 'Sheet1', 1, 0)
    
    def test_insert_rows_invalid_sheet(self, excel_tool, sample_file):
        """Test error for invalid sheet name."""
        with pytest.raises(ValueError, match="Sheet .* not found"):
            excel_tool.insert_rows(sample_file, 'NonExistent', 1, 1)
    
    def test_insert_rows_file_not_found(self, excel_tool):
        """Test error for non-existent file."""
        with pytest.raises(FileNotFoundError):
            excel_tool.insert_rows('nonexistent.xlsx', 'Sheet1', 1, 1)
    
    def test_insert_columns_invalid_start_col(self, excel_tool, sample_file):
        """Test error for invalid start_col."""
        with pytest.raises(ValueError, match="start_col must be >= 1"):
            excel_tool.insert_columns(sample_file, 'Sheet1', 0, 1)
    
    def test_insert_columns_invalid_count(self, excel_tool, sample_file):
        """Test error for invalid count."""
        with pytest.raises(ValueError, match="count must be >= 1"):
            excel_tool.insert_columns(sample_file, 'Sheet1', 1, 0)
    
    def test_insert_columns_invalid_type(self, excel_tool, sample_file):
        """Test error for invalid column type."""
        with pytest.raises(ValueError, match="Invalid column type"):
            excel_tool.insert_columns(sample_file, 'Sheet1', 3.14, 1)
    
    def test_insert_columns_invalid_sheet(self, excel_tool, sample_file):
        """Test error for invalid sheet name."""
        with pytest.raises(ValueError, match="Sheet .* not found"):
            excel_tool.insert_columns(sample_file, 'NonExistent', 'A', 1)
    
    def test_insert_columns_file_not_found(self, excel_tool):
        """Test error for non-existent file."""
        with pytest.raises(FileNotFoundError):
            excel_tool.insert_columns('nonexistent.xlsx', 'Sheet1', 'A', 1)


class TestInsertIntegration:
    """Test integration scenarios."""
    
    def test_insert_rows_then_add_data(self, excel_tool, sample_file):
        """Test inserting rows and then adding data."""
        # Insert row
        excel_tool.insert_rows(sample_file, 'Sheet1', 3, 1)
        
        # Add data to new row
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        ws.cell(3, 1, "NEW")
        ws.cell(3, 2, "DATA")
        wb.save(sample_file)
        wb.close()
        
        # Verify
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.cell(3, 1).value == "NEW"
        assert ws.cell(4, 1).value == "A3"  # Original row 3 shifted down
        wb.close()
    
    def test_insert_columns_then_add_data(self, excel_tool, sample_file):
        """Test inserting columns and then adding data."""
        # Insert column
        excel_tool.insert_columns(sample_file, 'Sheet1', 'C', 1)
        
        # Add data to new column
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        ws.cell(1, 3, "NEW")
        ws.cell(2, 3, "COL")
        wb.save(sample_file)
        wb.close()
        
        # Verify
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.cell(1, 3).value == "NEW"
        assert ws.cell(1, 4).value == "C1"  # Original column C shifted right
        wb.close()
    
    def test_insert_multiple_times(self, excel_tool, sample_file):
        """Test inserting rows multiple times."""
        excel_tool.insert_rows(sample_file, 'Sheet1', 2, 1)
        excel_tool.insert_rows(sample_file, 'Sheet1', 4, 1)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Should have 7 rows (5 + 2)
        assert ws.max_row == 7
        
        wb.close()
    
    def test_insert_rows_and_columns(self, excel_tool, sample_file):
        """Test inserting both rows and columns."""
        excel_tool.insert_rows(sample_file, 'Sheet1', 3, 1)
        excel_tool.insert_columns(sample_file, 'Sheet1', 'C', 1)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Should have 6 rows and 6 columns
        assert ws.max_row == 6
        assert ws.max_column == 6
        
        # Original A1 should still be at A1
        assert ws.cell(1, 1).value == "A1"
        
        # Original C3 should now be at D4 (column shifted right, row shifted down)
        assert ws.cell(4, 4).value == "C3"
        
        wb.close()
    
    def test_default_count_parameter(self, excel_tool, sample_file):
        """Test that count defaults to 1 when not specified."""
        # This is tested implicitly by the agent integration
        # but we verify the method works without explicit count
        excel_tool.insert_rows(sample_file, 'Sheet1', 3)
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Should have 6 rows (5 + 1)
        assert ws.max_row == 6
        
        wb.close()
