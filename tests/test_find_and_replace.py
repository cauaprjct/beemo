"""Tests for find_and_replace functionality in ExcelTool."""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from src.excel_tool import ExcelTool
from src.exceptions import CorruptedFileError


@pytest.fixture
def excel_tool():
    """Create ExcelTool instance."""
    return ExcelTool()


@pytest.fixture
def sample_file(tmp_path):
    """Create a sample Excel file with text data."""
    file_path = tmp_path / "test.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add sample data with various text
    ws.append(["Product", "Status", "Year", "Notes"])
    ws.append(["Product A", "Pending", "2024", "Review in 2024"])
    ws.append(["Product B", "Active", "2024", "Updated 2024"])
    ws.append(["Product A", "Pending", "2023", "Old product"])
    ws.append(["Product C", "Completed", "2024", "Done"])
    
    # Add second sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Item", "Price"])
    ws2.append(["Product A", "100"])
    ws2.append(["Product B", "200"])
    
    wb.save(file_path)
    wb.close()
    
    return str(file_path)


class TestFindAndReplaceBasic:
    """Test basic find and replace functionality."""
    
    def test_replace_simple_text(self, excel_tool, sample_file):
        """Test replacing simple text."""
        result = excel_tool.find_and_replace(sample_file, "2024", "2025")
        
        # Should replace 5 occurrences total (3 in Year column + 2 in Notes column)
        assert result['replacements_count'] == 5
        assert len(result['sheets_processed']) == 2
        
        # Verify replacements
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.cell(2, 3).value == "2025"  # Year column
        assert ws.cell(2, 4).value == "Review in 2025"  # Notes
        assert ws.cell(3, 3).value == "2025"
        wb.close()
    
    def test_replace_in_specific_sheet(self, excel_tool, sample_file):
        """Test replacing in specific sheet only."""
        result = excel_tool.find_and_replace(
            sample_file, "Product A", "Product Alpha", sheet="Sheet1"
        )
        
        # Should replace 2 occurrences in Sheet1 only
        assert result['replacements_count'] == 2
        assert result['sheets_processed'] == ["Sheet1"]
        
        # Verify Sheet1 was updated
        wb = load_workbook(sample_file)
        ws1 = wb['Sheet1']
        assert ws1.cell(2, 1).value == "Product Alpha"
        assert ws1.cell(4, 1).value == "Product Alpha"
        
        # Verify Sheet2 was NOT updated
        ws2 = wb['Sheet2']
        assert ws2.cell(2, 1).value == "Product A"  # Unchanged
        wb.close()
    
    def test_replace_all_sheets(self, excel_tool, sample_file):
        """Test replacing across all sheets."""
        result = excel_tool.find_and_replace(
            sample_file, "Product A", "Product Alpha"
        )
        
        # Should replace in both sheets
        assert result['replacements_count'] == 3  # 2 in Sheet1, 1 in Sheet2
        assert len(result['sheets_processed']) == 2
        
        # Verify both sheets updated
        wb = load_workbook(sample_file)
        assert wb['Sheet1'].cell(2, 1).value == "Product Alpha"
        assert wb['Sheet2'].cell(2, 1).value == "Product Alpha"
        wb.close()


class TestFindAndReplaceOptions:
    """Test find and replace with different options."""
    
    def test_case_sensitive_match(self, excel_tool, sample_file):
        """Test case-sensitive replacement."""
        # Add mixed case data
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        ws.cell(6, 1, "product a")  # lowercase
        ws.cell(7, 1, "PRODUCT A")  # uppercase
        wb.save(sample_file)
        wb.close()
        
        # Replace only exact case
        result = excel_tool.find_and_replace(
            sample_file, "Product A", "Product Alpha",
            match_case=True
        )
        
        # Should only replace exact case matches
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.cell(2, 1).value == "Product Alpha"  # Replaced
        assert ws.cell(6, 1).value == "product a"  # Not replaced (lowercase)
        assert ws.cell(7, 1).value == "PRODUCT A"  # Not replaced (uppercase)
        wb.close()
    
    def test_case_insensitive_match(self, excel_tool, sample_file):
        """Test case-insensitive replacement (default)."""
        # Add mixed case data
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        ws.cell(6, 1, "product a")
        ws.cell(7, 1, "PRODUCT A")
        wb.save(sample_file)
        wb.close()
        
        # Replace all cases
        result = excel_tool.find_and_replace(
            sample_file, "Product A", "Product Alpha",
            match_case=False
        )
        
        # Should replace all case variations
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.cell(2, 1).value == "Product Alpha"
        assert ws.cell(6, 1).value == "Product Alpha"
        assert ws.cell(7, 1).value == "Product Alpha"
        wb.close()
    
    def test_match_entire_cell(self, excel_tool, sample_file):
        """Test matching entire cell content only."""
        result = excel_tool.find_and_replace(
            sample_file, "Product A", "Product Alpha",
            match_entire_cell=True
        )
        
        # Should only replace cells with exactly "Product A"
        # Not cells like "Review in 2024" that contain other text
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.cell(2, 1).value == "Product Alpha"  # Exact match
        assert ws.cell(4, 1).value == "Product Alpha"  # Exact match
        wb.close()
    
    def test_substring_match(self, excel_tool, sample_file):
        """Test substring matching (default)."""
        result = excel_tool.find_and_replace(
            sample_file, "Prod", "Item"
        )
        
        # Should replace substring in all occurrences
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.cell(1, 1).value == "Itemuct"  # "Product" -> "Itemuct"
        assert ws.cell(2, 1).value == "Itemuct A"
        wb.close()


class TestFindAndReplaceEdgeCases:
    """Test edge cases."""
    
    def test_no_matches_found(self, excel_tool, sample_file):
        """Test when no matches are found."""
        result = excel_tool.find_and_replace(
            sample_file, "NonExistent", "Replacement"
        )
        
        assert result['replacements_count'] == 0
        assert len(result['sheets_processed']) == 2
    
    def test_replace_with_empty_string(self, excel_tool, sample_file):
        """Test replacing with empty string (deletion)."""
        result = excel_tool.find_and_replace(
            sample_file, "Product ", ""
        )
        
        # Should remove "Product " prefix
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.cell(2, 1).value == "A"  # "Product A" -> "A"
        assert ws.cell(3, 1).value == "B"
        wb.close()
    
    def test_replace_numbers(self, excel_tool, sample_file):
        """Test replacing numeric values."""
        result = excel_tool.find_and_replace(
            sample_file, "100", "150"
        )
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet2']
        assert ws.cell(2, 2).value == "150"
        wb.close()
    
    def test_replace_in_empty_cells_ignored(self, excel_tool, sample_file):
        """Test that empty cells are ignored."""
        # Add empty cells
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        ws.cell(10, 1, None)
        ws.cell(11, 1, "")
        wb.save(sample_file)
        wb.close()
        
        # Should not error on empty cells
        result = excel_tool.find_and_replace(
            sample_file, "test", "replacement"
        )
        
        assert result['replacements_count'] == 0


class TestFindAndReplaceValidation:
    """Test validation and error handling."""
    
    def test_empty_find_text(self, excel_tool, sample_file):
        """Test error for empty find_text."""
        with pytest.raises(ValueError, match="find_text cannot be empty"):
            excel_tool.find_and_replace(sample_file, "", "replacement")
    
    def test_invalid_sheet_name(self, excel_tool, sample_file):
        """Test error for invalid sheet name."""
        with pytest.raises(ValueError, match="Sheet .* not found"):
            excel_tool.find_and_replace(
                sample_file, "test", "replacement", sheet="NonExistent"
            )
    
    def test_file_not_found(self, excel_tool):
        """Test error for non-existent file."""
        with pytest.raises(FileNotFoundError):
            excel_tool.find_and_replace(
                "nonexistent.xlsx", "test", "replacement"
            )


class TestFindAndReplaceIntegration:
    """Test integration scenarios."""
    
    def test_multiple_replacements_same_cell(self, excel_tool, sample_file):
        """Test multiple occurrences in same cell."""
        # Add cell with multiple occurrences
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        ws.cell(10, 1, "test test test")
        wb.save(sample_file)
        wb.close()
        
        result = excel_tool.find_and_replace(
            sample_file, "test", "demo"
        )
        
        # Should replace all occurrences in the cell
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        assert ws.cell(10, 1).value == "demo demo demo"
        wb.close()
    
    def test_replace_preserves_other_data(self, excel_tool, sample_file):
        """Test that replacement preserves other data."""
        result = excel_tool.find_and_replace(
            sample_file, "Pending", "In Progress"
        )
        
        wb = load_workbook(sample_file)
        ws = wb['Sheet1']
        
        # Verify replacement
        assert ws.cell(2, 2).value == "In Progress"
        
        # Verify other data unchanged
        assert ws.cell(2, 1).value == "Product A"
        assert ws.cell(2, 3).value == "2024"
        assert ws.cell(3, 2).value == "Active"  # Different status unchanged
        
        wb.close()
    
    def test_replace_then_verify(self, excel_tool, sample_file):
        """Test replacement and verification workflow."""
        # First replacement
        result1 = excel_tool.find_and_replace(
            sample_file, "2024", "2025"
        )
        assert result1['replacements_count'] == 5
        
        # Second replacement (should find 0 since already replaced)
        result2 = excel_tool.find_and_replace(
            sample_file, "2024", "2025"
        )
        assert result2['replacements_count'] == 0
        
        # Verify new value exists
        result3 = excel_tool.find_and_replace(
            sample_file, "2025", "2026"
        )
        assert result3['replacements_count'] == 5
    
    def test_result_details_structure(self, excel_tool, sample_file):
        """Test that result contains proper details."""
        result = excel_tool.find_and_replace(
            sample_file, "Product A", "Product Alpha"
        )
        
        # Check result structure
        assert 'replacements_count' in result
        assert 'sheets_processed' in result
        assert 'details' in result
        
        # Check details per sheet
        assert len(result['details']) == 2
        for detail in result['details']:
            assert 'sheet' in detail
            assert 'replacements' in detail
            assert isinstance(detail['replacements'], int)
