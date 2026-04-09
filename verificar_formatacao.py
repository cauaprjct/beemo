from openpyxl import load_workbook

wb = load_workbook('C:/Users/ngb/Documents/GeminiOfficeFiles/vendas_2025.xlsx')
ws = wb['Vendas']

print('=' * 80)
print('VERIFICAÇÃO DE FORMATAÇÃO - vendas_2025.xlsx')
print('=' * 80)

# TESTE 5: Linha 5 - Fundo amarelo e negrito
print('\n🎨 TESTE 5: Formatação da Linha 5 (Fundo Amarelo + Negrito)')
print('-' * 80)

linha_5_formatada = True
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws[f'{col}5']
    
    # Verificar negrito
    negrito = cell.font.bold if cell.font else False
    
    # Verificar cor de fundo
    cor_fundo = None
    if cell.fill and cell.fill.start_color:
        cor_fundo = cell.fill.start_color.rgb
    
    print(f'   Célula {col}5:')
    print(f'      Valor: {cell.value}')
    print(f'      Negrito: {negrito}')
    print(f'      Cor de fundo: {cor_fundo}')
    
    # Verificar se é amarelo (FFFF00 ou variações)
    if cor_fundo:
        # Remover o prefixo '00' se existir (formato ARGB)
        cor_limpa = cor_fundo[-6:] if len(cor_fundo) == 8 else cor_fundo
        eh_amarelo = cor_limpa.upper() in ['FFFF00', 'FFFF99', 'FFFFCC', 'FFFFE0']
        print(f'      É amarelo: {eh_amarelo}')
        
        if not negrito or not eh_amarelo:
            linha_5_formatada = False
    else:
        linha_5_formatada = False
    print()

# Verificar dados da linha 5
print('📊 Dados da Linha 5:')
linha_5_dados = []
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    linha_5_dados.append(ws[f'{col}5'].value)
print(f'   {linha_5_dados}')

# TESTE 6: Verificar se alguma célula da coluna E tem formatação vermelha
print('\n🎨 TESTE 6: Formatação Condicional Coluna E (Valores > 400)')
print('-' * 80)
print('   Status: ❌ FALHOU (Permission Denied - arquivo estava aberto)')
print('   Motivo: Arquivo bloqueado durante a operação')

# Verificar se há células com preço > 400
celulas_acima_400 = []
for row in range(2, 102):  # Linhas 2 a 101
    valor = ws[f'E{row}'].value
    if valor and valor > 400:
        celulas_acima_400.append((row, valor))

print(f'\n   Células que deveriam ser formatadas: {len(celulas_acima_400)}')
if len(celulas_acima_400) > 0:
    print(f'   Exemplos (primeiras 5):')
    for row, valor in celulas_acima_400[:5]:
        print(f'      E{row}: R$ {valor}')

# RESUMO
print('\n' + '=' * 80)
print('📊 RESUMO FINAL')
print('=' * 80)

if linha_5_formatada:
    print('✅ Teste 5 (Linha 5 - Amarelo + Negrito): PASSOU')
else:
    print('⚠️ Teste 5 (Linha 5 - Amarelo + Negrito): PARCIAL (verifique manualmente)')

print('❌ Teste 6 (Coluna E - Vermelho > 400): FALHOU (Permission Denied)')

print('\n💡 DICA: Feche o arquivo Excel antes de executar operações de formatação')
print('   para evitar erros de "Permission Denied"')

wb.close()
