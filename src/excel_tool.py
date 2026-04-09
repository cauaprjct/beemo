"""Excel file manipulation tool for Gemini Office Agent.

This module provides the ExcelTool class for reading, creating, and modifying
Excel files (.xlsx) using the openpyxl library.
"""

import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter, column_index_from_string

from src.exceptions import CorruptedFileError
from src.logging_config import get_logger, log_file_access_error, log_error_with_traceback

logger = get_logger(__name__)


class ExcelTool:
    """Tool for manipulating Excel files (.xlsx).
    
    Provides methods for reading, creating, updating, and adding sheets
    to Excel workbooks using openpyxl.
    """

    def read_excel(self, file_path: str) -> Dict[str, Any]:
        """Lê arquivo Excel e retorna dados estruturados.
        
        Args:
            file_path: Path to the Excel file to read
            
        Returns:
            Dictionary containing:
                - sheets: Dict mapping sheet names to their data (list of rows)
                - metadata: Dict with file metadata
                
        Raises:
            FileNotFoundError: If the file does not exist
            CorruptedFileError: If the file is corrupted or cannot be read
        """
        logger.info(f"Reading Excel file: {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path, data_only=True)
            
            # Extract data from all sheets
            sheets_data = {}
            sheets_highlighted = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    sheet_data.append(list(row))
                
                sheets_data[sheet_name] = sheet_data
                
                # Detect highlighted rows (cells with non-default background color)
                highlighted = {}
                for row in sheet.iter_rows():
                    for cell in row:
                        try:
                            fill = cell.fill
                            if fill and fill.fill_type == 'solid':
                                rgb = fill.fgColor.rgb if fill.fgColor else None
                                if rgb and rgb not in ('00000000', 'FFFFFFFF', 'FF000000'):
                                    if cell.row not in highlighted:
                                        highlighted[cell.row] = rgb
                        except Exception:
                            pass
                sheets_highlighted[sheet_name] = highlighted
            
            # Detect charts in each sheet
            sheets_charts = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_chart_list = []
                for i, chart in enumerate(sheet._charts):
                    try:
                        chart_type = type(chart).__name__.replace('Chart', '') or 'Unknown'
                        title = chart.title
                        if title is None:
                            title = f"Chart {i + 1}"
                        elif not isinstance(title, str):
                            try:
                                title = str(title)
                            except Exception:
                                title = f"Chart {i + 1}"
                        sheet_chart_list.append({
                            "index": i,
                            "type": chart_type,
                            "title": title,
                            "position": self._get_chart_position(chart),
                        })
                    except Exception:
                        sheet_chart_list.append({"index": i, "type": "Unknown", "title": f"Chart {i + 1}"})
                sheets_charts[sheet_name] = sheet_chart_list
            
            # Extract metadata
            metadata = {
                "creator": workbook.properties.creator,
                "last_modified_by": workbook.properties.lastModifiedBy,
                "created": str(workbook.properties.created) if workbook.properties.created else None,
                "modified": str(workbook.properties.modified) if workbook.properties.modified else None,
                "sheet_count": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames
            }
            
            total_charts = sum(len(v) for v in sheets_charts.values())
            logger.info(f"Successfully read Excel file with {len(sheets_data)} sheets and {total_charts} chart(s)")
            return {
                "sheets": sheets_data,
                "highlighted_rows": sheets_highlighted,
                "charts": sheets_charts,
                "metadata": metadata
            }
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except Exception as e:
            logger.error(f"Error reading Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def create_excel(self, file_path: str, data: Dict[str, List[List[Any]]]) -> None:
        """Cria novo arquivo Excel com dados fornecidos.
        
        Args:
            file_path: Path where the Excel file should be created
            data: Dictionary mapping sheet names to their data (list of rows)
            
        Raises:
            IOError: If the file cannot be created
            ValidationError: If the data structure is invalid
        """
        logger.info(f"Creating Excel file: {file_path}")
        
        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        
        # Normalize data: accept list (flat rows) or dict (sheet_name → rows)
        if isinstance(data, list):
            data = {"Sheet1": data}
        elif hasattr(data, 'items') and not isinstance(data, dict):
            # Handle proto MapComposite objects from Gemini SDK
            data = dict(data)
        
        workbook = None
        try:
            workbook = Workbook()
            
            # Se não há dados, criar uma sheet padrão vazia
            if not data:
                sheet = workbook.active
                sheet.title = "Sheet1"
            else:
                # Remove the default sheet created by Workbook()
                default_sheet = workbook.active
                workbook.remove(default_sheet)
                
                # Create sheets with data
                for sheet_name, sheet_data in data.items():
                    sheet = workbook.create_sheet(title=str(sheet_name))
                    # Normalize sheet_data: could be list of lists, or list of dicts, etc.
                    if isinstance(sheet_data, list):
                        for row_idx, row_data in enumerate(sheet_data, start=1):
                            if isinstance(row_data, (list, tuple)):
                                for col_idx, cell_value in enumerate(row_data, start=1):
                                    sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                            elif isinstance(row_data, dict):
                                # First row of dicts → write headers, then values
                                if row_idx == 1:
                                    for col_idx, key in enumerate(row_data.keys(), start=1):
                                        sheet.cell(row=1, column=col_idx, value=key)
                                for col_idx, val in enumerate(row_data.values(), start=1):
                                    sheet.cell(row=row_idx + 1, column=col_idx, value=val)
                            else:
                                # Single value per row
                                sheet.cell(row=row_idx, column=1, value=row_data)
                    elif hasattr(sheet_data, 'items'):
                        # Nested dict → keys as headers, values as data
                        for col_idx, (key, val) in enumerate(sheet_data.items(), start=1):
                            sheet.cell(row=1, column=col_idx, value=key)
                            if isinstance(val, list):
                                for row_idx, v in enumerate(val, start=2):
                                    sheet.cell(row=row_idx, column=col_idx, value=v)
            
            # Save the workbook
            workbook.save(file_path)
            
            num_sheets = len(data) if data else 1
            logger.info(f"Successfully created Excel file with {num_sheets} sheet(s)")
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {file_path} - {str(e)}")
            raise IOError(f"Failed to create Excel file: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def update_cell(self, file_path: str, sheet: str, row: int, col: int, value: Any) -> None:
        """Atualiza célula específica.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet containing the cell
            row: Row number (1-based index)
            col: Column number (1-based index)
            value: New value for the cell
            
        Raises:
            FileNotFoundError: If the file does not exist
            CorruptedFileError: If the file is corrupted
            ValueError: If the sheet name is invalid
        """
        logger.info(f"Updating cell in {file_path}: sheet={sheet}, row={row}, col={col}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                logger.error(f"Sheet '{sheet}' not found in {file_path}")
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            worksheet.cell(row=row, column=col, value=value)
            
            workbook.save(file_path)
            
            logger.info(f"Successfully updated cell at ({row}, {col}) to '{value}'")
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error updating cell in Excel file: {file_path} - {str(e)}")
            raise IOError(f"Failed to update cell in Excel file: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def add_sheet(self, file_path: str, sheet_name: str, data: List[List[Any]]) -> None:
        """Adiciona nova planilha ao arquivo.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name for the new sheet
            data: Data for the new sheet (list of rows)
            
        Raises:
            FileNotFoundError: If the file does not exist
            CorruptedFileError: If the file is corrupted
            ValueError: If a sheet with the same name already exists
        """
        logger.info(f"Adding sheet '{sheet_name}' to {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet_name in workbook.sheetnames:
                logger.error(f"Sheet '{sheet_name}' already exists in {file_path}")
                raise ValueError(f"Sheet '{sheet_name}' already exists in workbook")
            
            # Create new sheet
            new_sheet = workbook.create_sheet(title=sheet_name)
            
            # Populate with data
            for row_idx, row_data in enumerate(data, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    new_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            
            workbook.save(file_path)
            
            logger.info(f"Successfully added sheet '{sheet_name}' with {len(data)} rows")
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error adding sheet to Excel file: {file_path} - {str(e)}")
            raise IOError(f"Failed to add sheet to Excel file: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def append_rows(self, file_path: str, sheet: str, rows: List[List[Any]]) -> None:
        """Adiciona linhas ao final de uma sheet existente.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet to append rows to
            rows: List of rows to append, each row is a list of cell values
            
        Raises:
            FileNotFoundError: If the file does not exist
            CorruptedFileError: If the file is corrupted
            ValueError: If the sheet name is invalid
        """
        logger.info(f"Appending {len(rows)} rows to sheet '{sheet}' in {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                logger.error(f"Sheet '{sheet}' not found in {file_path}")
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            
            for row_data in rows:
                worksheet.append(row_data)
            
            workbook.save(file_path)
            
            logger.info(f"Successfully appended {len(rows)} rows to sheet '{sheet}'")
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error appending rows to Excel file: {file_path} - {str(e)}")
            raise IOError(f"Failed to append rows to Excel file: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def update_range(self, file_path: str, sheet: str, updates: List[Dict[str, Any]]) -> None:
        """Atualiza múltiplas células de uma vez (abre e salva o arquivo uma única vez).
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet containing the cells
            updates: List of dicts, each with keys 'row', 'col', 'value'
            
        Raises:
            FileNotFoundError: If the file does not exist
            CorruptedFileError: If the file is corrupted
            ValueError: If the sheet name is invalid or updates is empty
        """
        logger.info(f"Updating {len(updates)} cells in sheet '{sheet}' of {file_path}")
        
        if not updates:
            raise ValueError("Updates list cannot be empty")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                logger.error(f"Sheet '{sheet}' not found in {file_path}")
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            
            for i, cell_update in enumerate(updates):
                try:
                    row = cell_update['row']
                    col = cell_update['col']
                    value = cell_update['value']
                except KeyError as e:
                    missing_field = str(e).strip("'")
                    logger.error(f"Update at position {i} is missing required field '{missing_field}'")
                    logger.error(f"Invalid update: {cell_update}")
                    raise ValueError(
                        f"Update at position {i} is missing required field '{missing_field}'. "
                        f"Expected format: {{'row': int, 'col': int, 'value': any}}. "
                        f"Received: {cell_update}"
                    ) from e
                
                worksheet.cell(row=row, column=col, value=value)
            
            workbook.save(file_path)
            
            logger.info(f"Successfully updated {len(updates)} cells in sheet '{sheet}'")
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except (ValueError, KeyError):
            raise
        except Exception as e:
            logger.error(f"Error updating range in Excel file: {file_path} - {str(e)}")
            raise IOError(f"Failed to update range in Excel file: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def delete_sheet(self, file_path: str, sheet_name: str) -> None:
        """Remove uma sheet do arquivo Excel.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to delete
            
        Raises:
            FileNotFoundError: If the file does not exist
            CorruptedFileError: If the file is corrupted
            ValueError: If the sheet name is invalid or is the only sheet
        """
        logger.info(f"Deleting sheet '{sheet_name}' from {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet_name not in workbook.sheetnames:
                logger.error(f"Sheet '{sheet_name}' not found in {file_path}")
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            
            if len(workbook.sheetnames) == 1:
                raise ValueError("Cannot delete the only sheet in the workbook")
            
            del workbook[sheet_name]
            workbook.save(file_path)
            
            logger.info(f"Successfully deleted sheet '{sheet_name}'")
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error deleting sheet from Excel file: {file_path} - {str(e)}")
            raise IOError(f"Failed to delete sheet from Excel file: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def delete_rows(self, file_path: str, sheet: str, start_row: int, count: int = 1) -> None:
        """Remove linhas de uma sheet.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet
            start_row: First row to delete (1-based index)
            count: Number of rows to delete (default: 1)
            
        Raises:
            FileNotFoundError: If the file does not exist
            CorruptedFileError: If the file is corrupted
            ValueError: If the sheet name is invalid or row index is out of range
        """
        logger.info(f"Deleting {count} row(s) starting at row {start_row} in sheet '{sheet}' of {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        if start_row < 1:
            raise ValueError(f"start_row must be >= 1, got {start_row}")
        if count < 1:
            raise ValueError(f"count must be >= 1, got {count}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                logger.error(f"Sheet '{sheet}' not found in {file_path}")
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            worksheet.delete_rows(start_row, count)
            
            workbook.save(file_path)
            
            logger.info(f"Successfully deleted {count} row(s) from row {start_row}")
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error deleting rows from Excel file: {file_path} - {str(e)}")
            raise IOError(f"Failed to delete rows from Excel file: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def format_cells(self, file_path: str, sheet: str, formatting: Dict[str, Any]) -> None:
        """Aplica formatação a células ou ranges de uma sheet.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet
            formatting: Dict com as configurações de formatação:
                - range: str com range no formato "A1:C10" ou "A1" (obrigatório)
                - bold: bool (opcional)
                - italic: bool (opcional)
                - font_size: int (opcional)
                - font_color: str hex sem # ex: "FF0000" (opcional)
                - bg_color: str hex sem # ex: "FFFF00" (opcional)
                - number_format: str ex: '#,##0.00', '0%', 'dd/mm/yyyy', '"R$"#,##0.00' (opcional)
                - alignment: str 'left'|'center'|'right' (opcional)
                - border: str 'thin'|'medium'|'thick' (opcional)
                - wrap_text: bool (opcional)
            
        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If the sheet or range is invalid
        """
        cell_range = formatting.get('range', '')
        logger.info(f"Formatting cells {cell_range} in sheet '{sheet}' of {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        if not cell_range:
            raise ValueError("'range' is required in formatting parameters")
        
        # Normalize color names to hex
        COLOR_MAP = {
            'blue': '0000FF', 'darkblue': '00008B', 'lightblue': 'ADD8E6',
            'red': 'FF0000', 'darkred': '8B0000', 'lightred': 'FFB6C1',
            'green': '00FF00', 'darkgreen': '006400', 'lightgreen': '90EE90',
            'yellow': 'FFFF00', 'gold': 'FFD700',
            'black': '000000', 'white': 'FFFFFF', 'gray': '808080', 'grey': '808080',
            'orange': 'FFA500', 'purple': '800080', 'pink': 'FFC0CB',
            'brown': 'A52A2A', 'cyan': '00FFFF', 'magenta': 'FF00FF'
        }
        
        if 'bg_color' in formatting:
            color = formatting['bg_color'].lower().strip()
            if color in COLOR_MAP:
                formatting['bg_color'] = COLOR_MAP[color]
                logger.info(f"Normalized bg_color '{color}' to hex '{formatting['bg_color']}'")
        
        if 'font_color' in formatting:
            color = formatting['font_color'].lower().strip()
            if color in COLOR_MAP:
                formatting['font_color'] = COLOR_MAP[color]
                logger.info(f"Normalized font_color '{color}' to hex '{formatting['font_color']}'")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            
            # Construir objetos de estilo
            font_kwargs = {}
            if 'bold' in formatting:
                font_kwargs['bold'] = formatting['bold']
            if 'italic' in formatting:
                font_kwargs['italic'] = formatting['italic']
            if 'font_size' in formatting:
                font_kwargs['size'] = formatting['font_size']
            if 'font_color' in formatting:
                font_kwargs['color'] = formatting['font_color']
            
            font = Font(**font_kwargs) if font_kwargs else None
            
            fill = None
            if 'bg_color' in formatting:
                fill = PatternFill(start_color=formatting['bg_color'],
                                   end_color=formatting['bg_color'],
                                   fill_type='solid')
            
            alignment = None
            align_value = formatting.get('alignment')
            wrap = formatting.get('wrap_text', False)
            if align_value or wrap:
                alignment = Alignment(
                    horizontal=align_value if align_value else None,
                    wrap_text=wrap
                )
            
            border = None
            border_style = formatting.get('border')
            if border_style:
                side = Side(style=border_style)
                border = Border(left=side, right=side, top=side, bottom=side)
            
            num_format = formatting.get('number_format')
            
            # Aplicar a todas as células no range
            for row in worksheet[cell_range]:
                cells = row if isinstance(row, tuple) else (row,)
                for cell in cells:
                    if font:
                        cell.font = font
                    if fill:
                        cell.fill = fill
                    if alignment:
                        cell.alignment = alignment
                    if border:
                        cell.border = border
                    if num_format:
                        cell.number_format = num_format
            
            workbook.save(file_path)
            logger.info(f"Successfully formatted cells {cell_range}")
            
        except InvalidFileException as e:
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error formatting cells: {file_path} - {str(e)}")
            raise IOError(f"Failed to format cells: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def auto_width(self, file_path: str, sheet: str) -> None:
        """Ajusta automaticamente a largura das colunas baseado no conteúdo.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet
            
        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If the sheet is invalid
        """
        logger.info(f"Auto-adjusting column widths in sheet '{sheet}' of {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            
            for col_cells in worksheet.columns:
                max_length = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    if cell.value is not None:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
                # Adiciona margem
                adjusted_width = min(max_length + 3, 60)
                worksheet.column_dimensions[col_letter].width = adjusted_width
            
            workbook.save(file_path)
            logger.info(f"Successfully auto-adjusted column widths")
            
        except InvalidFileException as e:
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error auto-adjusting widths: {file_path} - {str(e)}")
            raise IOError(f"Failed to auto-adjust widths: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def set_formula(self, file_path: str, sheet: str, row: int, col: int, formula: str) -> None:
        """Define uma fórmula em uma célula específica.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet
            row: Row number (1-based)
            col: Column number (1-based)
            formula: Fórmula Excel (ex: '=SUM(A1:A10)', '=AVERAGE(B2:B20)')
            
        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If the sheet is invalid or formula doesn't start with =
        """
        logger.info(f"Setting formula in ({row},{col}) of sheet '{sheet}' in {file_path}")
        
        # Validate row and col are not None
        if row is None or col is None:
            raise ValueError(f"Row and column must be specified (got row={row}, col={col})")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        if not formula.startswith('='):
            formula = '=' + formula
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            worksheet.cell(row=row, column=col, value=formula)
            
            workbook.save(file_path)
            logger.info(f"Successfully set formula at ({row},{col}): {formula}")
            
        except InvalidFileException as e:
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error setting formula: {file_path} - {str(e)}")
            raise IOError(f"Failed to set formula: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def set_formulas(self, file_path: str, sheet: str, formulas: List[Dict[str, Any]]) -> None:
        """Define múltiplas fórmulas de uma vez (1 open/save).
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet
            formulas: List of dicts with 'row', 'col', 'formula'
            
        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If the sheet is invalid
        """
        logger.info(f"Setting {len(formulas)} formulas in sheet '{sheet}' of {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        if not formulas:
            raise ValueError("Formulas list cannot be empty")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            
            for f in formulas:
                row = f['row']
                col = f['col']
                formula = f['formula']
                if not formula.startswith('='):
                    formula = '=' + formula
                worksheet.cell(row=row, column=col, value=formula)
            
            workbook.save(file_path)
            logger.info(f"Successfully set {len(formulas)} formulas")
            
        except InvalidFileException as e:
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except (ValueError, KeyError):
            raise
        except Exception as e:
            logger.error(f"Error setting formulas: {file_path} - {str(e)}")
            raise IOError(f"Failed to set formulas: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def merge_cells(self, file_path: str, sheet: str, cell_range: str) -> None:
        """Mescla células em um range.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet
            cell_range: Range de células no formato "A1:C1"
            
        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If the sheet or range is invalid
        """
        logger.info(f"Merging cells {cell_range} in sheet '{sheet}' of {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            worksheet.merge_cells(cell_range)
            
            workbook.save(file_path)
            logger.info(f"Successfully merged cells {cell_range}")
            
        except InvalidFileException as e:
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error merging cells: {file_path} - {str(e)}")
            raise IOError(f"Failed to merge cells: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def add_chart(self, file_path: str, sheet: str, chart_config: Dict[str, Any]) -> None:
        """Adiciona um gráfico a uma planilha Excel.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet to add the chart to
            chart_config: Dictionary with chart configuration:
                - type: Chart type ('bar', 'column', 'line', 'pie', 'area', 'scatter')
                - data_range: Data range for the chart (e.g., 'A1:B10')
                - title: Chart title (optional)
                - position: Cell where chart will be placed (e.g., 'H2', default: 'H2')
                - x_axis_title: Title for X axis (optional)
                - y_axis_title: Title for Y axis (optional)
                - style: Chart style number 1-48 (optional, default: 10)
                - width: Chart width in cm (optional, default: 15)
                - height: Chart height in cm (optional, default: 10)
                - categories: Range for category labels (optional, e.g., 'A2:A10')
                - values: Range for values (optional, e.g., 'B2:B10')
                - replace_existing: If True, replaces chart at same position (optional, default: False)
        
        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If the sheet or chart type is invalid, or if position is occupied
            IOError: If the chart cannot be created
        """
        from openpyxl.chart import (
            BarChart, LineChart, PieChart, AreaChart, 
            ScatterChart, Reference
        )
        
        logger.info(f"Adding {chart_config.get('type', 'unknown')} chart to sheet '{sheet}' in {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Validate chart type
        chart_type = chart_config.get('type', '').lower()
        valid_types = ['bar', 'column', 'line', 'pie', 'area', 'scatter']
        if chart_type not in valid_types:
            raise ValueError(
                f"Invalid chart type: '{chart_type}'. "
                f"Must be one of: {', '.join(valid_types)}"
            )
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            
            # Get desired position
            position = chart_config.get('position') or 'H2'
            replace_existing = chart_config.get('replace_existing', False)
            
            # Validate position is not occupied (unless replace_existing=True)
            if not replace_existing:
                existing_chart = self._find_chart_at_position(worksheet, position)
                if existing_chart:
                    # Get chart title for better error message
                    chart_title = self._get_chart_title(existing_chart)
                    raise ValueError(
                        f"Position '{position}' is already occupied by a chart"
                        f"{f' titled \"{chart_title}\"' if chart_title else ''}. "
                        f"Please choose a different position or use replace_existing=True to replace it."
                    )
            
            # Create appropriate chart type
            if chart_type == 'bar':
                chart = BarChart()
                chart.type = "bar"  # Horizontal bars
            elif chart_type == 'column':
                chart = BarChart()
                chart.type = "col"  # Vertical columns
            elif chart_type == 'line':
                chart = LineChart()
            elif chart_type == 'pie':
                chart = PieChart()
            elif chart_type == 'area':
                chart = AreaChart()
            elif chart_type == 'scatter':
                chart = ScatterChart()
            
            # Set chart title
            if 'title' in chart_config:
                chart.title = chart_config['title']
            
            # Set axis titles (not applicable for pie charts)
            if chart_type != 'pie':
                if 'x_axis_title' in chart_config:
                    chart.x_axis.title = chart_config['x_axis_title']
                if 'y_axis_title' in chart_config:
                    chart.y_axis.title = chart_config['y_axis_title']
            
            # Set chart style
            if 'style' in chart_config:
                chart.style = chart_config['style']
            else:
                chart.style = 10  # Default style
            
            # Add data to chart
            if chart_config.get('data_range'):
                # Simple data range (includes both categories and values)
                data_range = chart_config['data_range']
                # Add sheet name if not present
                if '!' not in data_range:
                    data_range = f"{sheet}!{data_range}"
                data = Reference(worksheet, range_string=data_range)
                chart.add_data(data, titles_from_data=True)
            else:
                # Separate categories and values
                if 'values' not in chart_config:
                    raise ValueError("Either 'data_range' or 'values' must be specified")
                
                values_range = chart_config['values']
                # Add sheet name if not present
                if '!' not in values_range:
                    values_range = f"{sheet}!{values_range}"
                values = Reference(worksheet, range_string=values_range)
                
                if 'categories' in chart_config:
                    categories_range = chart_config['categories']
                    # Add sheet name if not present
                    if '!' not in categories_range:
                        categories_range = f"{sheet}!{categories_range}"
                    categories = Reference(worksheet, range_string=categories_range)
                    
                    # For pie charts, use titles_from_data=False and set categories separately
                    if chart_type == 'pie':
                        chart.add_data(values, titles_from_data=False)
                        chart.set_categories(categories)
                    else:
                        chart.add_data(values, titles_from_data=True)
                        chart.set_categories(categories)
                else:
                    chart.add_data(values, titles_from_data=True)
            
            # Set chart size
            if 'width' in chart_config:
                chart.width = chart_config['width']
            else:
                chart.width = 15  # Default width in cm
            
            if 'height' in chart_config:
                chart.height = chart_config['height']
            else:
                chart.height = 10  # Default height in cm
            
            # If replacing, remove existing chart at position
            if replace_existing:
                existing_chart = self._find_chart_at_position(worksheet, position)
                if existing_chart:
                    worksheet._charts.remove(existing_chart)
                    logger.info(f"Removed existing chart at position {position}")
            
            # Position chart
            worksheet.add_chart(chart, position)
            
            workbook.save(file_path)
            
            logger.info(f"Successfully added {chart_type} chart at position {position}")
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error adding chart: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "add_chart")
            raise IOError(f"Failed to add chart: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def update_chart(self, file_path: str, sheet: str, chart_index: int,
                     chart_type: str, title: str = None) -> None:
        """Altera o tipo (e opcionalmente o título) de um gráfico existente.

        Args:
            file_path: Path to the Excel file
            sheet: Sheet name containing the chart
            chart_index: 0-based index of the chart to update
            chart_type: New chart type ('bar', 'column', 'line', 'pie', 'area', 'scatter')
            title: New chart title (optional; keeps existing title if not provided)

        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If the sheet, chart index, or chart type is invalid
            IOError: If the chart cannot be updated
        """
        from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, ScatterChart

        logger.info(f"Updating chart {chart_index} to type '{chart_type}' in {file_path}")

        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        chart_type = chart_type.lower()
        valid_types = ['bar', 'column', 'line', 'pie', 'area', 'scatter']
        if chart_type not in valid_types:
            raise ValueError(
                f"Invalid chart type: '{chart_type}'. "
                f"Must be one of: {', '.join(valid_types)}"
            )

        workbook = None
        try:
            workbook = load_workbook(filename=file_path)

            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in workbook")

            ws = workbook[sheet]

            if chart_index >= len(ws._charts):
                raise ValueError(
                    f"Chart index {chart_index} out of range. "
                    f"Sheet '{sheet}' has {len(ws._charts)} chart(s)."
                )

            old_chart = ws._charts[chart_index]

            # Preserve properties from old chart
            anchor = old_chart.anchor
            keep_title = title if title is not None else old_chart.title
            width = getattr(old_chart, 'width', 15)
            height = getattr(old_chart, 'height', 10)
            style = getattr(old_chart, 'style', 10)
            old_series = list(old_chart.series)

            # Build new chart of the target type
            if chart_type == 'bar':
                new_chart = BarChart()
                new_chart.type = "bar"
            elif chart_type == 'column':
                new_chart = BarChart()
                new_chart.type = "col"
            elif chart_type == 'line':
                new_chart = LineChart()
            elif chart_type == 'pie':
                new_chart = PieChart()
            elif chart_type == 'area':
                new_chart = AreaChart()
            else:
                new_chart = ScatterChart()

            # Transfer series data (best-effort; series structure is compatible for most types)
            for series in old_series:
                try:
                    new_chart.series.append(series)
                except Exception:
                    pass

            # Apply preserved properties
            new_chart.title = keep_title
            new_chart.width = width
            new_chart.height = height
            new_chart.style = style

            # Swap: remove old chart, add new chart at the same anchor
            ws._charts.remove(old_chart)
            new_chart.anchor = anchor
            ws._charts.append(new_chart)

            workbook.save(file_path)
            logger.info(
                f"Successfully changed chart {chart_index} to '{chart_type}' in '{sheet}'"
            )

        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except (ValueError, FileNotFoundError):
            raise
        except Exception as e:
            logger.error(f"Error updating chart: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "update_chart")
            raise IOError(f"Failed to update chart: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def move_chart(self, file_path: str, sheet: str, chart_index: int, position: str) -> None:
        """Move um gráfico existente para uma nova posição na planilha.

        Args:
            file_path: Path to the Excel file
            sheet: Sheet name containing the chart
            chart_index: 0-based index of the chart to move
            position: Target cell anchor, e.g. 'H2', 'K5', 'A1'

        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If sheet, chart index, or position is invalid
            IOError: If the chart cannot be moved
        """
        logger.info(f"Moving chart {chart_index} to '{position}' in {file_path}")

        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        if not position or not isinstance(position, str):
            raise ValueError("'position' must be a non-empty cell reference string (e.g. 'H2')")

        workbook = None
        try:
            workbook = load_workbook(filename=file_path)

            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in workbook")

            ws = workbook[sheet]

            if chart_index >= len(ws._charts):
                raise ValueError(
                    f"Chart index {chart_index} out of range. "
                    f"Sheet '{sheet}' has {len(ws._charts)} chart(s)."
                )

            chart = ws._charts[chart_index]
            anchor = chart.anchor

            # --- Preserve current dimensions before changing anchor ---
            # OneCellAnchor object: size stored in ext (extent) in EMU
            # TwoCellAnchor: compute size from cell positions
            # String anchor: use chart.width/height
            EMU_PER_CM = 360000
            if hasattr(anchor, 'ext') and anchor.ext is not None:
                # OneCellAnchor object with extent
                current_w = anchor.ext.cx / EMU_PER_CM
                current_h = anchor.ext.cy / EMU_PER_CM
                logger.info(f"Preserving size from OneCellAnchor extent: {current_w:.2f}x{current_h:.2f} cm")
            elif hasattr(anchor, '_from') and hasattr(anchor, 'to'):
                # TwoCellAnchor: compute from cell positions
                current_w, current_h = self._compute_chart_size_from_anchor(ws, anchor)
                logger.info(f"Preserving size from TwoCellAnchor: {current_w:.2f}x{current_h:.2f} cm")
            else:
                # String anchor or unknown: use chart's width/height
                current_w = getattr(chart, 'width', None) or 15.0
                current_h = getattr(chart, 'height', None) or 7.5
                logger.info(f"Preserving size from chart properties: {current_w}x{current_h} cm")

            # Set new position and restore dimensions
            chart.anchor = position.upper()
            chart.width = current_w
            chart.height = current_h

            workbook.save(file_path)
            logger.info(f"Successfully moved chart {chart_index} to '{position}' (size: {current_w:.2f}x{current_h:.2f} cm)")

        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except (ValueError, FileNotFoundError):
            raise
        except Exception as e:
            logger.error(f"Error moving chart: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "move_chart")
            raise IOError(f"Failed to move chart: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    @staticmethod
    def _compute_chart_size_from_anchor(ws, anchor):
        """Compute actual chart width and height in cm from a TwoCellAnchor object.

        openpyxl TwoCellAnchor encodes chart size via two cell markers plus EMU offsets.
        Default column width = 8.43 chars; default row height = 15 pts.
        Column px  = char_width * MAX_DIGIT_WIDTH + 5  (max_digit_width = 7)
        1 pixel    = 9525 EMU  (at 96 DPI: 914400/96)
        1 point    = 12700 EMU (914400/72)
        1 cm       = 360000 EMU
        """
        EMU_PER_CM = 360000
        DEFAULT_COL_CHARS = 8.43
        DEFAULT_ROW_PTS = 15.0
        MAX_DIGIT_W = 7
        EMU_PER_PX = 9525
        PTS_TO_EMU = 12700

        try:
            from_col = anchor._from.col
            from_col_off = anchor._from.colOff
            from_row = anchor._from.row
            from_row_off = anchor._from.rowOff
            to_col = anchor.to.col
            to_col_off = anchor.to.colOff
            to_row = anchor.to.row
            to_row_off = anchor.to.rowOff

            total_col_emu = 0
            for c in range(from_col, to_col):
                col_letter = get_column_letter(c + 1)
                cd = ws.column_dimensions.get(col_letter)
                char_w = (cd.width if cd and cd.width else DEFAULT_COL_CHARS)
                total_col_emu += int((char_w * MAX_DIGIT_W + 5) * EMU_PER_PX)

            total_w_emu = total_col_emu + (to_col_off - from_col_off)
            width_cm = total_w_emu / EMU_PER_CM

            total_row_emu = 0
            for r in range(from_row, to_row):
                rd = ws.row_dimensions.get(r + 1)
                pts = (rd.height if rd and rd.height else DEFAULT_ROW_PTS)
                total_row_emu += int(pts * PTS_TO_EMU)

            total_h_emu = total_row_emu + (to_row_off - from_row_off)
            height_cm = total_h_emu / EMU_PER_CM

            return max(round(width_cm, 2), 2.0), max(round(height_cm, 2), 2.0)
        except Exception:
            return 15.0, 7.5

    @staticmethod
    def _find_safe_chart_position(ws, current_anchor: str, new_w_cm: float, new_h_cm: float) -> str:
        """Return the same anchor if the chart won't cover data cells, or a safe one.

        Checks whether the chart's top-left anchor falls inside the data region.
        If it does, the chart is moved to the first column to the right of data.
        Charts that are already to the right of or below the data are left in place.
        """
        max_data_col = 0
        max_data_row = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    if cell.column > max_data_col:
                        max_data_col = cell.column
                    if cell.row > max_data_row:
                        max_data_row = cell.row

        if max_data_col == 0:
            return current_anchor

        m = re.match(r'([A-Za-z]+)(\d+)', current_anchor)
        if not m:
            return current_anchor

        chart_col = column_index_from_string(m.group(1).upper())
        chart_row = int(m.group(2))

        # Safe if chart origin is clearly outside data bounding box
        if chart_col > max_data_col or chart_row > max_data_row:
            return current_anchor

        # Overlap: move chart 2 columns to the right of the last data column
        safe_col = get_column_letter(max_data_col + 2)
        new_anchor = f"{safe_col}{chart_row}"
        logger.info(
            f"Smart placement: chart moved from {current_anchor} to {new_anchor} "
            f"to avoid covering data (data extends to col {max_data_col}, row {max_data_row})"
        )
        return new_anchor

    def resize_chart(self, file_path: str, sheet: str, chart_index: int,
                     scale: float = None, width: float = None, height: float = None) -> None:
        """Redimensiona um gráfico existente na planilha.

        Detecta automaticamente o tipo de âncora (OneCellAnchor/TwoCellAnchor),
        calcula as dimensões reais do gráfico e converte para OneCellAnchor para
        que width/height sejam respeitados ao salvar. Evita sobrepor dados.

        Args:
            file_path: Path to the Excel file
            sheet: Sheet name containing the chart
            chart_index: 0-based index of the chart to resize
            scale: Multiplier applied to both dimensions (e.g. 1.2 = 20% bigger, 0.8 = 20% smaller)
            width: Absolute width in cm
            height: Absolute height in cm

        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If sheet, chart index, or size arguments are invalid
            IOError: If the chart cannot be resized
        """
        if scale is None and width is None and height is None:
            raise ValueError("At least one of 'scale', 'width', or 'height' must be provided")

        logger.info(f"Resizing chart {chart_index} in {file_path} "
                    f"(scale={scale}, width={width}, height={height})")

        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        workbook = None
        try:
            workbook = load_workbook(filename=file_path)

            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in workbook")

            ws = workbook[sheet]

            if chart_index >= len(ws._charts):
                raise ValueError(
                    f"Chart index {chart_index} out of range. "
                    f"Sheet '{sheet}' has {len(ws._charts)} chart(s)."
                )

            chart = ws._charts[chart_index]
            anchor = chart.anchor

            # --- Resolve anchor type and real dimensions ---
            # openpyxl anchor types:
            #   str              → OneCellAnchor written by openpyxl (chart.width/height valid)
            #   obj with _from+to → TwoCellAnchor (size encoded in cell markers, not width/height)
            #   obj with _from+ext (no to) → OneCellAnchor object (chart.width/height valid)
            #   AbsoluteAnchor   → has pos+ext, no _from
            if isinstance(anchor, str):
                top_left = anchor
                current_w = getattr(chart, 'width', None) or 15.0
                current_h = getattr(chart, 'height', None) or 7.5
                logger.info(f"String anchor '{top_left}', size: {current_w}x{current_h} cm")

            elif hasattr(anchor, '_from') and hasattr(anchor, 'to'):
                # TwoCellAnchor: actual size encoded in two cell markers — must compute
                col_letter = get_column_letter(anchor._from.col + 1)
                row_num = anchor._from.row + 1
                top_left = f"{col_letter}{row_num}"
                current_w, current_h = self._compute_chart_size_from_anchor(ws, anchor)
                logger.info(
                    f"TwoCellAnchor → top-left: {top_left}, "
                    f"computed size: {current_w}x{current_h} cm"
                )
                chart.anchor = top_left  # convert: string anchor respects width/height on save

            elif hasattr(anchor, '_from'):
                # OneCellAnchor object — chart.width/height already correct
                col_letter = get_column_letter(anchor._from.col + 1)
                row_num = anchor._from.row + 1
                top_left = f"{col_letter}{row_num}"
                current_w = getattr(chart, 'width', None) or 15.0
                current_h = getattr(chart, 'height', None) or 7.5
                logger.info(
                    f"OneCellAnchor object → top-left: {top_left}, "
                    f"size: {current_w}x{current_h} cm"
                )
                chart.anchor = top_left  # normalise to string

            else:
                # AbsoluteAnchor or unknown — keep position, use chart dimensions
                top_left = "A1"
                current_w = getattr(chart, 'width', None) or 15.0
                current_h = getattr(chart, 'height', None) or 7.5
                chart.anchor = top_left
                logger.warning(f"Unknown anchor type ({type(anchor).__name__}), resetting to A1")

            # --- Compute new dimensions ---
            if scale is not None:
                new_w = round(current_w * scale, 2)
                new_h = round(current_h * scale, 2)
            else:
                new_w = float(width) if width is not None else current_w
                new_h = float(height) if height is not None else current_h

            # Note: Smart placement disabled - user controls position via move_chart
            # Resize keeps chart at current anchor position

            chart.width = new_w
            chart.height = new_h

            workbook.save(file_path)
            logger.info(
                f"Successfully resized chart {chart_index}: "
                f"{current_w}x{current_h} → {new_w}x{new_h} cm (anchor: {chart.anchor})"
            )

        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except (ValueError, FileNotFoundError):
            raise
        except Exception as e:
            logger.error(f"Error resizing chart: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "resize_chart")
            raise IOError(f"Failed to resize chart: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def sort_data(self, file_path: str, sheet: str, sort_config: Dict[str, Any]) -> None:
        """Ordena dados em uma sheet do Excel.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet to sort
            sort_config: Dictionary with sort configuration:
                - column: Column letter or number to sort by (e.g., 'A' or 1) (required)
                - start_row: First row to include in sort (default: 2, assumes row 1 is header)
                - end_row: Last row to include in sort (optional, defaults to last row with data)
                - order: 'asc' for ascending or 'desc' for descending (default: 'asc')
                - has_header: Whether first row is header (default: True)
        
        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If the sheet or parameters are invalid
            IOError: If the sort operation fails
        
        Example:
            sort_config = {
                'column': 'B',  # or 2
                'start_row': 2,
                'order': 'asc',
                'has_header': True
            }
        """
        logger.info(f"Sorting data in sheet '{sheet}' of {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Validate sort_config
        if 'column' not in sort_config:
            raise ValueError("'column' is required in sort_config")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            
            # Parse column (can be letter like 'A' or number like 1)
            column = sort_config['column']
            if isinstance(column, str):
                # Convert column letter to number (A=1, B=2, etc)
                column = column.upper()
                if len(column) == 1 and column.isalpha():
                    sort_col_idx = ord(column) - ord('A') + 1
                else:
                    raise ValueError(f"Invalid column letter: {column}")
            elif isinstance(column, int):
                sort_col_idx = column
            else:
                raise ValueError(f"Column must be string (letter) or int (number), got {type(column)}")
            
            # Get parameters
            has_header = sort_config.get('has_header', True)
            start_row = sort_config.get('start_row', 2 if has_header else 1)
            order = sort_config.get('order', 'asc').lower()
            
            if order not in ['asc', 'desc']:
                raise ValueError(f"Order must be 'asc' or 'desc', got '{order}'")
            
            # Find last row with data
            max_row = worksheet.max_row
            end_row = sort_config.get('end_row', max_row)
            
            if start_row < 1:
                raise ValueError(f"start_row must be >= 1, got {start_row}")
            if end_row < start_row:
                raise ValueError(f"end_row ({end_row}) must be >= start_row ({start_row})")
            
            # Read all data from the range
            data_rows = []
            max_col = worksheet.max_column
            
            for row_idx in range(start_row, end_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    row_data.append(cell_value)
                data_rows.append((row_idx, row_data))
            
            # Sort data by the specified column
            reverse = (order == 'desc')
            try:
                sorted_data = sorted(
                    data_rows,
                    key=lambda x: (x[1][sort_col_idx - 1] is None, x[1][sort_col_idx - 1]),
                    reverse=reverse
                )
            except (IndexError, TypeError) as e:
                raise ValueError(
                    f"Error sorting by column {column}: {str(e)}. "
                    f"Make sure the column exists and contains sortable data."
                ) from e
            
            # Write sorted data back to sheet
            for i, (original_row_idx, row_data) in enumerate(sorted_data):
                current_row = start_row + i
                for col_idx, cell_value in enumerate(row_data, start=1):
                    worksheet.cell(row=current_row, column=col_idx, value=cell_value)
            
            workbook.save(file_path)
            
            logger.info(
                f"Successfully sorted {len(sorted_data)} rows by column {column} "
                f"in {order}ending order"
            )
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error sorting data: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "sort_data")
            raise IOError(f"Failed to sort data: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()

    def _find_chart_at_position(self, worksheet, position: str):
        """
        Encontra um gráfico em uma posição específica da worksheet.
        
        Args:
            worksheet: Worksheet do openpyxl
            position: Posição da célula (e.g., 'H2')
        
        Returns:
            Chart object se encontrado, None caso contrário
        """
        # Converter posição para coordenadas (col, row)
        from openpyxl.utils import coordinate_to_tuple
        
        try:
            target_row, target_col = coordinate_to_tuple(position)
        except Exception:
            # Se posição inválida, não pode verificar
            return None
        
        # Verificar cada gráfico na worksheet
        for chart in worksheet._charts:
            if hasattr(chart, 'anchor') and hasattr(chart.anchor, '_from'):
                anchor_from = chart.anchor._from
                chart_col = anchor_from.col
                chart_row = anchor_from.row
                
                # Comparar posições (row é 0-indexed no anchor, 1-indexed na posição)
                if chart_col == target_col - 1 and chart_row == target_row - 1:
                    return chart
        
        return None
    
    def _get_chart_title(self, chart) -> Optional[str]:
        """
        Extrai o título de um gráfico.
        
        Args:
            chart: Chart object do openpyxl
        
        Returns:
            Título do gráfico ou None se não tiver título
        """
        try:
            if chart.title and chart.title.tx and chart.title.tx.rich:
                paragraphs = chart.title.tx.rich.p
                if paragraphs and len(paragraphs) > 0:
                    runs = paragraphs[0].r
                    if runs and len(runs) > 0:
                        return runs[0].t
        except Exception:
            pass
        
        return None

    def list_charts(self, file_path: str, sheet: Optional[str] = None) -> Dict[str, Any]:
        """
        Lista todos os gráficos em uma planilha Excel.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of specific sheet to list charts from (optional)
                   If None, lists charts from all sheets
        
        Returns:
            Dictionary containing:
                - charts: List of chart information dictionaries
                - total_count: Total number of charts found
                - sheets_analyzed: List of sheet names analyzed
            
            Each chart dictionary contains:
                - sheet: Sheet name where chart is located
                - title: Chart title (or "Untitled" if no title)
                - type: Chart type (BarChart, LineChart, PieChart, etc.)
                - position: Cell position (e.g., 'H2')
                - index: Index of chart in the sheet (0-based)
        
        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If specified sheet does not exist
            CorruptedFileError: If the file is corrupted
        
        Example:
            >>> tool = ExcelTool()
            >>> result = tool.list_charts('vendas.xlsx', 'Sheet1')
            >>> print(f"Found {result['total_count']} charts")
            >>> for chart in result['charts']:
            ...     print(f"  - {chart['title']} at {chart['position']}")
        """
        logger.info(f"Listing charts in {file_path}" + (f" (sheet: {sheet})" if sheet else " (all sheets)"))
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            # Determine which sheets to analyze
            if sheet:
                if sheet not in workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet}' not found in workbook")
                sheets_to_analyze = [sheet]
            else:
                sheets_to_analyze = workbook.sheetnames
            
            # Collect chart information
            charts_info = []
            
            for sheet_name in sheets_to_analyze:
                worksheet = workbook[sheet_name]
                
                for index, chart in enumerate(worksheet._charts):
                    # Extract chart information
                    chart_info = {
                        'sheet': sheet_name,
                        'title': self._get_chart_title(chart) or "Untitled",
                        'type': type(chart).__name__,
                        'position': self._get_chart_position(chart),
                        'index': index
                    }
                    
                    charts_info.append(chart_info)
            
            result = {
                'charts': charts_info,
                'total_count': len(charts_info),
                'sheets_analyzed': sheets_to_analyze
            }
            
            logger.info(
                f"Found {len(charts_info)} chart(s) in "
                f"{len(sheets_to_analyze)} sheet(s)"
            )
            
            return result
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error listing charts: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "list_charts")
            raise IOError(f"Failed to list charts: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()
    
    def _get_chart_position(self, chart) -> str:
        """
        Extrai a posição de um gráfico.
        
        Args:
            chart: Chart object do openpyxl
        
        Returns:
            Posição do gráfico (e.g., 'H2') ou 'Unknown' se não puder determinar
        """
        try:
            if hasattr(chart, 'anchor') and hasattr(chart.anchor, '_from'):
                anchor_from = chart.anchor._from
                col = anchor_from.col
                row = anchor_from.row
                
                # Converter coluna numérica para letra
                col_letter = get_column_letter(col + 1)  # col é 0-indexed
                position = f"{col_letter}{row + 1}"  # row é 0-indexed
                
                return position
        except Exception:
            pass
        
        return "Unknown"
    
    def delete_chart(self, file_path: str, sheet: str, identifier: Any) -> None:
        """
        Remove um gráfico de uma planilha Excel.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet containing the chart. If None and identifier="all",
                   deletes all charts from all sheets.
            identifier: Chart identifier - can be:
                       - "all": Delete all charts in the sheet (or entire workbook if sheet=None)
                       - int: Index of chart in the sheet (0-based)
                       - str: Title of the chart to delete
        
        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If the sheet doesn't exist or chart not found
            CorruptedFileError: If the file is corrupted
            IOError: If the chart cannot be deleted
        
        Example:
            >>> tool = ExcelTool()
            >>> # Delete all charts in a sheet
            >>> tool.delete_chart('vendas.xlsx', 'Sheet1', 'all')
            >>> # Delete all charts in entire workbook
            >>> tool.delete_chart('vendas.xlsx', None, 'all')
            >>> # Delete by index
            >>> tool.delete_chart('vendas.xlsx', 'Sheet1', 0)
            >>> # Delete by title
            >>> tool.delete_chart('vendas.xlsx', 'Sheet1', 'Vendas por Produto')
        """
        logger.info(f"Deleting chart '{identifier}' from sheet '{sheet}' in {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            # Delete ALL charts (all sheets if sheet not specified, else just the given sheet)
            if identifier == "all":
                if sheet is not None and sheet not in workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet}' not found in workbook")
                sheets_to_clear = workbook.sheetnames if sheet is None else [sheet]
                total_deleted = 0
                for sheet_name in sheets_to_clear:
                    ws = workbook[sheet_name]
                    total_deleted += len(ws._charts)
                    ws._charts.clear()
                if total_deleted == 0:
                    target = "workbook" if sheet is None else f"sheet '{sheet}'"
                    raise ValueError(f"No charts found in {target}")
                workbook.save(file_path)
                logger.info(f"Successfully deleted all {total_deleted} chart(s) from {len(sheets_to_clear)} sheet(s)")
                return

            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            
            # Find chart to delete
            chart_to_delete = None
            chart_index = None
            
            if isinstance(identifier, int):
                # Delete by index
                if identifier < 0 or identifier >= len(worksheet._charts):
                    raise ValueError(
                        f"Chart index {identifier} out of range. "
                        f"Sheet has {len(worksheet._charts)} chart(s) (indices 0-{len(worksheet._charts)-1})"
                    )
                chart_to_delete = worksheet._charts[identifier]
                chart_index = identifier
                logger.debug(f"Found chart at index {identifier}")
                
            elif isinstance(identifier, str):
                # Delete by title
                for index, chart in enumerate(worksheet._charts):
                    chart_title = self._get_chart_title(chart)
                    if chart_title and chart_title == identifier:
                        chart_to_delete = chart
                        chart_index = index
                        logger.debug(f"Found chart with title '{identifier}' at index {index}")
                        break
                
                if chart_to_delete is None:
                    # List available charts for better error message
                    available_titles = []
                    for chart in worksheet._charts:
                        title = self._get_chart_title(chart)
                        available_titles.append(title if title else "Untitled")
                    
                    raise ValueError(
                        f"Chart with title '{identifier}' not found in sheet '{sheet}'. "
                        f"Available charts: {', '.join(available_titles) if available_titles else 'none'}"
                    )
            else:
                raise ValueError(
                    f"Invalid identifier type: {type(identifier).__name__}. "
                    f"Must be int (index) or str (title)"
                )
            
            # Get chart info for logging
            chart_title = self._get_chart_title(chart_to_delete)
            chart_position = self._get_chart_position(chart_to_delete)
            
            # Remove chart
            worksheet._charts.remove(chart_to_delete)
            
            # Save workbook
            workbook.save(file_path)
            
            logger.info(
                f"Successfully deleted chart "
                f"'{chart_title if chart_title else 'Untitled'}' "
                f"(index {chart_index}) at position {chart_position}"
            )
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error deleting chart: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "delete_chart")
            raise IOError(f"Failed to delete chart: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()


    def remove_duplicates(self, file_path: str, sheet: str, config: Dict[str, Any]) -> Dict[str, Any]:
        """
        Remove linhas duplicadas de uma planilha Excel.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet to process
            config: Configuration dictionary with:
                - column: Optional[Union[str, int]] - Column to check for duplicates
                         (letter like 'A' or number like 1). If None, checks entire row.
                - has_header: bool - Whether first row is header (default: True)
                - keep: str - Which occurrence to keep: 'first' or 'last' (default: 'first')
        
        Returns:
            Dictionary with:
                - removed_count: Number of duplicate rows removed
                - remaining_count: Number of rows remaining
                - duplicates_found: List of duplicate values found
        
        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If parameters are invalid
            CorruptedFileError: If the file is corrupted
            IOError: If operation fails
        
        Example:
            >>> tool = ExcelTool()
            >>> # Remove duplicates by email column
            >>> result = tool.remove_duplicates('clientes.xlsx', 'Sheet1', 
            ...                                  {'column': 'C', 'has_header': True, 'keep': 'first'})
            >>> print(f"Removed {result['removed_count']} duplicates")
            >>> 
            >>> # Remove duplicate rows (entire row comparison)
            >>> result = tool.remove_duplicates('dados.xlsx', 'Sheet1', 
            ...                                  {'has_header': True, 'keep': 'first'})
        """
        logger.info(f"Removing duplicates from sheet '{sheet}' in {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Parse configuration
        column = config.get('column')
        has_header = config.get('has_header', True)
        keep = config.get('keep', 'first')
        
        # Validate keep parameter
        if keep not in ['first', 'last']:
            raise ValueError(f"Invalid 'keep' value: '{keep}'. Must be 'first' or 'last'")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            
            # Get all data
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Check if sheet is truly empty (openpyxl returns max_row=1 for empty sheets)
            if max_row == 0 or (max_row == 1 and worksheet.cell(1, 1).value is None):
                logger.warning(f"Sheet '{sheet}' is empty")
                return {
                    'removed_count': 0,
                    'remaining_count': 0,
                    'duplicates_found': []
                }
            
            # Determine start row (skip header if present)
            start_row = 2 if has_header else 1
            
            if start_row > max_row:
                logger.warning(f"Sheet '{sheet}' has only header row")
                return {
                    'removed_count': 0,
                    'remaining_count': 1 if has_header else 0,
                    'duplicates_found': []
                }
            
            # Convert column to index if specified
            col_idx = None
            if column is not None:
                if isinstance(column, str):
                    # Convert letter to number (A=1, B=2, etc.)
                    col_idx = 0
                    for char in column.upper():
                        col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
                elif isinstance(column, int):
                    col_idx = column
                else:
                    raise ValueError(f"Invalid column type: {type(column).__name__}. Must be str or int")
                
                if col_idx < 1 or col_idx > max_col:
                    raise ValueError(f"Column {column} out of range (1-{max_col})")
            
            # Read all data rows
            rows_data = []
            for row_idx in range(start_row, max_row + 1):
                row_values = []
                for col_idx_iter in range(1, max_col + 1):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx_iter).value
                    row_values.append(cell_value)
                rows_data.append((row_idx, tuple(row_values)))
            
            # Track seen values and rows to keep
            seen = set()
            rows_to_keep = []
            rows_to_delete = []
            duplicates_found = []
            
            # Process rows based on keep strategy
            if keep == 'first':
                # Keep first occurrence
                for row_idx, row_values in rows_data:
                    # Determine comparison key
                    if col_idx is not None:
                        # Compare specific column
                        key = row_values[col_idx - 1]
                    else:
                        # Compare entire row
                        key = row_values
                    
                    if key not in seen:
                        seen.add(key)
                        rows_to_keep.append(row_idx)
                    else:
                        rows_to_delete.append(row_idx)
                        if key not in duplicates_found:
                            duplicates_found.append(key)
            else:
                # Keep last occurrence - process in reverse
                for row_idx, row_values in reversed(rows_data):
                    # Determine comparison key
                    if col_idx is not None:
                        # Compare specific column
                        key = row_values[col_idx - 1]
                    else:
                        # Compare entire row
                        key = row_values
                    
                    if key not in seen:
                        seen.add(key)
                        rows_to_keep.append(row_idx)
                    else:
                        rows_to_delete.append(row_idx)
                        if key not in duplicates_found:
                            duplicates_found.append(key)
            
            # Delete duplicate rows (in reverse order to maintain indices)
            rows_to_delete.sort(reverse=True)
            for row_idx in rows_to_delete:
                worksheet.delete_rows(row_idx, 1)
            
            # Save workbook
            workbook.save(file_path)
            
            removed_count = len(rows_to_delete)
            remaining_count = max_row - removed_count
            
            # Convert duplicates to strings for logging
            duplicates_str = []
            for dup in duplicates_found[:10]:  # Limit to first 10
                if isinstance(dup, tuple):
                    duplicates_str.append(str(dup))
                else:
                    duplicates_str.append(str(dup))
            
            logger.info(
                f"Removed {removed_count} duplicate row(s) from sheet '{sheet}'. "
                f"Remaining: {remaining_count} row(s)"
            )
            
            if duplicates_str:
                logger.debug(f"Sample duplicates found: {', '.join(duplicates_str[:5])}")
            
            return {
                'removed_count': removed_count,
                'remaining_count': remaining_count,
                'duplicates_found': duplicates_str
            }
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error removing duplicates: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "remove_duplicates")
            raise IOError(f"Failed to remove duplicates: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()


    def filter_and_copy(self, file_path: str, source_sheet: str, config: Dict[str, Any]) -> Dict[str, Any]:
        """
        Filtra dados de uma planilha e copia para nova sheet ou arquivo.
        
        Args:
            file_path: Path to the Excel file
            source_sheet: Name of the source sheet to filter
            config: Configuration dictionary with:
                - column: Union[str, int] - Column to filter (letter like 'A' or number like 1)
                - operator: str - Comparison operator: '>', '<', '>=', '<=', '==', '!=', 'contains', 'starts_with', 'ends_with'
                - value: Any - Value to compare against
                - destination_sheet: Optional[str] - Name of destination sheet (creates new sheet in same file)
                - destination_file: Optional[str] - Path to destination file (creates new file)
                - has_header: bool - Whether first row is header (default: True)
                - copy_header: bool - Whether to copy header to destination (default: True if has_header)
        
        Returns:
            Dictionary with:
                - filtered_count: Number of rows that matched the filter
                - destination: Where data was copied to
                - destination_type: 'sheet' or 'file'
        
        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If parameters are invalid
            CorruptedFileError: If the file is corrupted
            IOError: If operation fails
        
        Example:
            >>> tool = ExcelTool()
            >>> # Filter sales > 1000 to new sheet
            >>> result = tool.filter_and_copy('vendas.xlsx', 'Sheet1',
            ...     {'column': 'F', 'operator': '>', 'value': 1000,
            ...      'destination_sheet': 'Vendas Altas', 'has_header': True})
            >>> 
            >>> # Filter by text contains
            >>> result = tool.filter_and_copy('clientes.xlsx', 'Sheet1',
            ...     {'column': 'C', 'operator': 'contains', 'value': '@gmail.com',
            ...      'destination_file': 'clientes_gmail.xlsx', 'has_header': True})
        """
        logger.info(f"Filtering and copying data from sheet '{source_sheet}' in {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Parse configuration
        column = config.get('column')
        operator = config.get('operator')
        value = config.get('value')
        destination_sheet = config.get('destination_sheet')
        destination_file = config.get('destination_file')
        has_header = config.get('has_header', True)
        copy_header = config.get('copy_header', has_header)
        
        # Validate required parameters
        if column is None:
            raise ValueError("'column' parameter is required")
        if operator is None:
            raise ValueError("'operator' parameter is required")
        if value is None:
            raise ValueError("'value' parameter is required")
        
        # Validate operator
        valid_operators = ['>', '<', '>=', '<=', '==', '!=', 'contains', 'starts_with', 'ends_with']
        if operator not in valid_operators:
            raise ValueError(
                f"Invalid operator '{operator}'. "
                f"Must be one of: {', '.join(valid_operators)}"
            )
        
        # Validate destination
        if not destination_sheet and not destination_file:
            raise ValueError("Either 'destination_sheet' or 'destination_file' must be specified")
        if destination_sheet and destination_file:
            raise ValueError("Cannot specify both 'destination_sheet' and 'destination_file'")
        
        workbook = None
        dest_workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if source_sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{source_sheet}' not found in workbook")
            
            worksheet = workbook[source_sheet]
            
            # Get all data
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Check if sheet is empty
            if max_row == 0 or (max_row == 1 and worksheet.cell(1, 1).value is None):
                logger.warning(f"Sheet '{source_sheet}' is empty")
                return {
                    'filtered_count': 0,
                    'destination': destination_sheet or destination_file,
                    'destination_type': 'sheet' if destination_sheet else 'file'
                }
            
            # Convert column to index
            if isinstance(column, str):
                # Convert letter to number (A=1, B=2, etc.)
                col_idx = 0
                for char in column.upper():
                    col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
            elif isinstance(column, int):
                col_idx = column
            else:
                raise ValueError(f"Invalid column type: {type(column).__name__}. Must be str or int")
            
            if col_idx < 1 or col_idx > max_col:
                raise ValueError(f"Column {column} out of range (1-{max_col})")
            
            # Determine start row
            start_row = 2 if has_header else 1
            
            # Read header if present
            header_row = None
            if has_header:
                header_row = []
                for col_idx_iter in range(1, max_col + 1):
                    header_row.append(worksheet.cell(row=1, column=col_idx_iter).value)
            
            # Filter rows
            filtered_rows = []
            for row_idx in range(start_row, max_row + 1):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                
                # Apply filter
                matches = self._apply_filter(cell_value, operator, value)
                
                if matches:
                    # Copy entire row
                    row_data = []
                    for col_idx_iter in range(1, max_col + 1):
                        row_data.append(worksheet.cell(row=row_idx, column=col_idx_iter).value)
                    filtered_rows.append(row_data)
            
            # Create destination
            if destination_sheet:
                # Create new sheet in same file
                if destination_sheet in workbook.sheetnames:
                    # Remove existing sheet
                    del workbook[destination_sheet]
                
                dest_ws = workbook.create_sheet(destination_sheet)
                
                # Copy header if needed
                row_offset = 1
                if copy_header and header_row:
                    for col_idx_iter, value in enumerate(header_row, start=1):
                        dest_ws.cell(row=1, column=col_idx_iter, value=value)
                    row_offset = 2
                
                # Copy filtered rows
                for row_idx, row_data in enumerate(filtered_rows, start=row_offset):
                    for col_idx_iter, value in enumerate(row_data, start=1):
                        dest_ws.cell(row=row_idx, column=col_idx_iter, value=value)
                
                # Save workbook
                workbook.save(file_path)
                
                logger.info(
                    f"Filtered {len(filtered_rows)} row(s) to sheet '{destination_sheet}'"
                )
                
                return {
                    'filtered_count': len(filtered_rows),
                    'destination': destination_sheet,
                    'destination_type': 'sheet'
                }
            
            else:
                # Create new file
                dest_workbook = Workbook()
                dest_ws = dest_workbook.active
                dest_ws.title = source_sheet
                
                # Copy header if needed
                row_offset = 1
                if copy_header and header_row:
                    for col_idx_iter, value in enumerate(header_row, start=1):
                        dest_ws.cell(row=1, column=col_idx_iter, value=value)
                    row_offset = 2
                
                # Copy filtered rows
                for row_idx, row_data in enumerate(filtered_rows, start=row_offset):
                    for col_idx_iter, value in enumerate(row_data, start=1):
                        dest_ws.cell(row=row_idx, column=col_idx_iter, value=value)
                
                # Save new file
                dest_workbook.save(destination_file)
                
                logger.info(
                    f"Filtered {len(filtered_rows)} row(s) to file '{destination_file}'"
                )
                
                return {
                    'filtered_count': len(filtered_rows),
                    'destination': destination_file,
                    'destination_type': 'file'
                }
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error filtering and copying: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "filter_and_copy")
            raise IOError(f"Failed to filter and copy: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()
            if dest_workbook is not None:
                dest_workbook.close()
    
    def _apply_filter(self, cell_value: Any, operator: str, filter_value: Any) -> bool:
        """
        Apply filter operator to cell value.
        
        Args:
            cell_value: Value from the cell
            operator: Comparison operator
            filter_value: Value to compare against
        
        Returns:
            True if cell_value matches the filter, False otherwise
        """
        # Handle None values
        if cell_value is None:
            return operator == '==' and filter_value is None
        
        try:
            # Numeric comparisons
            if operator in ['>', '<', '>=', '<=']:
                # Try to convert to numbers
                try:
                    cell_num = float(cell_value)
                    filter_num = float(filter_value)
                    
                    if operator == '>':
                        return cell_num > filter_num
                    elif operator == '<':
                        return cell_num < filter_num
                    elif operator == '>=':
                        return cell_num >= filter_num
                    elif operator == '<=':
                        return cell_num <= filter_num
                except (ValueError, TypeError):
                    return False
            
            # Equality comparisons
            elif operator == '==':
                return cell_value == filter_value
            elif operator == '!=':
                return cell_value != filter_value
            
            # String comparisons
            elif operator in ['contains', 'starts_with', 'ends_with']:
                cell_str = str(cell_value).lower()
                filter_str = str(filter_value).lower()
                
                if operator == 'contains':
                    return filter_str in cell_str
                elif operator == 'starts_with':
                    return cell_str.startswith(filter_str)
                elif operator == 'ends_with':
                    return cell_str.endswith(filter_str)
            
            return False
            
        except Exception as e:
            logger.debug(f"Error applying filter: {e}")
            return False


    def insert_rows(self, file_path: str, sheet: str, start_row: int, count: int = 1) -> None:
        """
        Insere linhas vazias em uma sheet.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet
            start_row: Position where to insert rows (1-based index)
            count: Number of rows to insert (default: 1)
            
        Raises:
            FileNotFoundError: If the file does not exist
            CorruptedFileError: If the file is corrupted
            ValueError: If the sheet name is invalid or row index is out of range
            IOError: If operation fails
        
        Example:
            >>> tool = ExcelTool()
            >>> # Insert 3 empty rows at position 5
            >>> tool.insert_rows('data.xlsx', 'Sheet1', 5, 3)
        """
        logger.info(f"Inserting {count} row(s) at position {start_row} in sheet '{sheet}' of {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        if start_row < 1:
            raise ValueError(f"start_row must be >= 1, got {start_row}")
        if count < 1:
            raise ValueError(f"count must be >= 1, got {count}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                logger.error(f"Sheet '{sheet}' not found in {file_path}")
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            
            # Insert rows (openpyxl shifts existing rows down)
            worksheet.insert_rows(start_row, count)
            
            workbook.save(file_path)
            
            logger.info(f"Successfully inserted {count} row(s) at position {start_row}")
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error inserting rows: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "insert_rows")
            raise IOError(f"Failed to insert rows: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()
    
    def insert_columns(self, file_path: str, sheet: str, start_col: int, count: int = 1) -> None:
        """
        Insere colunas vazias em uma sheet.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet
            start_col: Position where to insert columns (1-based index, or letter like 'A')
            count: Number of columns to insert (default: 1)
            
        Raises:
            FileNotFoundError: If the file does not exist
            CorruptedFileError: If the file is corrupted
            ValueError: If the sheet name is invalid or column index is out of range
            IOError: If operation fails
        
        Example:
            >>> tool = ExcelTool()
            >>> # Insert 2 empty columns at position C (3)
            >>> tool.insert_columns('data.xlsx', 'Sheet1', 'C', 2)
            >>> # Or using number
            >>> tool.insert_columns('data.xlsx', 'Sheet1', 3, 2)
        """
        logger.info(f"Inserting {count} column(s) at position {start_col} in sheet '{sheet}' of {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Convert column to index if it's a letter
        if isinstance(start_col, str):
            col_idx = 0
            for char in start_col.upper():
                col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
            start_col = col_idx
        elif not isinstance(start_col, int):
            raise ValueError(f"Invalid column type: {type(start_col).__name__}. Must be str or int")
        
        if start_col < 1:
            raise ValueError(f"start_col must be >= 1, got {start_col}")
        if count < 1:
            raise ValueError(f"count must be >= 1, got {count}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                logger.error(f"Sheet '{sheet}' not found in {file_path}")
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            
            # Insert columns (openpyxl shifts existing columns right)
            worksheet.insert_cols(start_col, count)
            
            workbook.save(file_path)
            
            col_letter = get_column_letter(start_col)
            logger.info(f"Successfully inserted {count} column(s) at position {col_letter} ({start_col})")
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error inserting columns: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "insert_columns")
            raise IOError(f"Failed to insert columns: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()


    def freeze_panes(self, file_path: str, sheet: str, row: Optional[int] = None, col: Optional[int] = None) -> None:
        """
        Congela painéis (linhas e/ou colunas) em uma sheet.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet
            row: Row number where to freeze (freezes rows above this row, 1-based)
                 Example: row=2 freezes row 1 (header)
            col: Column number or letter where to freeze (freezes columns left of this column)
                 Example: col='B' or col=2 freezes column A
            
        Note:
            - At least one of row or col must be specified
            - row=2, col=None: Freezes only row 1 (horizontal freeze)
            - row=None, col='B': Freezes only column A (vertical freeze)
            - row=2, col='B': Freezes row 1 and column A (both)
            
        Raises:
            FileNotFoundError: If the file does not exist
            CorruptedFileError: If the file is corrupted
            ValueError: If parameters are invalid
            IOError: If operation fails
        
        Example:
            >>> tool = ExcelTool()
            >>> # Freeze first row (header)
            >>> tool.freeze_panes('data.xlsx', 'Sheet1', row=2)
            >>> # Freeze first column
            >>> tool.freeze_panes('data.xlsx', 'Sheet1', col='B')
            >>> # Freeze both row 1 and column A
            >>> tool.freeze_panes('data.xlsx', 'Sheet1', row=2, col='B')
        """
        logger.info(f"Freezing panes in sheet '{sheet}' of {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Validate that at least one parameter is specified
        if row is None and col is None:
            raise ValueError("At least one of 'row' or 'col' must be specified")
        
        # Convert column to index if it's a letter
        col_idx = 1  # Default to column A (1)
        if col is not None:
            if isinstance(col, str):
                col_idx = 0
                for char in col.upper():
                    col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
            elif isinstance(col, int):
                col_idx = col
            else:
                raise ValueError(f"Invalid column type: {type(col).__name__}. Must be str or int")
            
            if col_idx < 1:
                raise ValueError(f"col must be >= 1, got {col_idx}")
        
        # Validate row
        row_idx = 1  # Default to row 1
        if row is not None:
            if not isinstance(row, int):
                raise ValueError(f"Invalid row type: {type(row).__name__}. Must be int")
            if row < 1:
                raise ValueError(f"row must be >= 1, got {row}")
            row_idx = row
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                logger.error(f"Sheet '{sheet}' not found in {file_path}")
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            
            # Determine freeze cell
            if row is not None and col is not None:
                # Freeze both rows and columns
                freeze_cell = f"{get_column_letter(col_idx)}{row_idx}"
            elif row is not None:
                # Freeze only rows (column A)
                freeze_cell = f"A{row_idx}"
            else:
                # Freeze only columns (row 1)
                freeze_cell = f"{get_column_letter(col_idx)}1"
            
            # Set freeze panes
            worksheet.freeze_panes = freeze_cell
            
            workbook.save(file_path)
            
            logger.info(f"Successfully froze panes at cell {freeze_cell}")
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error freezing panes: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "freeze_panes")
            raise IOError(f"Failed to freeze panes: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()
    
    def unfreeze_panes(self, file_path: str, sheet: str) -> None:
        """
        Remove o congelamento de painéis de uma sheet.
        
        Args:
            file_path: Path to the Excel file
            sheet: Name of the sheet
            
        Raises:
            FileNotFoundError: If the file does not exist
            CorruptedFileError: If the file is corrupted
            ValueError: If the sheet name is invalid
            IOError: If operation fails
        
        Example:
            >>> tool = ExcelTool()
            >>> tool.unfreeze_panes('data.xlsx', 'Sheet1')
        """
        logger.info(f"Unfreezing panes in sheet '{sheet}' of {file_path}")
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            if sheet not in workbook.sheetnames:
                logger.error(f"Sheet '{sheet}' not found in {file_path}")
                raise ValueError(f"Sheet '{sheet}' not found in workbook")
            
            worksheet = workbook[sheet]
            
            # Remove freeze panes
            worksheet.freeze_panes = None
            
            workbook.save(file_path)
            
            logger.info(f"Successfully unfroze panes")
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error unfreezing panes: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "unfreeze_panes")
            raise IOError(f"Failed to unfreeze panes: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()


    def find_and_replace(
        self,
        file_path: str,
        find_text: str,
        replace_text: str,
        sheet: Optional[str] = None,
        match_case: bool = False,
        match_entire_cell: bool = False
    ) -> Dict[str, Any]:
        """
        Busca e substitui texto em uma planilha Excel.
        
        Args:
            file_path: Path to the Excel file
            find_text: Text to find
            replace_text: Text to replace with
            sheet: Name of specific sheet (optional, if None applies to all sheets)
            match_case: Whether to match case (default: False)
            match_entire_cell: Whether to match entire cell content (default: False)
        
        Returns:
            Dictionary with:
                - replacements_count: Total number of replacements made
                - sheets_processed: List of sheet names processed
                - details: List of replacement details per sheet
        
        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If parameters are invalid
            CorruptedFileError: If the file is corrupted
            IOError: If operation fails
        
        Example:
            >>> tool = ExcelTool()
            >>> # Replace in all sheets (case-insensitive)
            >>> result = tool.find_and_replace('vendas.xlsx', '2024', '2025')
            >>> print(f"Made {result['replacements_count']} replacements")
            >>> 
            >>> # Replace in specific sheet (case-sensitive, entire cell)
            >>> result = tool.find_and_replace(
            ...     'vendas.xlsx', 'Produto A', 'Produto Alpha',
            ...     sheet='Sheet1', match_case=True, match_entire_cell=True
            ... )
        """
        logger.info(
            f"Finding and replacing '{find_text}' with '{replace_text}' in {file_path}"
            + (f" (sheet: {sheet})" if sheet else " (all sheets)")
        )
        
        path = Path(file_path)
        if not path.exists():
            log_file_access_error(logger, file_path, "arquivo não encontrado")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        if not find_text:
            raise ValueError("find_text cannot be empty")
        
        workbook = None
        try:
            workbook = load_workbook(filename=file_path)
            
            # Determine which sheets to process
            if sheet:
                if sheet not in workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet}' not found in workbook")
                sheets_to_process = [sheet]
            else:
                sheets_to_process = workbook.sheetnames
            
            # Track replacements
            total_replacements = 0
            details = []
            
            # Process each sheet
            for sheet_name in sheets_to_process:
                worksheet = workbook[sheet_name]
                sheet_replacements = 0
                
                # Iterate through all cells
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        
                        # Convert cell value to string
                        cell_str = str(cell.value)
                        
                        # Prepare search text
                        search_text = cell_str
                        find_str = find_text
                        
                        if not match_case:
                            search_text = search_text.lower()
                            find_str = find_str.lower()
                        
                        # Check if text is found
                        if match_entire_cell:
                            # Match entire cell content
                            if search_text == find_str:
                                cell.value = replace_text
                                sheet_replacements += 1
                        else:
                            # Match substring
                            if find_str in search_text:
                                # Replace (preserving case if match_case=False)
                                if match_case:
                                    new_value = cell_str.replace(find_text, replace_text)
                                else:
                                    # Case-insensitive replacement
                                    import re
                                    pattern = re.compile(re.escape(find_text), re.IGNORECASE)
                                    new_value = pattern.sub(replace_text, cell_str)
                                
                                cell.value = new_value
                                sheet_replacements += 1
                
                total_replacements += sheet_replacements
                
                details.append({
                    'sheet': sheet_name,
                    'replacements': sheet_replacements
                })
                
                logger.debug(f"Made {sheet_replacements} replacement(s) in sheet '{sheet_name}'")
            
            # Save workbook
            workbook.save(file_path)
            
            logger.info(
                f"Successfully made {total_replacements} replacement(s) "
                f"across {len(sheets_to_process)} sheet(s)"
            )
            
            return {
                'replacements_count': total_replacements,
                'sheets_processed': sheets_to_process,
                'details': details
            }
            
        except InvalidFileException as e:
            logger.error(f"Corrupted Excel file: {file_path} - {str(e)}")
            raise CorruptedFileError(f"Excel file is corrupted or invalid: {file_path}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Error in find and replace: {file_path} - {str(e)}")
            log_error_with_traceback(logger, e, "find_and_replace")
            raise IOError(f"Failed to find and replace: {file_path} - {str(e)}") from e
        finally:
            if workbook is not None:
                workbook.close()
