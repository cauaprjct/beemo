"""Agent principal para coordenação de workflow do Gemini Office Agent.

Este módulo contém a classe Agent que coordena o fluxo de trabalho entre
entrada do usuário, descoberta de arquivos, Gemini API e ferramentas de manipulação.
"""

import json
import logging
from pathlib import Path
from typing import List, Dict, Any

from config import Config
from src.file_scanner import FileScanner
from src.gemini_client import GeminiClient
from src.excel_tool import ExcelTool
from src.word_tool import WordTool
from src.powerpoint_tool import PowerPointTool
from src.pdf_tool import PdfTool
from src.prompt_templates import PromptTemplates
from src.response_parser import ResponseParser
from src.security_validator import SecurityValidator
from src.models import AgentResponse, ActionResult, BatchResult
from src.logging_config import get_logger, log_operation_start, log_operation_end, log_error_with_traceback
from src.exceptions import ValidationError
from src.html_to_elements import contains_html, html_to_elements

logger = get_logger(__name__)


class Agent:
    """Componente central que coordena o workflow entre componentes.
    
    O Agent é responsável por:
    - Identificar arquivos relevantes usando FileScanner
    - Ler conteúdo de arquivos usando ferramentas apropriadas
    - Construir prompts contextualizados para Gemini
    - Enviar prompts para GeminiClient
    - Interpretar respostas do Gemini e determinar ações
    - Executar ações usando ferramentas apropriadas
    - Retornar resultados de operações
    
    Attributes:
        config: Configuração do sistema
        file_scanner: Scanner para descoberta de arquivos
        gemini_client: Cliente para comunicação com Gemini API
        excel_tool: Ferramenta para manipulação de Excel
        word_tool: Ferramenta para manipulação de Word
        powerpoint_tool: Ferramenta para manipulação de PowerPoint
        security_validator: Validador de segurança para operações
    """
    
    def __init__(self, config: Config, version_manager=None, response_cache=None):
        """Inicializa o Agent com configuração.
        
        Args:
            config: Objeto Config com configurações do sistema
            version_manager: Gerenciador de versões (opcional)
            response_cache: Cache de respostas (opcional)
        """
        self.config = config
        self.version_manager = version_manager
        self.response_cache = response_cache
        
        # Inicializa componentes
        self.file_scanner = FileScanner(config.root_path)
        self.gemini_client = GeminiClient(
            config.api_key, config.model_name, response_cache,
            fallback_models=config.fallback_models
        )
        self.excel_tool = ExcelTool()
        self.word_tool = WordTool(gemini_client=self.gemini_client)
        self.powerpoint_tool = PowerPointTool()
        self.pdf_tool = PdfTool()
        self.security_validator = SecurityValidator(config.root_path)
        self._last_read_result = None
        
        logger.info("Agent inicializado com sucesso")

    
    def process_user_request(self, user_prompt: str, selected_files: List[str] = None) -> AgentResponse:
        """Processa solicitação do usuário e retorna resultado.
        
        Este é o método principal que coordena todo o workflow:
        1. Descobre arquivos disponíveis (ou usa selected_files se fornecido)
        2. Filtra arquivos relevantes baseado no prompt
        3. Lê conteúdo dos arquivos relevantes
        4. Constrói prompt contextualizado
        5. Envia para Gemini API
        6. Interpreta resposta e executa ações
        7. Retorna resultado
        
        Args:
            user_prompt: Descrição em linguagem natural da operação desejada
            selected_files: Lista opcional de arquivos pré-selecionados pelo usuário.
                            Se fornecida, ignora a descoberta automática.
        
        Returns:
            AgentResponse com status, mensagem e arquivos modificados
        
        Validates: Requirements 7.1, 7.2, 7.3, 7.4, 7.5, 7.6, 7.7, 7.8, 9.2
        """
        logger.info("=== Iniciando processamento de solicitação do usuário ===")
        logger.info(f"Prompt do usuário: {user_prompt}")
        log_operation_start(logger, "process_user_request")
        
        try:
            # 1. Usar arquivos selecionados ou descobrir automaticamente
            if selected_files is not None:
                available_files = selected_files
                logger.info(f"Usando {len(available_files)} arquivo(s) pré-selecionado(s)")
            else:
                available_files = self._discover_files()
                logger.info(f"Arquivos disponíveis descobertos: {len(available_files)}")
            
            # 2. Filtrar arquivos relevantes baseado no prompt
            relevant_files = self._filter_relevant_files(available_files, user_prompt)
            logger.info(f"Arquivos relevantes filtrados: {len(relevant_files)}")
            
            # 3. Ler conteúdo dos arquivos relevantes
            file_contexts = []
            for file_path in relevant_files:
                try:
                    content = self._read_file_content(file_path)
                    file_type = self._get_file_type(file_path)
                    
                    # Formatar conteúdo para o contexto
                    formatted = PromptTemplates.format_file_content(
                        file_path, file_type, content
                    )
                    file_contexts.append(formatted)
                    
                except Exception as e:
                    logger.warning(f"Erro ao ler arquivo {file_path}: {e}")
                    # Continua com outros arquivos
            
            # 4. Construir prompt contextualizado
            context_prompt = self._build_context_prompt(user_prompt, file_contexts)
            logger.debug(f"Prompt contextualizado construído (tamanho: {len(context_prompt)} caracteres)")
            
            # 5. Enviar para Gemini API
            logger.info("Enviando prompt para Gemini API")
            gemini_response = self.gemini_client.generate_response(context_prompt)
            logger.info("Resposta recebida do Gemini API")
            
            # 6. Executar ações baseadas na resposta
            result = self._execute_actions(gemini_response)
            
            logger.info("=== Processamento concluído com sucesso ===")
            log_operation_end(logger, "process_user_request")
            return result
            
        except Exception as e:
            log_error_with_traceback(logger, e, "process_user_request")
            return AgentResponse(
                success=False,
                message="Erro ao processar solicitação",
                files_modified=[],
                error=str(e)
            )

    
    def process_chat_message(self, user_message: str, chat_history: List[Dict] = None,
                              selected_files: List[str] = None) -> AgentResponse:
        """Process a chat message using native Gemini function calling.

        The model natively decides whether to call tools (file operations)
        or respond with plain text (chat/questions). No JSON parsing required.

        Args:
            user_message: The user's current message
            chat_history: Previous conversation turns [{role, content}]
            selected_files: Optional pre-selected files to include in context

        Returns:
            AgentResponse with is_chat=True for conversational replies,
            or file operation result when tools were executed
        """
        from google.genai import types as genai_types
        from src.tool_definitions import build_office_tools

        from src.exceptions import QuotaExceededError as _QuotaExceededError
        logger.info(f"[FC] Processando mensagem: {user_message[:80]}")

        try:
            # 1. Discover available files
            if selected_files is not None:
                available_files = selected_files
            else:
                available_files = self._discover_files()

            # 1b. Shortcut for pure conversational messages — bypass tool API to save quota
            if self._is_chat_only_message(user_message) and not available_files:
                _simple_prompt = (
                    "Você é o Beemo, um assistente simpático para arquivos Office. "
                    "Responda de forma breve e amigável em português. "
                    "Responda APENAS com texto natural, SEM JSON.\n\n"
                    f"Usuário: {user_message}"
                )
                try:
                    _txt = self.gemini_client.generate_text_response(_simple_prompt)
                    _txt = self._strip_json_wrapper(_txt)
                    return AgentResponse(success=True, message=_txt, files_modified=[], is_chat=True)
                except _QuotaExceededError:
                    raise  # let outer handler return friendly message, no need to try tools
                except Exception:
                    pass  # other errors: fall through to normal tool-enabled flow

            # 2. Build system instruction listing available files
            file_list_str = "\n".join(
                f"- {Path(fp).name}" for fp in available_files
            ) or "(nenhum arquivo disponível)"

            single_file_hint = (
                f" O único arquivo disponível é '{Path(available_files[0]).name}' — use-o diretamente."
                if len(available_files) == 1 else ""
            )
            system_instruction = (
                "Você é o Beemo, um assistente simpático e inteligente para gerenciar "
                "arquivos Office (Excel, Word, PDF, PowerPoint).\n\n"
                f"Arquivos disponíveis no diretório atual:\n{file_list_str}\n\n"
                "Regras OBRIGATÓRIAS — siga sempre, sem exceções:\n"
                "- Para perguntas, saudações e conversas → responda em texto natural, SEM usar ferramentas.\n"
                "- Para listar arquivos → use list_files.\n"
                "- Para ler/analisar/resumir um arquivo → use read_file.\n"
                "- Para criar ou modificar um arquivo → use a ferramenta correspondente IMEDIATAMENTE, "
                "SEM fazer perguntas ao usuário.\n"
                f"- Se há apenas 1 arquivo disponível, use-o DIRETAMENTE sem perguntar qual arquivo usar.{single_file_hint}\n"
                "- NUNCA pergunte qual arquivo usar, qual planilha usar, quais colunas usar ou detalhes "
                "técnicos de implementação. Descubra essas informações lendo o arquivo com read_file.\n"
                "- Para gráficos: leia o arquivo primeiro (read_file), analise o conteúdo, "
                "escolha colunas e posição automaticamente e crie o gráfico sem perguntar nada.\n"
                "- Para criar planilhas com muitas linhas (50–1000+): use operation=create com "
                "row_count=N e columns='Col1,Col2,...' — o sistema gera os dados automaticamente "
                "em Python sem sobrecarregar o modelo. NÃO gere as linhas manualmente como JSON.\n"
                "- Para volumes muito grandes (>500 linhas) OU quando o usuário especificar "
                "regras de dados complexas: crie em etapas usando create (primeiras linhas) "
                "seguido de múltiplos append até completar o total. Informe o usuário que está "
                "fazendo em etapas e execute tudo sem perguntar.\n"
                "- Para posicionar gráficos, use estas palavras-chave no campo 'position' — "
                "o sistema resolve automaticamente para a célula correta:\n"
                "  beside_content → ao lado dos dados (padrão)\n"
                "  below_content  → abaixo dos dados\n"
                "  after_last_chart → após o último gráfico existente\n"
                "  right:N / left:N / down:N / up:N → N colunas/linhas relativo ao gráfico atual\n"
                "  Ex: 'mova uma casa para direita' → position='right:1'; "
                "'abaixo do conteúdo' → position='below_content'.\n"
                "  NÃO peça esclarecimentos — escolha a keyword correta e execute direto.\n"
                "- Se o contexto da conversa indicar um arquivo já mencionado, use esse arquivo.\n"
                "- Ao ler um arquivo, resuma o conteúdo de forma clara e amigável.\n"
                "- Responda sempre no mesmo idioma do usuário (normalmente português).\n"
                "- IMPORTANTE: Quando o usuário pedir múltiplas operações (criar + formatar + gráfico), "
                "tente executar TODAS as ferramentas necessárias em PARALELO na mesma resposta, "
                "em vez de uma por vez. Isso economiza chamadas à API e é mais eficiente.\n"
                "- Seja conciso e direto, mas amigável e com a personalidade curiosa do Beemo."
            )

            # 3. Build conversation contents from history + current message
            contents = []
            for turn in (chat_history or [])[-10:]:
                role = "user" if turn.get("role") == "user" else "model"
                contents.append(
                    genai_types.Content(
                        role=role,
                        parts=[genai_types.Part(text=str(turn.get("content", "")))],
                    )
                )
            contents.append(
                genai_types.Content(
                    role="user",
                    parts=[genai_types.Part(text=user_message)],
                )
            )

            # 4. Build tools
            tools = build_office_tools()

            # 5. Agentic loop — model calls tools until it produces a text answer
            _READ_ONLY_OPS = {
                'read', 'get_info', 'list_charts', 'analyze_document',
                'analyze_word_count', 'analyze_section_length',
                'get_document_statistics', 'analyze_tone', 'identify_jargon',
                'analyze_readability', 'check_term_consistency',
                'extract_tables', 'get_page_layout',
            }
            files_modified = []
            tool_errors = []

            # Limit agentic loop to 5 turns to avoid API quota exhaustion
            # Each turn = 1 API call, so 5 turns = max 5 calls per user message
            MAX_TURNS = 5
            
            for turn_idx in range(MAX_TURNS):
                response = self.gemini_client.generate_with_tools(
                    contents, tools, system_instruction
                )

                candidate = response.candidates[0]
                
                # Handle case where content is None (blocked by safety filters or API error)
                if candidate.content is None:
                    logger.error(f"[FC] Resposta bloqueada ou inválida no turn {turn_idx + 1}")
                    logger.error(f"[FC] Finish reason: {candidate.finish_reason}")
                    logger.error(f"[FC] Safety ratings: {candidate.safety_ratings if hasattr(candidate, 'safety_ratings') else 'N/A'}")
                    return AgentResponse(
                        success=False,
                        message=(
                            "❌ A API bloqueou a resposta. Isso pode acontecer se:\n"
                            "- O prompt foi considerado muito longo ou complexo\n"
                            "- Houve um problema temporário na API\n"
                            "- O conteúdo acionou filtros de segurança\n\n"
                            "Tente simplificar a solicitação ou dividir em etapas menores."
                        ),
                        files_modified=files_modified,
                        is_chat=True,
                        error=f"Content blocked: {candidate.finish_reason}"
                    )
                
                parts = candidate.content.parts or []

                # Detect function calls in parts
                fc_parts = [
                    p for p in parts
                    if getattr(p, 'function_call', None) and p.function_call
                ]

                if not fc_parts:
                    # No tool calls — collect response text robustly
                    try:
                        text = response.text or ""
                    except Exception:
                        text = " ".join(
                            p.text for p in parts if getattr(p, 'text', None)
                        ).strip()
                    # Strip accidental JSON wrappers the model may produce
                    text = self._strip_json_wrapper(text)
                    logger.info(f"[FC] Resposta em texto recebida (turn {turn_idx + 1})")
                    had_errors = bool(tool_errors) and not files_modified
                    return AgentResponse(
                        success=not had_errors,
                        message=text or (tool_errors[-1] if had_errors else "Operação concluída."),
                        error=tool_errors[-1] if had_errors else None,
                        files_modified=files_modified,
                        is_chat=not bool(files_modified),
                    )

                # Execute each function call
                logger.info(f"[FC] Executando {len(fc_parts)} chamada(s) de ferramenta")
                contents.append(candidate.content)
                fn_result_parts = []

                for fc_part in fc_parts:
                    fc     = fc_part.function_call
                    fn_name = fc.name
                    fn_args = dict(fc.args) if fc.args else {}
                    logger.info(f"[FC] {fn_name}({list(fn_args.keys())})")

                    # Extract thought_signature if present (required by Gemini 3.x)
                    thought_sig = getattr(fc_part, 'thought_signature', None)

                    try:
                        result = self._execute_tool_call(fn_name, fn_args, available_files)
                        if isinstance(result, dict):
                            modified = result.get("modified_file")
                            op = fn_args.get("operation", "read")
                            if modified and op not in _READ_ONLY_OPS:
                                files_modified.append(modified)
                        fn_resp = genai_types.Part.from_function_response(
                            name=fn_name,
                            response={"result": str(result)[:3000]},
                        )
                    except Exception as tool_err:
                        logger.error(f"[FC] Erro em {fn_name}: {tool_err}", exc_info=True)
                        tool_errors.append(str(tool_err))
                        fn_resp = genai_types.Part.from_function_response(
                            name=fn_name,
                            response={"error": str(tool_err)},
                        )

                    # Propagate thought_signature for Gemini 3.x compatibility
                    if thought_sig is not None:
                        try:
                            fn_resp.thought_signature = thought_sig
                        except Exception:
                            pass  # SDK version may not support this field yet
                    fn_result_parts.append(fn_resp)

                contents.append(
                    genai_types.Content(role="tool", parts=fn_result_parts)
                )

            # If we exit the loop without returning, we hit the turn limit
            logger.warning(f"[FC] Atingido limite de {MAX_TURNS} turnos sem resposta final")
            return AgentResponse(
                success=bool(files_modified),
                message=(
                    f"Operação parcialmente concluída após {MAX_TURNS} etapas. "
                    f"Arquivos modificados: {len(files_modified)}. "
                    "Algumas operações podem não ter sido finalizadas."
                ),
                files_modified=files_modified,
                is_chat=False,
            )

            return AgentResponse(
                success=True,
                message="Operação concluída.",
                files_modified=files_modified,
                is_chat=not bool(files_modified),
            )

        except _QuotaExceededError:
            logger.warning("Quota/rate limit esgotada em todos os modelos")
            return AgentResponse(
                success=False,
                message=(
                    "⏳ Limite de requisições atingido em todos os modelos. "
                    "Aguarde alguns minutos e tente novamente. "
                    "Se o erro persistir, a cota diária gratuita foi esgotada — "
                    "tente novamente amanhã ou atualize para um plano pago."
                ),
                files_modified=[],
                is_chat=True,
                error=None,
            )
        except Exception as e:
            log_error_with_traceback(logger, e, "process_chat_message")
            return AgentResponse(
                success=False,
                message="Erro ao processar mensagem",
                files_modified=[],
                error=str(e)
            )

    def _execute_tool_call(self, tool_name: str, args: Dict[str, Any],
                           available_files: List[str]) -> Dict[str, Any]:
        """Execute a native Gemini function call and return a result dict.

        Args:
            tool_name: Name of the declared function (e.g. 'read_file')
            args: Arguments passed by the model
            available_files: Currently available file paths for resolution

        Returns:
            Dict with at least one of: 'content', 'files', 'success', 'modified_file'
        """
        if tool_name == "list_files":
            return {
                "files": [Path(f).name for f in available_files],
                "count": len(available_files),
            }

        filename = args.get("filename", "").strip()
        if not filename:
            raise ValueError("Parâmetro 'filename' é obrigatório")

        if tool_name == "read_file":
            file_path = self._resolve_filename(filename, available_files)
            content = self._read_file_content(file_path)
            self._last_read_result = content
            ext = Path(file_path).suffix.lower()
            return {
                "filename": Path(file_path).name,
                "content": self._format_file_content_for_llm(content, ext),
            }

        _TOOL_TYPE_MAP = {
            "excel_operation":       "excel",
            "word_operation":        "word",
            "pdf_operation":         "pdf",
            "powerpoint_operation":  "powerpoint",
        }
        tool_type = _TOOL_TYPE_MAP.get(tool_name)
        if not tool_type:
            raise ValueError(f"Ferramenta desconhecida: {tool_name}")

        operation = args.get("operation", "read")

        # For create/merge operations the target file doesn't exist yet — resolve to root_path
        _CREATE_OPS = {'create', 'merge'}
        if operation in _CREATE_OPS:
            if Path(filename).is_absolute():
                file_path = filename
            else:
                file_path = str(Path(self.config.root_path) / filename)
        else:
            file_path = self._resolve_filename(filename, available_files)

        # Normalize parameters: proto MapComposite or JSON string → plain dict
        raw_params = args.get("parameters") or {}
        if isinstance(raw_params, str):
            try:
                parameters = json.loads(raw_params)
            except Exception:
                parameters = {}
        elif hasattr(raw_params, 'items'):
            parameters = {k: (dict(v) if hasattr(v, 'items') else v)
                          for k, v in raw_params.items()}
        else:
            parameters = {}

        self.security_validator.validate_operation(operation)
        validated_path = self.security_validator.validate_file_path(file_path)
        validated_path_str = str(validated_path)

        self._last_read_result = None
        self._execute_single_action(tool_type, operation, validated_path_str, parameters)

        if self._last_read_result is not None:
            return {
                "filename": Path(file_path).name,
                "operation": operation,
                "content": str(self._last_read_result)[:4000],
            }

        return {
            "filename": Path(file_path).name,
            "operation": operation,
            "success": True,
            "modified_file": validated_path_str,
        }

    def _resolve_filename(self, filename: str, available_files: List[str]) -> str:
        """Resolve a bare filename to its full path from available files.

        Args:
            filename: File name (with or without extension or full path)
            available_files: List of full file paths to search

        Returns:
            Full path of the matched file

        Raises:
            ValueError: If the file cannot be found
        """
        if not filename:
            raise ValueError("Nome de arquivo vazio")

        if Path(filename).exists():
            return filename

        fn_lower = filename.strip().lower()

        for fp in available_files:
            if Path(fp).name.lower() == fn_lower:
                return fp

        fn_stem = Path(filename).stem.lower()
        for fp in available_files:
            if Path(fp).stem.lower() == fn_stem:
                return fp

        for fp in available_files:
            if fn_lower in Path(fp).name.lower():
                return fp

        all_scanned = self._discover_files()
        for fp in all_scanned:
            if Path(fp).name.lower() == fn_lower:
                return fp

        available_names = ", ".join(Path(f).name for f in available_files)
        raise ValueError(
            f"Arquivo '{filename}' não encontrado. "
            f"Disponíveis: {available_names}"
        )

    @staticmethod
    def _strip_json_wrapper(text: str) -> str:
        """Strip accidental JSON wrappers from model responses.

        Some Gemini models may wrap plain-text replies in JSON like:
            {"response": "Olá!"} or {"message": "Olá!"}
        This method extracts the inner text when that happens.
        """
        if not text:
            return text
        stripped = text.strip()
        if stripped.startswith("{") and stripped.endswith("}"):
            try:
                import json as _json
                data = _json.loads(stripped)
                if isinstance(data, dict):
                    # Try common keys the models use
                    for key in ("response", "message", "text", "reply", "answer", "content"):
                        if key in data and isinstance(data[key], str):
                            return data[key]
                    # If it's a single-key dict with a string value, extract it
                    if len(data) == 1:
                        val = next(iter(data.values()))
                        if isinstance(val, str):
                            return val
            except (ValueError, TypeError):
                pass
        return text

    @staticmethod
    def _generate_sample_data(col_names: list, row_count: int, filename: str, id_offset: int = 0) -> list:
        """Generate sample/simulated data rows for Excel creation.

        Uses the column names to infer data types and generate realistic values.
        Falls back to generic numeric data if column types can't be determined.

        Args:
            col_names: List of column header names (e.g. ['Date', 'Revenue', 'Expenses'])
                       If None, infers from filename.
            row_count: Number of data rows to generate
            filename: Target filename (used for context if col_names is empty)

        Returns:
            List of lists including header row + data rows
        """
        import random
        from datetime import datetime, timedelta

        if not col_names:
            # Infer columns from filename context
            fn_lower = filename.lower()
            if any(kw in fn_lower for kw in ('financ', 'revenue', 'receita', 'despesa', 'lucro')):
                col_names = ['Data', 'Receita', 'Despesas', 'Lucro']
            elif any(kw in fn_lower for kw in ('vendas', 'sales', 'venda')):
                col_names = ['Data', 'Produto', 'Quantidade', 'Valor_Unitario', 'Total']
            elif any(kw in fn_lower for kw in ('funcionario', 'employe', 'rh', 'pessoal')):
                col_names = ['Nome', 'Cargo', 'Departamento', 'Salario']
            else:
                col_names = ['ID', 'Valor_A', 'Valor_B', 'Total']

        rows = [col_names]  # header row

        # Column type inference based on common patterns
        _NAMES = ['Ana Silva', 'Bruno Costa', 'Carla Souza', 'Daniel Lima', 'Elena Santos',
                  'Fábio Oliveira', 'Gabriela Rocha', 'Henrique Alves', 'Isabela Ferreira',
                  'João Mendes', 'Karine Dias', 'Lucas Barbosa', 'Mariana Nunes', 'Nelson Ramos',
                  'Olívia Cardoso', 'Pedro Gomes', 'Rafaela Ribeiro', 'Sérgio Pereira',
                  'Tatiana Araújo', 'Vinícius Martins']
        _PRODUCTS = ['Produto A', 'Produto B', 'Produto C', 'Produto D', 'Produto E',
                     'Serviço X', 'Serviço Y', 'Serviço Z', 'Plano Basic', 'Plano Premium']
        _DEPTS = ['TI', 'RH', 'Marketing', 'Financeiro', 'Comercial', 'Operações']
        _CARGOS = ['Analista', 'Coordenador', 'Gerente', 'Diretor', 'Estagiário', 'Assistente']
        _CITIES = ['São Paulo', 'Rio de Janeiro', 'Belo Horizonte', 'Curitiba', 'Porto Alegre',
                   'Salvador', 'Fortaleza', 'Recife', 'Brasília', 'Manaus']

        base_date = datetime(2024, 1, 1)

        for i in range(row_count):
            i = i + id_offset
            row = []
            receita = despesa = None
            qtd = preco = None
            for col in col_names:
                cl = col.lower().replace('_', ' ')
                if any(k in cl for k in ('data', 'date', 'mes', 'periodo')):
                    row.append((base_date + timedelta(days=i)).strftime('%Y-%m-%d'))
                elif any(k in cl for k in ('nome', 'name', 'funcionario', 'cliente', 'vendedor', 'representante', 'responsavel')):
                    row.append(random.choice(_NAMES))
                elif any(k in cl for k in ('produto', 'product', 'item')):
                    row.append(random.choice(_PRODUCTS))
                elif any(k in cl for k in ('categoria', 'category', 'tipo', 'type', 'grupo', 'segmento')):
                    row.append(random.choice(['Eletrônicos', 'Roupas', 'Alimentos', 'Serviços', 'Software', 'Hardware', 'Consultoria', 'Treinamento']))
                elif any(k in cl for k in ('departamento', 'dept', 'setor', 'area')):
                    row.append(random.choice(_DEPTS))
                elif any(k in cl for k in ('cargo', 'position', 'funcao')):
                    row.append(random.choice(_CARGOS))
                elif any(k in cl for k in ('cidade', 'city', 'estado')):
                    row.append(random.choice(_CITIES))
                elif any(k in cl for k in ('receita', 'revenue', 'faturamento', 'renda')):
                    receita = round(random.uniform(10000, 500000), 2)
                    row.append(receita)
                elif any(k in cl for k in ('despesa', 'expense', 'custo', 'gasto')):
                    despesa = round(random.uniform(5000, 300000), 2)
                    row.append(despesa)
                elif any(k in cl for k in ('desconto', 'discount', 'abatimento', 'reducao')):
                    row.append(round(random.uniform(0, 30), 2))
                elif any(k in cl for k in ('status', 'situacao', 'estado', 'condicao')):
                    row.append(random.choice(['Ativo', 'Inativo', 'Pendente', 'Concluído', 'Cancelado', 'Em andamento']))
                elif any(k in cl for k in ('lucro', 'profit', 'resultado', 'margem')):
                    if receita is not None and despesa is not None:
                        row.append(round(receita - despesa, 2))
                    else:
                        row.append(round(random.uniform(-50000, 200000), 2))
                elif any(k in cl for k in ('salario', 'salary', 'remuneracao')):
                    row.append(round(random.uniform(2000, 25000), 2))
                elif any(k in cl for k in ('quantidade', 'qty', 'qtd', 'unidades')):
                    qtd = random.randint(1, 500)
                    row.append(qtd)
                elif any(k in cl for k in ('valor unitario', 'preco', 'price', 'valor unit')):
                    preco = round(random.uniform(10, 5000), 2)
                    row.append(preco)
                elif any(k in cl for k in ('total', 'subtotal', 'valor total')):
                    if qtd is not None and preco is not None:
                        row.append(round(qtd * preco, 2))
                    else:
                        row.append(round(random.uniform(100, 100000), 2))
                elif any(k in cl for k in ('id', 'codigo', 'code', 'numero')):
                    row.append(i + 1)  # id_offset already applied to i
                elif any(k in cl for k in ('idade', 'age')):
                    row.append(random.randint(18, 65))
                elif any(k in cl for k in ('email',)):
                    name = random.choice(_NAMES).split()[0].lower()
                    row.append(f"{name}{random.randint(1,99)}@email.com")
                elif any(k in cl for k in ('telefone', 'phone', 'tel')):
                    row.append(f"(11) 9{random.randint(1000,9999)}-{random.randint(1000,9999)}")
                else:
                    # Generic column: vary type based on column name/position for realism
                    col_idx = col_names.index(col)
                    _is_single_letter = len(col.strip()) == 1 and col.strip().isalpha()
                    _is_generic = _is_single_letter or cl.startswith('col') or cl.startswith('campo')
                    if _is_generic:
                        # Cycle through int / float / text / bool for variety
                        cycle = col_idx % 4
                        if cycle == 0:
                            row.append(random.randint(1, 1000))
                        elif cycle == 1:
                            row.append(round(random.uniform(0, 10000), 2))
                        elif cycle == 2:
                            row.append(random.choice(_PRODUCTS + _CITIES + _NAMES))
                        else:
                            row.append(random.choice(['Sim', 'Não', 'Pendente', 'Ativo', 'Inativo']))
                    else:
                        row.append(round(random.uniform(0, 10000), 2))
            rows.append(row)

        return rows

    @staticmethod
    def _is_chat_only_message(message: str) -> bool:
        """Detecta mensagens puramente conversacionais que não envolvem arquivos.

        Retorna True para saudações, agradecimentos, perguntas sobre o agente
        e mensagens curtas sem verbos de ação sobre arquivos.
        """
        import re
        msg = message.strip().lower()

        # Greetings and salutations
        _GREETINGS = {
            'ola', 'olá', 'oi', 'oie', 'eai', 'e ai', 'e aí',
            'hey', 'hi', 'hello',
            'bom dia', 'boa tarde', 'boa noite',
            'tudo bem', 'tudo bom', 'como vai', 'como você está',
            'tudo certo', 'td bem', 'td bom',
        }
        if msg in _GREETINGS or any(
            msg.startswith(g + ' ') or msg.startswith(g + '!')
            or msg.startswith(g + '?') or msg.startswith(g + ',')
            or msg == g
            for g in _GREETINGS
        ):
            return True

        # Gratitude / feedback
        _THANKS = {'obrigado', 'obrigada', 'valeu', 'vlw', 'thanks', 'thank you', 'thx', 'brigado', 'brigada'}
        if msg in _THANKS or any(msg.startswith(t) for t in _THANKS):
            return True

        # Positive feedback without action
        _FEEDBACK = {'ótimo', 'otimo', 'perfeito', 'show', 'legal', 'massa', 'incrível', 'incrivel', 'muito bom', 'excelente', 'parabens', 'parabéns'}
        if msg in _FEEDBACK:
            return True

        # Questions about the agent itself (not about a file)
        _ABOUT_AGENT = [
            r'o que (você|voce) (faz|pode|consegue|sabe)',
            r'(quem|what) (é|are) (você|voce|you)',
            r'me (ajude|ajuda|fale sobre você|fale sobre voce)',
            r'(pode|can you) me ajudar',
            r'what can you do',
            r'how (do|can) (i|you)',
            r'(suas|tuas) (funcionalidades|capacidades|habilidades)',
            r'o que (é|e) (você|voce|o gemini)',
        ]
        for pattern in _ABOUT_AGENT:
            if re.search(pattern, msg):
                return True

        # Action VERBS that indicate file operations — if present, NOT chat-only.
        # Intentionally excludes nouns (pdf, excel, arquivo) since they appear in questions too.
        _ACTION_KEYWORDS = [
            'crie', 'cria', 'criar', 'faça', 'faz', 'fazer',
            'abra', 'abre', 'abrir', 'edite', 'edita', 'editar',
            'adicione', 'adiciona', 'adicionar', 'remova', 'remove', 'remover',
            'delete', 'deleta', 'deletar', 'exclua', 'exclui', 'excluir',
            'altere', 'altera', 'alterar', 'mude', 'muda', 'mudar',
            'atualize', 'atualiza', 'atualizar', 'formate', 'formata', 'formatar',
            'converta', 'converte', 'converter', 'gere', 'gera', 'gerar',
            'ordene', 'ordena', 'ordenar', 'filtre', 'filtra', 'filtrar',
            'insira', 'insere', 'inserir', 'salve', 'salva', 'salvar',
            'escreva', 'escreve', 'escrever',
            'create', 'make', 'generate', 'add', 'delete',
            'update', 'edit', 'format', 'convert', 'sort', 'filter',
            'insert', 'save', 'write', 'merge', 'split',
            '.xlsx', '.docx', '.pptx', '.pdf',
        ]
        for kw in _ACTION_KEYWORDS:
            if kw in msg:
                return False

        # Short messages (≤3 words) with no action keywords → likely casual
        word_count = len(msg.split())
        if word_count <= 3:
            return True

        return False

    def _discover_files(self) -> List[str]:
        """Descobre arquivos Office disponíveis usando FileScanner.
        
        Returns:
            Lista de caminhos completos dos arquivos Office encontrados
        
        Validates: Requirements 7.1
        """
        logger.info("Descobrindo arquivos Office disponíveis")
        files = self.file_scanner.scan_office_files()
        logger.debug(f"Arquivos descobertos: {files}")
        return files
    
    def _filter_relevant_files(self, available_files: List[str], user_prompt: str) -> List[str]:
        """Filtra arquivos relevantes baseado no prompt do usuário.
        
        Analisa o prompt do usuário para identificar menções a nomes de arquivos
        ou tipos de arquivos específicos.
        
        Args:
            available_files: Lista de todos os arquivos disponíveis
            user_prompt: Prompt do usuário
        
        Returns:
            Lista de arquivos relevantes para o prompt
        
        Validates: Requirements 7.1
        """
        logger.info("Filtrando arquivos relevantes baseado no prompt")
        
        # Converte prompt para minúsculas para comparação
        prompt_lower = user_prompt.lower()
        
        relevant_files = []
        
        for file_path in available_files:
            file_name = Path(file_path).name.lower()
            file_stem = Path(file_path).stem.lower()
            
            # Verifica se o nome do arquivo é mencionado no prompt
            if file_name in prompt_lower or file_stem in prompt_lower:
                relevant_files.append(file_path)
                logger.debug(f"Arquivo relevante por nome: {file_path}")
                continue
            
            # Verifica menções a tipos de arquivo
            extension = Path(file_path).suffix.lower()
            if extension == '.xlsx' and any(word in prompt_lower for word in ['excel', 'planilha', 'spreadsheet']):
                relevant_files.append(file_path)
                logger.debug(f"Arquivo relevante por tipo (Excel): {file_path}")
            elif extension == '.docx' and any(word in prompt_lower for word in ['word', 'documento', 'document']):
                relevant_files.append(file_path)
                logger.debug(f"Arquivo relevante por tipo (Word): {file_path}")
            elif extension == '.pptx' and any(word in prompt_lower for word in ['powerpoint', 'apresentação', 'presentation', 'slide']):
                relevant_files.append(file_path)
                logger.debug(f"Arquivo relevante por tipo (PowerPoint): {file_path}")
            elif extension == '.pdf' and any(word in prompt_lower for word in ['pdf', 'documento pdf']):
                relevant_files.append(file_path)
                logger.debug(f"Arquivo relevante por tipo (PDF): {file_path}")
        
        # Se nenhum arquivo específico foi identificado, retorna todos
        # (deixa o Gemini decidir quais são relevantes)
        if not relevant_files:
            logger.info("Nenhum arquivo específico identificado, retornando todos")
            return available_files
        
        return relevant_files

    
    def _read_file_content(self, file_path: str) -> Any:
        """Lê conteúdo de arquivo usando ferramenta apropriada baseada na extensão.
        
        Seleciona automaticamente a ferramenta correta (Excel, Word ou PowerPoint)
        baseado na extensão do arquivo.
        
        Args:
            file_path: Caminho do arquivo a ser lido
        
        Returns:
            Conteúdo do arquivo em formato estruturado (depende do tipo)
        
        Raises:
            ValueError: Se a extensão do arquivo não for suportada
            FileNotFoundError: Se o arquivo não existir
            CorruptedFileError: Se o arquivo estiver corrompido
        
        Validates: Requirements 7.2
        """
        logger.info(f"Lendo conteúdo do arquivo: {file_path}")
        
        extension = Path(file_path).suffix.lower()
        
        if extension == '.xlsx':
            logger.debug("Usando ExcelTool para ler arquivo")
            return self.excel_tool.read_excel(file_path)
        
        elif extension == '.docx':
            logger.debug("Usando WordTool para ler arquivo")
            # Retorna estrutura compatível com format_file_content
            text = self.word_tool.read_word(file_path)
            tables = self.word_tool.extract_tables(file_path)
            return {
                'paragraphs': text.split('\n'),
                'tables': tables,
                'metadata': {}
            }
        
        elif extension == '.pptx':
            logger.debug("Usando PowerPointTool para ler arquivo")
            # Retorna estrutura compatível com format_file_content
            slides_data = self.powerpoint_tool.read_powerpoint(file_path)
            return {
                'slides': slides_data,
                'metadata': {}
            }
        
        elif extension == '.pdf':
            logger.debug("Usando PdfTool para ler arquivo")
            return self.pdf_tool.read_pdf(file_path)
        
        else:
            logger.error(f"Extensão de arquivo não suportada: {extension}")
            raise ValueError(f"Extensão de arquivo não suportada: {extension}")
    
    def _get_file_type(self, file_path: str) -> str:
        """Determina o tipo de arquivo baseado na extensão.
        
        Args:
            file_path: Caminho do arquivo
        
        Returns:
            Tipo do arquivo: 'excel', 'word' ou 'powerpoint'
        """
        extension = Path(file_path).suffix.lower()
        
        if extension == '.xlsx':
            return 'excel'
        elif extension == '.docx':
            return 'word'
        elif extension == '.pptx':
            return 'powerpoint'
        elif extension == '.pdf':
            return 'pdf'
        else:
            return 'unknown'

    
    def _resolve_chart_position(self, file_path: str, sheet: str,
                                chart_index: int = None, raw_position: str = None) -> str:
        """Resolve posição relativa/especial para referência absoluta de célula.

        Palavras-chave aceitas (case-insensitive):
          beside_content / next_to_data   → coluna(max_col + 2), linha 1
          below_content / abaixo          → coluna A, linha(max_row + 2)
          right:N                         → coluna atual + N (requer chart_index)
          left:N                          → coluna atual - N (requer chart_index)
          down:N / baixo:N                → linha atual + N (requer chart_index)
          up:N / cima:N                   → linha atual - N (requer chart_index)
          after_last_chart                → posição após o último gráfico da planilha
          Qualquer outra string           → retorna como está (posição absoluta)
        """
        import openpyxl
        from openpyxl.utils import get_column_letter

        if not raw_position:
            raw_position = "beside_content"

        pos = raw_position.strip().lower()

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb[sheet] if sheet in wb.sheetnames else wb.active
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            wb.close()
        except Exception:
            max_row, max_col = 1, 1

        # --- Special keywords ---
        if pos in ("beside_content", "next_to_data", "ao_lado", "beside"):
            return f"{get_column_letter(max_col + 2)}1"

        if pos in ("below_content", "abaixo", "below", "abaixo_do_conteudo"):
            return f"A{max_row + 2}"

        # --- Relative: right:N / left:N / down:N / up:N ---
        for prefix, direction in (
            ("right:", "right"), ("left:", "left"),
            ("down:", "down"), ("up:", "up"),
            ("baixo:", "down"), ("cima:", "up"),
            ("direita:", "right"), ("esquerda:", "left"),
        ):
            if pos.startswith(prefix):
                try:
                    n = int(pos[len(prefix):])
                except ValueError:
                    n = 1
                # Need current chart position
                cur_col, cur_row = 8, 2  # fallback H2
                if chart_index is not None:
                    try:
                        charts_info = self.excel_tool.list_charts(file_path, sheet)
                        for c in charts_info.get("charts", []):
                            if c["index"] == chart_index and c["position"] != "Unknown":
                                cpos = c["position"]
                                from openpyxl.utils import coordinate_to_tuple
                                cur_row, cur_col = coordinate_to_tuple(cpos)
                                break
                    except Exception:
                        pass
                if direction == "right":
                    new_col = max(1, cur_col + n)
                    return f"{get_column_letter(new_col)}{cur_row}"
                elif direction == "left":
                    new_col = max(1, cur_col - n)
                    return f"{get_column_letter(new_col)}{cur_row}"
                elif direction == "down":
                    return f"{get_column_letter(cur_col)}{max(1, cur_row + n)}"
                elif direction == "up":
                    return f"{get_column_letter(cur_col)}{max(1, cur_row - n)}"

        if pos == "after_last_chart":
            try:
                charts_info = self.excel_tool.list_charts(file_path, sheet)
                if charts_info.get("charts"):
                    positions = [c["position"] for c in charts_info["charts"]
                                 if c.get("position") and c["position"] != "Unknown"]
                    if positions:
                        from openpyxl.utils import coordinate_to_tuple
                        max_c = max(coordinate_to_tuple(p)[1] for p in positions)
                        return f"{get_column_letter(max_c + 2)}1"
            except Exception:
                pass
            return f"{get_column_letter(max_col + 2)}1"

        # Absolute position — return as-is
        return raw_position

    def _format_excel_for_llm(self, content: dict, max_chars: int = 50000) -> str:
        """Formata conteúdo Excel de forma compacta e informativa para o modelo.

        Muito mais eficiente que str(dict)[:4000]: preserva estrutura de tabela,
        inclui posições de gráficos e marca claramente se houve truncamento.
        """
        parts = []
        sheets = content.get("sheets", {})
        charts = content.get("charts", {})
        meta = content.get("metadata", {})

        if meta.get("sheet_names"):
            parts.append(f"Planilhas: {', '.join(meta['sheet_names'])}")

        for sheet_name, rows in sheets.items():
            sheet_charts = charts.get(sheet_name, [])
            non_empty = [r for r in rows if any(v is not None for v in r)]
            total_rows = len(non_empty)
            total_cols = max((len(r) for r in non_empty), default=0)
            parts.append(f"\n--- {sheet_name} ({total_rows} linhas × {total_cols} colunas) ---")

            # Table as pipe-separated text (header + rows, truncated)
            MAX_ROWS_PER_SHEET = 500
            rows_to_show = non_empty[:MAX_ROWS_PER_SHEET]
            for row in rows_to_show:
                parts.append(" | ".join("" if v is None else str(v) for v in row))
            if total_rows > MAX_ROWS_PER_SHEET:
                parts.append(f"... [{total_rows - MAX_ROWS_PER_SHEET} linhas omitidas]")

            if sheet_charts:
                parts.append(f"Gráficos em {sheet_name}:")
                for c in sheet_charts:
                    pos = c.get("position", "?")
                    parts.append(f"  [{c['index']}] {c['type']} '{c['title']}' @ {pos}")

        result = "\n".join(parts)
        if len(result) > max_chars:
            result = result[:max_chars] + f"\n... [truncado em {max_chars} chars]"
        return result

    def _format_file_content_for_llm(self, content: Any, extension: str) -> str:
        """Converte conteúdo de arquivo para string otimizada para o modelo."""
        if extension == ".xlsx" and isinstance(content, dict) and "sheets" in content:
            return self._format_excel_for_llm(content)
        text = str(content)
        if len(text) > 50000:
            return text[:50000] + "\n... [truncado]"
        return text

    def _build_context_prompt(self, user_prompt: str, file_contexts: List[Dict[str, Any]]) -> str:
        """Constrói prompt contextualizado usando prompt_templates.
        
        Combina o prompt do usuário com o conteúdo dos arquivos relevantes
        para criar um prompt completo para o Gemini.
        
        Args:
            user_prompt: Prompt original do usuário
            file_contexts: Lista de contextos de arquivos formatados
        
        Returns:
            Prompt contextualizado completo
        
        Validates: Requirements 7.3, 7.4
        """
        logger.info("Construindo prompt contextualizado")
        
        prompt = PromptTemplates.build_context_prompt(user_prompt, file_contexts)
        
        logger.debug(f"Prompt construído com {len(file_contexts)} arquivo(s) no contexto")
        
        return prompt

    
    def _execute_actions(self, gemini_response: str) -> AgentResponse:
        """Interpreta resposta do Gemini e executa ações usando response_parser e validação de segurança.
        
        Parse a resposta do Gemini, valida as ações propostas e executa-as
        usando as ferramentas apropriadas. Suporta operações em lote com
        tratamento de erros parciais.
        
        Args:
            gemini_response: Resposta em texto do Gemini API
        
        Returns:
            AgentResponse com resultado da execução (incluindo BatchResult para múltiplas ações)
        
        Validates: Requirements 7.5, 7.6, 7.7
        """
        logger.info("Interpretando resposta do Gemini e executando ações")
        
        try:
            # Parse da resposta usando ResponseParser
            parsed_response = ResponseParser.parse_response(gemini_response)
            
            # Extrai ações e explicação
            actions = ResponseParser.extract_actions(parsed_response)
            explanation = ResponseParser.extract_explanation(parsed_response)
            
            logger.info(f"Resposta parseada: {len(actions)} ação(ões) identificada(s)")
            logger.info(f"Explicação: {explanation}")
            
            # Log detalhado das ações para debug
            for idx, action in enumerate(actions):
                logger.debug(f"Ação {idx + 1}: {json.dumps(action, indent=2, ensure_ascii=False)}")
            
            # Deduplica ações 'create' repetidas no mesmo arquivo: o modelo às
            # vezes gera múltiplos update_content (→ create) para o mesmo arquivo,
            # o que causaria recriações redundantes. Mantém apenas a 1ª ocorrência.
            _OVERWRITE_OPS = {'create'}
            _seen_create = set()
            deduplicated = []
            for _a in actions:
                _key = (_a.get('tool'), _a.get('operation'), _a.get('target_file'))
                if _a.get('operation') in _OVERWRITE_OPS:
                    if _key in _seen_create:
                        logger.info(
                            f"Ação create duplicada ignorada para: {_a.get('target_file')}"
                        )
                        continue
                    _seen_create.add(_key)
                deduplicated.append(_a)
            if len(deduplicated) < len(actions):
                logger.info(
                    f"Deduplicação: {len(actions)} → {len(deduplicated)} ação(ões)"
                )
            actions = deduplicated

            # Determinar se é operação em lote
            is_batch = len(actions) > 1
            
            # Executa cada ação e coleta resultados
            files_modified = []
            action_results = []
            successful_count = 0
            failed_count = 0
            
            for i, action in enumerate(actions):
                logger.info(f"Executando ação {i + 1}/{len(actions)}")
                
                tool = action['tool']
                operation = action['operation']
                target_file = action['target_file']
                parameters = action['parameters']
                
                logger.debug(f"Ação: tool={tool}, operation={operation}, file={target_file}")
                
                action_success = False
                action_error = None
                validated_path_str = target_file
                
                try:
                    # Valida operação usando SecurityValidator
                    self.security_validator.validate_operation(operation)
                    
                    # Valida caminho do arquivo
                    validated_path = self.security_validator.validate_file_path(target_file)
                    validated_path_str = str(validated_path)
                    
                    # Executa ação baseada no tool e operation
                    self._execute_single_action(tool, operation, validated_path_str, parameters)

                    # Read-only operations don't modify the file
                    _READ_ONLY_OPS = {
                        'read', 'get_info', 'list_charts', 'analyze_document',
                        'analyze_word_count', 'analyze_section_length',
                        'get_document_statistics', 'analyze_tone', 'identify_jargon',
                        'analyze_readability', 'check_term_consistency',
                        'extract_tables', 'get_page_layout',
                    }
                    if operation not in _READ_ONLY_OPS:
                        files_modified.append(validated_path_str)
                    action_success = True
                    successful_count += 1
                    logger.info(f"Ação {i + 1} executada com sucesso")
                    
                except ValidationError as e:
                    action_error = f"Erro de validação: {str(e)}"
                    failed_count += 1
                    logger.error(f"Erro de validação na ação {i + 1}: {e}", exc_info=True)
                    
                    # Para operação única, retornar erro imediatamente
                    if not is_batch:
                        return AgentResponse(
                            success=False,
                            message=f"Erro de validação na ação: {e}",
                            files_modified=files_modified,
                            error=str(e)
                        )
                
                except (ValueError, KeyError) as e:
                    action_error = f"Erro nos parâmetros: {str(e)}"
                    failed_count += 1
                    logger.error(f"Erro nos parâmetros da ação {i + 1}: {e}", exc_info=True)
                    
                    # Para operação única, retornar erro imediatamente
                    if not is_batch:
                        return AgentResponse(
                            success=False,
                            message=f"Erro nos parâmetros da ação: {e}",
                            files_modified=files_modified,
                            error=str(e)
                        )
                
                except Exception as e:
                    action_error = str(e)
                    failed_count += 1
                    logger.error(f"Erro ao executar ação {i + 1}: {e}", exc_info=True)
                    
                    # Para operação única, retornar erro imediatamente
                    if not is_batch:
                        return AgentResponse(
                            success=False,
                            message=f"Erro ao executar ação: {e}",
                            files_modified=files_modified,
                            error=str(e)
                        )
                
                # Registrar resultado da ação
                if is_batch:
                    action_results.append(ActionResult(
                        action_index=i,
                        tool=tool,
                        operation=operation,
                        target_file=validated_path_str,
                        success=action_success,
                        error=action_error
                    ))
            
            # Construir resposta
            if is_batch:
                batch_result = BatchResult(
                    total_actions=len(actions),
                    successful_actions=successful_count,
                    failed_actions=failed_count,
                    action_results=action_results,
                    overall_success=successful_count > 0
                )
                
                # Mensagem de resumo
                if failed_count == 0:
                    message = f"✅ Todas as {successful_count} operações foram concluídas com sucesso"
                elif successful_count == 0:
                    message = f"❌ Todas as {failed_count} operações falharam"
                else:
                    message = f"⚠️ {successful_count} de {len(actions)} operações bem-sucedidas ({failed_count} falharam)"
                
                if explanation:
                    message = f"{explanation}\n{message}"
                
                return AgentResponse(
                    success=batch_result.overall_success,
                    message=message,
                    files_modified=files_modified,
                    error=None if successful_count > 0 else "Todas as operações falharam",
                    batch_result=batch_result
                )
            else:
                # Operação única bem-sucedida
                success_message = explanation if explanation else "Operação concluída com sucesso"
                
                return AgentResponse(
                    success=True,
                    message=success_message,
                    files_modified=files_modified,
                    error=None
                )
            
        except ValidationError as e:
            logger.error(f"Erro de validação: {e}", exc_info=True)
            return AgentResponse(
                success=False,
                message=f"Erro de validação: {e}",
                files_modified=[],
                error=str(e)
            )
        
        except Exception as e:
            logger.error(f"Erro ao executar ações: {e}", exc_info=True)
            return AgentResponse(
                success=False,
                message=f"Erro ao executar ações: {e}",
                files_modified=[],
                error=str(e)
            )

    
    def _execute_single_action(self, tool: str, operation: str, target_file: str, parameters: Dict[str, Any]) -> None:
        """Executa uma única ação usando a ferramenta apropriada.
        
        Args:
            tool: Nome da ferramenta ('excel', 'word', 'powerpoint')
            operation: Tipo de operação ('read', 'create', 'update', 'add')
            target_file: Caminho do arquivo alvo
            parameters: Parâmetros específicos da operação
        
        Raises:
            ValueError: Se tool ou operation não forem reconhecidos
            Exception: Erros específicos das ferramentas
        
        Validates: Requirements 7.6
        """
        logger.info(f"Executando ação: {tool}.{operation} em {target_file}")

        # Auto-resolve sheet name when not provided (common issue with models)
        if tool == 'excel' and operation not in ('create', 'read'):
            sheet_param = parameters.get('sheet') or parameters.get('sheet_name')
            if not sheet_param or sheet_param == 'None':
                try:
                    import openpyxl as _opx
                    _wb = _opx.load_workbook(target_file, read_only=True)
                    _resolved = _wb.active.title if _wb.active else _wb.sheetnames[0]
                    _wb.close()
                    parameters['sheet'] = _resolved
                    if parameters.get('sheet_name') == 'None':
                        parameters['sheet_name'] = _resolved
                    logger.info(f"[FC] Auto-resolved sheet name: {_resolved}")
                except Exception:
                    pass
        
        # Criar backup antes de operações que modificam arquivos
        version_id = None
        if self.version_manager and operation in ['update', 'add', 'create', 'add_chart', 'delete_chart', 'sort', 'delete_sheet', 'delete_rows', 'format', 'formula', 'merge']:
            user_prompt = parameters.get('_user_prompt', 'Operação do usuário')
            version_id = self.version_manager.create_backup(
                target_file, operation, user_prompt
            )
            if version_id:
                logger.info(f"Backup criado: {version_id}")
        
        try:
            # Excel operations
            if tool == 'excel':
                if operation == 'read':
                    result = self.excel_tool.read_excel(target_file)
                    self._last_read_result = result
                    
                elif operation == 'create':
                    data = parameters.get('data', {})
                    if isinstance(data, str):
                        try:
                            data = json.loads(data)
                        except Exception:
                            data = {}
                    # Normalize: if model sent a flat list of rows, wrap as single sheet
                    if isinstance(data, list):
                        data = {"Sheet1": data}

                    # Server-side data generation: if row_count is provided but data
                    # is empty/sparse, generate sample rows using Python
                    row_count = parameters.get('row_count') or parameters.get('count')
                    if row_count and isinstance(row_count, (int, float)) and int(row_count) > 0:
                        row_count = int(row_count)
                        columns_str = parameters.get('columns', '')
                        if isinstance(columns_str, str) and columns_str:
                            col_names = [c.strip() for c in columns_str.split(',') if c.strip()]
                        else:
                            col_names = None

                        # Check if data is empty or only has headers
                        _sheet_data = next(iter(data.values()), []) if data else []
                        _has_enough_data = len(_sheet_data) > max(2, row_count // 2)

                        if not _has_enough_data:
                            generated = self._generate_sample_data(col_names, row_count, target_file)
                            if generated:
                                data = {"Sheet1": generated}
                                logger.info(f"[FC] Server-side generated {row_count} sample rows")

                    self.excel_tool.create_excel(target_file, data)
                    
                elif operation == 'update':
                    # Suporta update_range (múltiplas células) ou update_cell (célula única)
                    updates = parameters.get('updates')
                    if updates:
                        sheet = parameters.get('sheet')
                        self.excel_tool.update_range(target_file, sheet, updates)
                    else:
                        sheet = parameters.get('sheet')
                        row = parameters.get('row')
                        col = parameters.get('col')
                        if col is None and parameters.get('column'):
                            from openpyxl.utils import column_index_from_string
                            col = column_index_from_string(str(parameters['column']))
                        value = parameters.get('value') or parameters.get('content')
                        self.excel_tool.update_cell(target_file, sheet, row, col, value)
                    
                elif operation == 'add':
                    sheet_name = parameters.get('sheet_name') or parameters.get('sheet')
                    data = parameters.get('data', [])
                    if isinstance(data, str):
                        try:
                            data = json.loads(data)
                        except Exception:
                            data = []
                    self.excel_tool.add_sheet(target_file, sheet_name, data)
                
                elif operation == 'append':
                    sheet = parameters.get('sheet')
                    rows = parameters.get('rows', [])
                    if isinstance(rows, str):
                        try:
                            rows = json.loads(rows)
                        except Exception:
                            rows = []
                    # Server-side generation: if row_count provided but rows empty/small
                    row_count = parameters.get('row_count') or parameters.get('count')
                    if row_count and isinstance(row_count, (int, float)) and int(row_count) > 0:
                        row_count = int(row_count)
                        if len(rows) < max(1, row_count // 2):
                            # Infer columns from the existing sheet header
                            col_names = None
                            try:
                                import openpyxl as _opx
                                _wb2 = _opx.load_workbook(target_file, read_only=True)
                                _ws2 = _wb2[sheet] if sheet and sheet in _wb2.sheetnames else _wb2.active
                                col_names = [_ws2.cell(1, c).value for c in range(1, (_ws2.max_column or 1) + 1)
                                             if _ws2.cell(1, c).value is not None]
                                _wb2.close()
                            except Exception:
                                pass
                            columns_str = parameters.get('columns', '')
                            if isinstance(columns_str, str) and columns_str:
                                col_names = [c.strip() for c in columns_str.split(',') if c.strip()]
                            # Calculate current data row count for ID continuity
                            id_offset = 0
                            try:
                                import openpyxl as _opx3
                                _wb3 = _opx3.load_workbook(target_file, read_only=True)
                                _ws3 = _wb3[sheet] if sheet and sheet in _wb3.sheetnames else _wb3.active
                                id_offset = max(0, (_ws3.max_row or 1) - 1)  # subtract header
                                _wb3.close()
                            except Exception:
                                pass
                            generated = self._generate_sample_data(col_names, row_count, target_file, id_offset=id_offset)
                            rows = generated[1:]  # skip header row
                            logger.info(f"[FC] Server-side generated {len(rows)} rows for append (id_offset={id_offset})")
                    self.excel_tool.append_rows(target_file, sheet, rows)
                
                elif operation == 'delete_sheet':
                    sheet_name = parameters.get('sheet_name') or parameters.get('sheet')
                    self.excel_tool.delete_sheet(target_file, sheet_name)
                
                elif operation == 'delete_rows':
                    sheet = parameters.get('sheet')
                    start_row = parameters.get('start_row') or parameters.get('row')
                    count = parameters.get('count', 1)
                    self.excel_tool.delete_rows(target_file, sheet, start_row, count)
                
                elif operation == 'format':
                    sheet = parameters.get('sheet')
                    formatting = parameters.get('formatting') or {}
                    if isinstance(formatting, str):
                        try:
                            formatting = json.loads(formatting)
                        except Exception:
                            formatting = {}
                    if not formatting.get('range'):
                        cell_range = (
                            parameters.get('range')
                            or parameters.get('cell_range')
                        )
                        if cell_range:
                            formatting['range'] = cell_range
                    for k in ('bold', 'italic', 'font_size', 'font_color', 'bg_color',
                              'number_format', 'alignment', 'border', 'wrap_text'):
                        if parameters.get(k) is not None and k not in formatting:
                            formatting[k] = parameters[k]
                    self.excel_tool.format_cells(target_file, sheet, formatting)
                
                elif operation == 'auto_width':
                    sheet = parameters.get('sheet')
                    self.excel_tool.auto_width(target_file, sheet)
                
                elif operation == 'formula':
                    sheet = parameters.get('sheet')
                    formulas = parameters.get('formulas')
                    if formulas:
                        self.excel_tool.set_formulas(target_file, sheet, formulas)
                    else:
                        row = parameters.get('row')
                        col = parameters.get('col')
                        if col is None and parameters.get('column'):
                            from openpyxl.utils import column_index_from_string
                            col = column_index_from_string(str(parameters['column']))
                        formula = parameters.get('formula') or parameters.get('content')
                        self.excel_tool.set_formula(target_file, sheet, row, col, formula)
                
                elif operation == 'merge':
                    sheet = parameters.get('sheet')
                    cell_range = parameters.get('range')
                    self.excel_tool.merge_cells(target_file, sheet, cell_range)
                
                elif operation == 'add_chart':
                    sheet = parameters.get('sheet')
                    chart_config = parameters.get('chart_config') or {
                        'type': parameters.get('chart_type', 'bar'),
                        'title': parameters.get('title'),
                        'position': parameters.get('position'),
                        'data_range': parameters.get('data_range') or parameters.get('range'),
                    }
                    # Auto-infer chart data if not provided — read sheet dimensions
                    if not chart_config.get('data_range') and not chart_config.get('values'):
                        _wb = None
                        try:
                            import openpyxl
                            from openpyxl.utils import get_column_letter
                            _wb = openpyxl.load_workbook(target_file, read_only=True)
                            _ws = _wb[sheet] if sheet and sheet in _wb.sheetnames else _wb.active
                            if not sheet:
                                sheet = _ws.title
                            max_row = _ws.max_row or 1
                            max_col = _ws.max_column or 1

                            chart_type = (chart_config.get('type') or 'bar').lower()
                            if chart_type == 'pie' and max_col >= 2:
                                # Pie chart: categories=col A (labels), values=col B (numbers)
                                # Find first numeric column for values
                                val_col = 2  # default: column B
                                cat_col = 1  # default: column A
                                for c in range(1, max_col + 1):
                                    cell_val = _ws.cell(row=2, column=c).value
                                    if isinstance(cell_val, (int, float)):
                                        val_col = c
                                        break
                                # Categories = labels column (skip header)
                                cat_letter = get_column_letter(cat_col)
                                val_letter = get_column_letter(val_col)
                                chart_config['categories'] = f"{cat_letter}2:{cat_letter}{max_row}"
                                chart_config['values'] = f"{val_letter}1:{val_letter}{max_row}"
                                logger.info(f"[FC] Pie chart: categories={chart_config['categories']}, values={chart_config['values']}")
                            else:
                                # Bar/line/etc: use full data range
                                end_col = get_column_letter(max_col)
                                chart_config['data_range'] = f"A1:{end_col}{max_row}"
                                logger.info(f"[FC] Auto-inferred data_range: {chart_config['data_range']}")
                        except Exception as dr_err:
                            logger.warning(f"Could not auto-infer chart data: {dr_err}")
                        finally:
                            if _wb is not None:
                                _wb.close()

                    # Always resolve position (handles keywords + None + absolute refs)
                    raw_pos = chart_config.get('position')
                    resolved_pos = self._resolve_chart_position(
                        target_file, sheet or "", None, raw_pos or 'beside_content'
                    )
                    chart_config['position'] = resolved_pos
                    logger.info(f"[FC] add_chart resolved position: {resolved_pos}")
                    self.excel_tool.add_chart(target_file, sheet, chart_config)
                
                elif operation == 'update_chart':
                    chart_index = parameters.get('chart_index', 0)
                    chart_type = parameters.get('chart_type', '')
                    title = parameters.get('title', None)
                    # Resolve sheet: use provided value, or fall back to first sheet
                    sheet = parameters.get('sheet')
                    if not sheet:
                        wb_tmp = __import__('openpyxl').load_workbook(target_file, read_only=True)
                        sheet = wb_tmp.sheetnames[0]
                        wb_tmp.close()
                    self.excel_tool.update_chart(target_file, sheet, chart_index, chart_type, title)

                elif operation == 'move_chart':
                    chart_index = parameters.get('chart_index', 0)
                    position = parameters.get('position', '')
                    sheet = parameters.get('sheet')
                    if not sheet:
                        wb_tmp = __import__('openpyxl').load_workbook(target_file, read_only=True)
                        sheet = wb_tmp.sheetnames[0]
                        wb_tmp.close()
                    position = self._resolve_chart_position(
                        target_file, sheet, chart_index, position
                    )
                    logger.info(f"[FC] move_chart resolved position: {position}")
                    self.excel_tool.move_chart(target_file, sheet, chart_index, position)

                elif operation == 'resize_chart':
                    chart_index = parameters.get('chart_index', 0)
                    scale = parameters.get('scale')
                    width = parameters.get('width')
                    height = parameters.get('height')
                    # Convert to float if present
                    if scale is not None:
                        scale = float(scale)
                    if width is not None:
                        width = float(width)
                    if height is not None:
                        height = float(height)
                    sheet = parameters.get('sheet')
                    if not sheet:
                        wb_tmp = __import__('openpyxl').load_workbook(target_file, read_only=True)
                        sheet = wb_tmp.sheetnames[0]
                        wb_tmp.close()
                    self.excel_tool.resize_chart(
                        target_file, sheet, chart_index,
                        scale=scale, width=width, height=height
                    )

                elif operation == 'list_charts':
                    sheet = parameters.get('sheet')  # Optional, can be None
                    result = self.excel_tool.list_charts(target_file, sheet)
                    # Store result for user to see
                    return result
                
                elif operation == 'delete_chart':
                    sheet = parameters.get('sheet')
                    identifier = parameters.get('identifier')
                    self.excel_tool.delete_chart(target_file, sheet, identifier)
                
                elif operation == 'sort':
                    sheet = parameters.get('sheet')
                    sort_config = parameters.get('sort_config') or {
                        'column': parameters.get('column'),
                        'order': parameters.get('order', 'asc'),
                    }
                    self.excel_tool.sort_data(target_file, sheet, sort_config)
                
                elif operation == 'remove_duplicates':
                    sheet = parameters.get('sheet')
                    config = parameters.get('config') or {}
                    if isinstance(config, str):
                        try:
                            config = json.loads(config)
                        except Exception:
                            config = {}
                    if not config:
                        config = {
                            k: parameters[k] for k in ('column', 'has_header', 'keep')
                            if parameters.get(k) is not None
                        }
                    result = self.excel_tool.remove_duplicates(target_file, sheet, config)
                    logger.info(f"Removed {result['removed_count']} duplicate(s), {result['remaining_count']} row(s) remaining")
                
                elif operation == 'filter_and_copy':
                    source_sheet = parameters.get('sheet')
                    config = parameters.get('config') or {}
                    if isinstance(config, str):
                        try:
                            config = json.loads(config)
                        except Exception:
                            config = {}
                    if not config:
                        config = {
                            k: parameters[k] for k in (
                                'column', 'operator', 'value',
                                'destination_sheet', 'destination_file',
                                'has_header'
                            )
                            if parameters.get(k) is not None
                        }
                    result = self.excel_tool.filter_and_copy(target_file, source_sheet, config)
                    logger.info(f"Filtered {result['filtered_count']} row(s) to {result['destination_type']}: {result['destination']}")
                
                elif operation == 'insert_rows':
                    sheet = parameters.get('sheet')
                    start_row = parameters.get('start_row')
                    count = parameters.get('count', 1)
                    self.excel_tool.insert_rows(target_file, sheet, start_row, count)
                
                elif operation == 'insert_columns':
                    sheet = parameters.get('sheet')
                    start_col = parameters.get('start_col') or parameters.get('column')
                    count = parameters.get('count', 1)
                    self.excel_tool.insert_columns(target_file, sheet, start_col, count)
                
                elif operation == 'freeze_panes':
                    sheet = parameters.get('sheet')
                    row = parameters.get('row')
                    col = parameters.get('col') or parameters.get('column')
                    self.excel_tool.freeze_panes(target_file, sheet, row, col)
                
                elif operation == 'unfreeze_panes':
                    sheet = parameters.get('sheet')
                    self.excel_tool.unfreeze_panes(target_file, sheet)
                
                elif operation == 'find_and_replace':
                    find_text = parameters.get('find_text') or parameters.get('old_text')
                    replace_text = parameters.get('replace_text') or parameters.get('new_text')
                    sheet = parameters.get('sheet')
                    match_case = parameters.get('match_case', False)
                    match_entire_cell = parameters.get('match_entire_cell', False)
                    result = self.excel_tool.find_and_replace(
                        target_file, find_text, replace_text,
                        sheet=sheet, match_case=match_case, match_entire_cell=match_entire_cell
                    )
                    
                else:
                    raise ValueError(f"Operação Excel não reconhecida: {operation}")
            
            # Word operations
            elif tool == 'word':
                if operation == 'read':
                    result = self.word_tool.read_word(target_file)
                    self._last_read_result = result
                    
                elif operation == 'create':
                    elements = parameters.get('elements')
                    if isinstance(elements, str):
                        try:
                            elements = json.loads(elements)
                        except Exception:
                            elements = None
                    if elements:
                        self.word_tool.create_structured(target_file, elements)
                    else:
                        content = parameters.get('content', '')
                        if content:
                            if contains_html(content):
                                logger.warning("HTML detectado no conteúdo Word — convertendo para elementos.")
                                self.word_tool.create_structured(target_file, html_to_elements(content))
                                return
                            content_lower = content.lower().replace(' ', '').replace('\n', '')
                            suspicious_patterns = [
                                'typetextstyle', 'typetext', 'typecolumns',
                                'typeparagraph', 'typeheading', 'typetable',
                                'typelist', 'typebold', 'typeitalic',
                                'alignmentcenter', 'fontsizebold',
                            ]
                            if any(p in content_lower for p in suspicious_patterns):
                                raise ValueError(
                                    "Conteúdo malformado para criação Word. Tente novamente."
                                )
                        self.word_tool.create_word(target_file, content)
                    
                elif operation == 'update':
                    index = parameters.get('index')
                    new_text = parameters.get('new_text') or parameters.get('content', '')
                    self.word_tool.update_paragraph(target_file, index, new_text)
                    
                elif operation == 'add':
                    text = parameters.get('text', '')
                    self.word_tool.add_paragraph(target_file, text)
                
                elif operation == 'add_paragraph':
                    text = parameters.get('text', '')
                    self.word_tool.add_paragraph(target_file, text)
                
                elif operation == 'add_heading':
                    text = parameters.get('text', '')
                    level = parameters.get('level', 1)
                    self.word_tool.add_heading(target_file, text, level)
                
                elif operation == 'add_table':
                    headers = parameters.get('headers', [])
                    if isinstance(headers, str):
                        try:
                            headers = json.loads(headers)
                        except Exception:
                            headers = []
                    rows = parameters.get('rows', [])
                    if isinstance(rows, str):
                        try:
                            rows = json.loads(rows)
                        except Exception:
                            rows = []
                    
                    # Fallback: try to extract from 'data' or 'content' parameters
                    if not headers:
                        data = parameters.get('data', '')
                        if isinstance(data, str) and data:
                            try:
                                parsed_data = json.loads(data)
                                if isinstance(parsed_data, list) and len(parsed_data) > 0:
                                    if isinstance(parsed_data[0], dict):
                                        # List of dicts: keys = headers, values = rows
                                        headers = list(parsed_data[0].keys())
                                        rows = [list(item.values()) for item in parsed_data]
                                    elif isinstance(parsed_data[0], list):
                                        # List of lists: first = headers, rest = rows
                                        headers = [str(h) for h in parsed_data[0]]
                                        rows = parsed_data[1:]
                            except Exception:
                                pass
                    
                    # Last resort: if still no headers, log error and skip
                    if not headers:
                        logger.error(f"add_table chamado sem headers válidos. parameters={parameters}")
                        raise ValueError(
                            "Parâmetros 'headers' e 'rows' são obrigatórios para add_table. "
                            "Envie headers como JSON array (ex: '[\"Col1\", \"Col2\"]') e "
                            "rows como JSON array de arrays (ex: '[[\"val1\", \"val2\"]]')"
                        )
                    
                    style = parameters.get('style', 'Table Grid')
                    self.word_tool.add_table(target_file, headers, rows, style)
                
                elif operation == 'add_list':
                    items = parameters.get('items', [])
                    if isinstance(items, str):
                        try:
                            items = json.loads(items)
                        except Exception:
                            items = [s.strip() for s in items.split('\n') if s.strip()]
                    if not items:
                        raw = parameters.get('content') or parameters.get('text', '')
                        try:
                            items = json.loads(raw)
                        except Exception:
                            items = [s.strip() for s in raw.split('\n') if s.strip()]
                    ordered = parameters.get('ordered', False)
                    self.word_tool.add_list(target_file, items, ordered)
                
                elif operation == 'delete_paragraph':
                    index = parameters.get('index')
                    self.word_tool.delete_paragraph(target_file, index)
                
                elif operation == 'replace':
                    old_text = parameters.get('old_text', '')
                    new_text = parameters.get('new_text', '')
                    self.word_tool.replace_text(target_file, old_text, new_text)
                
                elif operation == 'format':
                    index = parameters.get('index')
                    formatting = parameters.get('formatting') or {
                        k: parameters[k] for k in ('style', 'bold', 'italic', 'font_size', 'alignment')
                        if parameters.get(k) is not None
                    }
                    self.word_tool.format_paragraph(target_file, index, formatting)
                
                elif operation == 'format_paragraph':
                    index = parameters.get('index')
                    formatting = parameters.get('formatting') or {
                        k: parameters[k] for k in ('style', 'bold', 'italic', 'font_size', 'alignment')
                        if parameters.get(k) is not None
                    }
                    self.word_tool.format_paragraph(target_file, index, formatting)
                
                elif operation == 'update_paragraph':
                    index = parameters.get('index')
                    new_text = parameters.get('new_text') or parameters.get('content', '')
                    self.word_tool.update_paragraph(target_file, index, new_text)
                
                elif operation == 'improve_text':
                    improvement_type = parameters.get('improvement_type', 'clarity')
                    target = parameters.get('target', 'document')
                    index = parameters.get('index')
                    tone = parameters.get('tone')
                    self.word_tool.improve_text(target_file, improvement_type, target, index, tone)
                
                elif operation == 'correct_grammar':
                    target = parameters.get('target', 'document')
                    index = parameters.get('index')
                    self.word_tool.correct_grammar(target_file, target, index)
                
                elif operation == 'improve_clarity':
                    target = parameters.get('target', 'document')
                    index = parameters.get('index')
                    self.word_tool.improve_clarity(target_file, target, index)
                
                elif operation == 'adjust_tone':
                    tone = parameters.get('tone', 'formal')
                    target = parameters.get('target', 'document')
                    index = parameters.get('index')
                    self.word_tool.adjust_tone(target_file, tone, target, index)
                
                elif operation == 'simplify_language':
                    target = parameters.get('target', 'document')
                    index = parameters.get('index')
                    self.word_tool.simplify_language(target_file, target, index)
                
                elif operation == 'rewrite_professional':
                    target = parameters.get('target', 'document')
                    index = parameters.get('index')
                    self.word_tool.rewrite_professional(target_file, target, index)
                
                elif operation == 'generate_summary':
                    output_mode = parameters.get('output_mode', 'new_section')
                    section_title = parameters.get('section_title')
                    self.word_tool.generate_summary(target_file, output_mode, section_title)
                
                elif operation == 'extract_key_points':
                    num_points = parameters.get('num_points', 5)
                    output_mode = parameters.get('output_mode', 'new_section')
                    section_title = parameters.get('section_title')
                    self.word_tool.extract_key_points(target_file, num_points, output_mode, section_title)
                
                elif operation == 'create_resume':
                    size = parameters.get('size', '1_paragraph')
                    output_mode = parameters.get('output_mode', 'new_section')
                    section_title = parameters.get('section_title')
                    self.word_tool.create_resume(target_file, size, output_mode, section_title)
                
                elif operation == 'generate_conclusions':
                    num_conclusions = parameters.get('num_conclusions', 3)
                    output_mode = parameters.get('output_mode', 'new_section')
                    section_title = parameters.get('section_title')
                    self.word_tool.generate_conclusions(target_file, num_conclusions, output_mode, section_title)
                
                elif operation == 'create_faq':
                    num_questions = parameters.get('num_questions', 5)
                    output_mode = parameters.get('output_mode', 'new_section')
                    section_title = parameters.get('section_title')
                    self.word_tool.create_faq(target_file, num_questions, output_mode, section_title)
                
                # Phase 3: Format Conversion and Transformation
                elif operation == 'convert_list_to_table':
                    list_index = parameters.get('list_index', 0)
                    include_header = parameters.get('include_header', False)
                    header_text = parameters.get('header_text')
                    self.word_tool.convert_list_to_table(target_file, list_index, include_header, header_text)
                
                elif operation == 'convert_table_to_list':
                    table_index = parameters.get('table_index', 0)
                    list_type = parameters.get('list_type', 'numbered')
                    skip_header = parameters.get('skip_header', False)
                    separator = parameters.get('separator', ' - ')
                    self.word_tool.convert_table_to_list(target_file, table_index, list_type, skip_header, separator)
                
                elif operation == 'extract_tables_to_excel':
                    output_path = parameters.get('output_path')
                    sheet_names = parameters.get('sheet_names')
                    if not output_path:
                        raise ValueError("output_path é obrigatório para extract_tables_to_excel")
                    self.word_tool.extract_tables_to_excel(target_file, output_path, sheet_names)
                
                elif operation == 'export_to_txt':
                    output_path = parameters.get('output_path')
                    if not output_path:
                        raise ValueError("output_path é obrigatório para export_to_txt")
                    self.word_tool.export_to_txt(target_file, output_path)
                
                elif operation == 'export_to_markdown':
                    output_path = parameters.get('output_path')
                    if not output_path:
                        raise ValueError("output_path é obrigatório para export_to_markdown")
                    self.word_tool.export_to_markdown(target_file, output_path)
                
                elif operation == 'export_to_html':
                    output_path = parameters.get('output_path')
                    if not output_path:
                        raise ValueError("output_path é obrigatório para export_to_html")
                    self.word_tool.export_to_html(target_file, output_path)
                
                elif operation == 'export_to_pdf':
                    output_path = parameters.get('output_path')
                    if not output_path:
                        # Infer output_path from target_file: replace .docx with .pdf
                        from pathlib import Path as _Path
                        output_path = str(_Path(target_file).with_suffix('.pdf'))
                        logger.info(f"output_path não fornecido, inferido automaticamente: {output_path}")
                    self.word_tool.export_to_pdf(target_file, output_path)
                
                # Phase 4: Document Analysis and Insights
                elif operation == 'analyze_word_count':
                    result = self.word_tool.analyze_word_count(target_file)
                    self._last_read_result = result
                
                elif operation == 'analyze_section_length':
                    max_words = parameters.get('max_words', 500)
                    result = self.word_tool.analyze_section_length(target_file, max_words)
                    self._last_read_result = result
                
                elif operation == 'get_document_statistics':
                    result = self.word_tool.get_document_statistics(target_file)
                    self._last_read_result = result
                
                elif operation == 'analyze_tone':
                    result = self.word_tool.analyze_tone(target_file)
                    self._last_read_result = result
                
                elif operation == 'identify_jargon':
                    result = self.word_tool.identify_jargon(target_file)
                    self._last_read_result = result
                
                elif operation == 'analyze_readability':
                    result = self.word_tool.analyze_readability(target_file)
                    self._last_read_result = result
                
                elif operation == 'check_term_consistency':
                    result = self.word_tool.check_term_consistency(target_file)
                    self._last_read_result = result
                
                elif operation == 'analyze_document':
                    include_ai = parameters.get('include_ai_analysis', True)
                    result = self.word_tool.analyze_document(target_file, include_ai)
                    self._last_read_result = result
                
                # Image operations
                elif operation == 'add_image':
                    image_path = parameters.get('image_path')
                    if not image_path:
                        raise ValueError("image_path é obrigatório para add_image")
                    width = parameters.get('width')
                    height = parameters.get('height')
                    alignment = parameters.get('alignment', 'left')
                    caption = parameters.get('caption')
                    self.word_tool.add_image(target_file, image_path, width, height, alignment, caption)
                
                elif operation == 'add_image_at_position':
                    image_path = parameters.get('image_path')
                    if not image_path:
                        raise ValueError("image_path é obrigatório para add_image_at_position")
                    paragraph_index = parameters.get('paragraph_index', 0)
                    width = parameters.get('width')
                    height = parameters.get('height')
                    self.word_tool.add_image_at_position(target_file, image_path, paragraph_index, width, height)

                # Hyperlink operations
                elif operation == 'add_hyperlink':
                    text = parameters.get('text')
                    url = parameters.get('url')
                    if not text:
                        raise ValueError("text é obrigatório para add_hyperlink")
                    if not url:
                        raise ValueError("url é obrigatório para add_hyperlink")
                    bold = parameters.get('bold', False)
                    italic = parameters.get('italic', False)
                    font_size = parameters.get('font_size')
                    color = parameters.get('color', '0563C1')
                    self.word_tool.add_hyperlink(target_file, text, url, bold, italic, font_size, color)

                elif operation == 'add_hyperlink_to_paragraph':
                    text = parameters.get('text')
                    url = parameters.get('url')
                    if not text:
                        raise ValueError("text é obrigatório para add_hyperlink_to_paragraph")
                    if not url:
                        raise ValueError("url é obrigatório para add_hyperlink_to_paragraph")
                    paragraph_index = parameters.get('paragraph_index', 0)
                    bold = parameters.get('bold', False)
                    italic = parameters.get('italic', False)
                    color = parameters.get('color', '0563C1')
                    self.word_tool.add_hyperlink_to_paragraph(target_file, paragraph_index, text, url, bold, italic, color)

                # Header and footer operations
                elif operation == 'add_header':
                    text = parameters.get('text', '')
                    alignment = parameters.get('alignment', 'center')
                    bold = parameters.get('bold', False)
                    italic = parameters.get('italic', False)
                    font_size = parameters.get('font_size')
                    font_name = parameters.get('font_name')
                    include_page_number = parameters.get('include_page_number', False)
                    page_number_position = parameters.get('page_number_position', 'right')
                    include_total_pages = parameters.get('include_total_pages', False)
                    use_tab_stops = parameters.get('use_tab_stops', False)
                    self.word_tool.add_header(target_file, text, alignment, bold, italic, font_size, font_name, include_page_number, page_number_position, include_total_pages, use_tab_stops)

                elif operation == 'add_footer':
                    text = parameters.get('text', '')
                    alignment = parameters.get('alignment', 'center')
                    bold = parameters.get('bold', False)
                    italic = parameters.get('italic', False)
                    font_size = parameters.get('font_size')
                    font_name = parameters.get('font_name')
                    include_page_number = parameters.get('include_page_number', True)
                    page_number_position = parameters.get('page_number_position', 'center')
                    include_total_pages = parameters.get('include_total_pages', False)
                    self.word_tool.add_footer(target_file, text, alignment, bold, italic, font_size, font_name, include_page_number, page_number_position, include_total_pages)

                elif operation == 'remove_header':
                    self.word_tool.remove_header(target_file)

                elif operation == 'remove_footer':
                    self.word_tool.remove_footer(target_file)

                # Page layout operations
                elif operation == 'set_page_margins':
                    top = parameters.get('top', 2.5)
                    bottom = parameters.get('bottom', 2.5)
                    left = parameters.get('left', 3.0)
                    right = parameters.get('right', 2.0)
                    unit = parameters.get('unit', 'cm')
                    self.word_tool.set_page_margins(target_file, top, bottom, left, right, unit)

                elif operation == 'set_page_size':
                    size = parameters.get('size', 'A4')
                    orientation = parameters.get('orientation', 'portrait')
                    self.word_tool.set_page_size(target_file, size, orientation)

                elif operation == 'get_page_layout':
                    result = self.word_tool.get_page_layout(target_file)
                    self._last_read_result = result

                # Page break and section break operations
                elif operation == 'add_page_break':
                    position = parameters.get('position')
                    self.word_tool.add_page_break(target_file, position)

                elif operation == 'add_section_break':
                    break_type = parameters.get('break_type', 'new_page')
                    position = parameters.get('position')
                    self.word_tool.add_section_break(target_file, break_type, position)
                    
                else:
                    raise ValueError(f"Operação Word não reconhecida: {operation}")
            
            # PowerPoint operations
            elif tool == 'powerpoint':
                if operation == 'read':
                    result = self.powerpoint_tool.read_powerpoint(target_file)
                    self._last_read_result = result
                    
                elif operation == 'create':
                    slides = parameters.get('slides', [])
                    if isinstance(slides, str):
                        try:
                            slides = json.loads(slides)
                        except Exception:
                            slides = []
                    self.powerpoint_tool.create_powerpoint(target_file, slides)
                    
                elif operation == 'update':
                    slide_index = parameters.get('slide_index')
                    new_content = parameters.get('new_content') or {}
                    if isinstance(new_content, str):
                        try:
                            new_content = json.loads(new_content)
                        except Exception:
                            new_content = {}
                    if not new_content:
                        new_content = {k: parameters[k] for k in ('title', 'content', 'text') if parameters.get(k)}
                    self.powerpoint_tool.update_slide(target_file, slide_index, new_content)
                    
                elif operation == 'add':
                    slide_data = parameters.get('slide_data') or {}
                    if isinstance(slide_data, str):
                        try:
                            slide_data = json.loads(slide_data)
                        except Exception:
                            slide_data = {}
                    if not slide_data:
                        slide_data = {k: parameters[k] for k in ('title', 'content', 'text', 'layout') if parameters.get(k)}
                    self.powerpoint_tool.add_slide(target_file, slide_data)
                
                elif operation == 'delete_slide':
                    slide_index = parameters.get('slide_index')
                    self.powerpoint_tool.delete_slide(target_file, slide_index)
                
                elif operation == 'duplicate_slide':
                    slide_index = parameters.get('slide_index')
                    self.powerpoint_tool.duplicate_slide(target_file, slide_index)
                
                elif operation == 'add_textbox':
                    slide_index = parameters.get('slide_index')
                    text = parameters.get('text', '')
                    self.powerpoint_tool.add_textbox(
                        target_file, slide_index, text,
                        left=parameters.get('left', 1.0),
                        top=parameters.get('top', 1.0),
                        width=parameters.get('width', 5.0),
                        height=parameters.get('height', 1.0),
                        font_size=parameters.get('font_size', 18),
                        bold=parameters.get('bold', False),
                        font_color=parameters.get('font_color')
                    )
                
                elif operation == 'add_table':
                    slide_index = parameters.get('slide_index')
                    headers = parameters.get('headers', [])
                    if isinstance(headers, str):
                        try:
                            headers = json.loads(headers)
                        except Exception:
                            headers = []
                    rows = parameters.get('rows', [])
                    if isinstance(rows, str):
                        try:
                            rows = json.loads(rows)
                        except Exception:
                            rows = []
                    self.powerpoint_tool.add_table_to_slide(
                        target_file, slide_index, headers, rows
                    )
                
                elif operation == 'replace':
                    old_text = parameters.get('old_text', '')
                    new_text = parameters.get('new_text', '')
                    self.powerpoint_tool.replace_text(target_file, old_text, new_text)
                    
                else:
                    raise ValueError(f"Operação PowerPoint não reconhecida: {operation}")
            
            # PDF operations
            elif tool == 'pdf':
                if operation == 'read':
                    result = self.pdf_tool.read_pdf(target_file)
                    self._last_read_result = result
                
                elif operation == 'create':
                    elements = parameters.get('elements', [])
                    if isinstance(elements, str):
                        try:
                            elements = json.loads(elements)
                        except Exception:
                            elements = []
                    page_size = parameters.get('page_size', 'A4')
                    self.pdf_tool.create_pdf(target_file, elements, page_size)
                
                elif operation == 'merge':
                    file_paths = parameters.get('file_paths', [])
                    if isinstance(file_paths, str):
                        try:
                            file_paths = json.loads(file_paths)
                        except Exception:
                            file_paths = [file_paths] if file_paths else []
                    resolved = []
                    for fp in file_paths:
                        validated = self.security_validator.validate_file_path(fp)
                        resolved.append(str(validated))
                    self.pdf_tool.merge_pdfs(resolved, target_file)
                
                elif operation == 'split':
                    output_path = parameters.get('output_path', target_file)
                    validated_output = str(self.security_validator.validate_file_path(output_path))
                    start_page = parameters.get('start_page')
                    end_page = parameters.get('end_page')
                    if start_page is None or end_page is None:
                        raise ValueError(
                            "Parâmetros 'start_page' e 'end_page' são obrigatórios para split de PDF."
                        )
                    self.pdf_tool.split_pdf(target_file, validated_output, int(start_page), int(end_page))
                
                elif operation == 'add_text':
                    text = parameters.get('text', '')
                    self.pdf_tool.add_text_overlay(
                        target_file, text,
                        x=parameters.get('x', 200),
                        y=parameters.get('y', 400),
                        font_size=parameters.get('font_size', 40),
                        color=parameters.get('color', 'CCCCCC'),
                        opacity=parameters.get('opacity', 0.3),
                        pages=parameters.get('pages')
                    )
                
                elif operation == 'get_info':
                    result = self.pdf_tool.get_info(target_file)
                    self._last_read_result = result
                
                elif operation == 'rotate':
                    rotation = parameters.get('rotation', 90)
                    pages = parameters.get('pages')
                    self.pdf_tool.rotate_pages(target_file, rotation, pages)
                
                elif operation == 'extract_tables':
                    result = self.pdf_tool.extract_tables(target_file)
                    self._last_read_result = result
                
                else:
                    raise ValueError(f"Operação PDF não reconhecida: {operation}")
            
            else:
                raise ValueError(f"Ferramenta não reconhecida: {tool}")
            
            # Salvar estado pós-operação para redo
            if self.version_manager and version_id:
                self.version_manager.save_after_state(target_file, version_id)
            
            logger.debug(f"Ação executada com sucesso: {tool}.{operation}")
            
        except Exception as e:
            logger.error(f"Erro ao executar ação: {e}")
            raise

    
    def undo_file(self, file_path: str) -> AgentResponse:
        """Desfaz a última operação em um arquivo.
        
        Args:
            file_path: Caminho do arquivo
        
        Returns:
            AgentResponse com resultado da operação
        """
        if not self.version_manager:
            return AgentResponse(
                success=False,
                message="Versionamento não está habilitado",
                files_modified=[],
                error="VersionManager não inicializado"
            )
        
        try:
            self.version_manager.undo(file_path)
            return AgentResponse(
                success=True,
                message=f"Operação desfeita com sucesso: {Path(file_path).name}",
                files_modified=[file_path],
                error=None
            )
        except Exception as e:
            logger.error(f"Erro ao desfazer operação: {e}")
            return AgentResponse(
                success=False,
                message=f"Erro ao desfazer operação: {e}",
                files_modified=[],
                error=str(e)
            )
    
    def redo_file(self, file_path: str) -> AgentResponse:
        """Refaz uma operação desfeita em um arquivo.
        
        Args:
            file_path: Caminho do arquivo
        
        Returns:
            AgentResponse com resultado da operação
        """
        if not self.version_manager:
            return AgentResponse(
                success=False,
                message="Versionamento não está habilitado",
                files_modified=[],
                error="VersionManager não inicializado"
            )
        
        try:
            self.version_manager.redo(file_path)
            return AgentResponse(
                success=True,
                message=f"Operação refeita com sucesso: {Path(file_path).name}",
                files_modified=[file_path],
                error=None
            )
        except Exception as e:
            logger.error(f"Erro ao refazer operação: {e}")
            return AgentResponse(
                success=False,
                message=f"Erro ao refazer operação: {e}",
                files_modified=[],
                error=str(e)
            )
