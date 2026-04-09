import openpyxl

wb = openpyxl.load_workbook('C:/Users/ngb/Documents/GeminiOfficeFiles/vendas_simples.xlsx')
ws = wb.active

print(f'Total rows: {ws.max_row}')
print(f'Total columns: {ws.max_column}')
print(f'Header: {[cell.value for cell in ws[1]]}')
print(f'First data row: {[cell.value for cell in ws[2]]}')
print(f'Last data row: {[cell.value for cell in ws[ws.max_row]]}')
print(f'\nHeader formatting:')
print(f'  A1 bold: {ws["A1"].font.bold}')
print(f'  A1 bg_color: {ws["A1"].fill.start_color.rgb if ws["A1"].fill.start_color else None}')
print(f'  A1 fill_type: {ws["A1"].fill.fill_type}')

wb.close()
