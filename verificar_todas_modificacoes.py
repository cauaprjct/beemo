from src.excel_tool import ExcelTool
from openpyxl import load_workbook

tool = ExcelTool()
data = tool.read_excel('C:/Users/ngb/Documents/GeminiOfficeFiles/vendas_2025.xlsx')

sheet_name = list(data['sheets'].keys())[0]
rows = data['sheets'][sheet_name]

print('=' * 80)
print('VERIFICAÇÃO DE TODAS AS MODIFICAÇÕES - vendas_2025.xlsx')
print('=' * 80)

# TESTE 2: Linha 5 - Quantidade alterada para 100
print('\n🔍 TESTE 2: Modificação na Linha 5 (Quantidade)')
print('-' * 80)
linha_5 = rows[4]  # Índice 4 = linha 5 (linha 1 é cabeçalho)
print(f'   Linha 5 completa: {linha_5}')
print(f'   ID: {linha_5[0]}')
print(f'   Data: {linha_5[1]}')
print(f'   Produto: {linha_5[2]}')
print(f'   Quantidade (D5): {linha_5[3]}')
print(f'   Preço Unitário (E5): {linha_5[4]}')
print(f'   Total (F5): {linha_5[5]}')

if linha_5[3] == 100:
    print('\n   ✅ SUCESSO! Quantidade alterada para 100')
else:
    print(f'\n   ❌ FALHA! Quantidade ainda é {linha_5[3]} (esperado: 100)')

# Verificar se a fórmula está presente
wb = load_workbook('C:/Users/ngb/Documents/GeminiOfficeFiles/vendas_2025.xlsx')
ws = wb[sheet_name]
formula_f5 = ws['F5'].value
print(f'   Fórmula em F5: {formula_f5}')
if formula_f5 and '=' in str(formula_f5):
    print('   ✅ Fórmula preservada!')
wb.close()

# TESTE 3: Linha 10 - Quantidade = 25 e Preço = 150
print('\n🔍 TESTE 3: Modificações na Linha 10 (Quantidade e Preço)')
print('-' * 80)
linha_10 = rows[9]  # Índice 9 = linha 10
print(f'   Linha 10 completa: {linha_10}')
print(f'   ID: {linha_10[0]}')
print(f'   Data: {linha_10[1]}')
print(f'   Produto: {linha_10[2]}')
print(f'   Quantidade (D10): {linha_10[3]}')
print(f'   Preço Unitário (E10): {linha_10[4]}')
print(f'   Total (F10): {linha_10[5]}')

sucesso_quantidade = linha_10[3] == 25
sucesso_preco = linha_10[4] == 150

if sucesso_quantidade:
    print('\n   ✅ SUCESSO! Quantidade alterada para 25')
else:
    print(f'\n   ❌ FALHA! Quantidade ainda é {linha_10[3]} (esperado: 25)')

if sucesso_preco:
    print('   ✅ SUCESSO! Preço alterado para 150')
else:
    print(f'   ❌ FALHA! Preço ainda é {linha_10[4]} (esperado: 150)')

# TESTE 4: Linha com ID=50 - Produto alterado para "Produto VIP"
print('\n🔍 TESTE 4: Busca e Modificação (ID=50, Produto="Produto VIP")')
print('-' * 80)

# Encontrar linha com ID=50
linha_id50 = None
indice_id50 = None
for i, row in enumerate(rows[1:], start=2):  # Pula cabeçalho
    if row[0] == 50:
        linha_id50 = row
        indice_id50 = i
        break

if linha_id50:
    print(f'   Linha {indice_id50} (ID=50) encontrada!')
    print(f'   Linha completa: {linha_id50}')
    print(f'   ID: {linha_id50[0]}')
    print(f'   Data: {linha_id50[1]}')
    print(f'   Produto: {linha_id50[2]}')
    print(f'   Quantidade: {linha_id50[3]}')
    print(f'   Preço Unitário: {linha_id50[4]}')
    
    if linha_id50[2] == 'Produto VIP':
        print('\n   ✅ SUCESSO! Produto alterado para "Produto VIP"')
    else:
        print(f'\n   ❌ FALHA! Produto ainda é "{linha_id50[2]}" (esperado: "Produto VIP")')
else:
    print('   ❌ ERRO! Linha com ID=50 não encontrada')

# RESUMO FINAL
print('\n' + '=' * 80)
print('📊 RESUMO FINAL')
print('=' * 80)

testes = {
    'Teste 2 (Linha 5 - Quantidade=100)': linha_5[3] == 100,
    'Teste 3 (Linha 10 - Quantidade=25)': linha_10[3] == 25,
    'Teste 3 (Linha 10 - Preço=150)': linha_10[4] == 150,
    'Teste 4 (ID=50 - Produto="Produto VIP")': linha_id50 and linha_id50[2] == 'Produto VIP'
}

for teste, resultado in testes.items():
    simbolo = '✅' if resultado else '❌'
    status = 'PASSOU' if resultado else 'FALHOU'
    print(f'{simbolo} {teste}: {status}')

total_sucesso = sum(testes.values())
total_testes = len(testes)
print(f'\n🎯 SCORE: {total_sucesso}/{total_testes} testes passaram ({int(total_sucesso/total_testes*100)}%)')

if total_sucesso == total_testes:
    print('\n🎉 PERFEITO! Todas as modificações foram aplicadas corretamente!')
elif total_sucesso > 0:
    print(f'\n⚠️ PARCIAL! {total_sucesso} de {total_testes} modificações aplicadas.')
else:
    print('\n❌ FALHA TOTAL! Nenhuma modificação foi aplicada.')
